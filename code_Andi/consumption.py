import psycopg2 as pg
import psycopg2.extras as pge
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.cm as cm
import pandas as pd
from matplotlib.backends.backend_pdf import PdfPages
import os
from PyPDF2 import PdfFileMerger, PdfFileReader
from sklearn import preprocessing
from sklearn import linear_model
from sklearn import model_selection
from sklearn import ensemble
from sklearn import neighbors
from sklearn import decomposition
from sklearn.cluster import KMeans, DBSCAN
from sklearn.metrics import silhouette_samples, silhouette_score, calinski_harabaz_score, mean_squared_error
from sklearn.model_selection import StratifiedKFold, train_test_split
from sklearn.metrics import cohen_kappa_score, precision_recall_fscore_support, accuracy_score, classification_report
# from sklearn.pipeline import make_pipeline
import xlrd
import copy
import time
import sompy as smp
from collections import Counter
import pickle
import pyprind
import sys
from scipy.cluster.hierarchy import cophenet
from scipy.spatial.distance import pdist
from scipy.cluster.hierarchy import inconsistent
from mpl_toolkits.axes_grid.anchored_artists import AnchoredText
import csv
from scipy import stats
from openpyxl import load_workbook
import seaborn as sns
sns.set(color_codes=True)
import brightway2 as bw2
# try:
#     from Clustering_Tools import *
# except:
#     sys.path.insert(0, r"D:\froemelt\Documents\Andi\02_Doktorat\03_Projects\05_HEIA\03_Computations\04_HEIA-Tools")
#     from Clustering_Tools import *
from sompy.visualization.bmuhits import BmuHitsView
# from bokeh.io import output_file, save
# from bokeh.plotting import figure, show, ColumnDataSource
# from bokeh.models import LinearAxis, Range1d, HoverTool

def get_time(printstring, tic):
    """
    This is a helper-function and prints the elapsed time taken from "tic" to present in a nice format
    :param printstring: String which will be displayed e.g.: 'Total Time' (note that ':' is not needed)
    :param tic: start time (needs to be taken as time.time())
    """
    elapsed = time.time() - tic  # elapsed time between tic and present
    m, s = divmod(elapsed, 60)  # get minutes and seconds by division and remainder
    h, m = divmod(m, 60)  # get hours and minutes

    # prepare nice numbers to print
    h = "0" + str(round(h)) if h < 10 else str(round(h))
    m = "0" + str(round(m)) if m < 10 else str(round(m))
    s = "0" + str(round(s)) if s < 10 else str(round(s))

    # print elapsed time and printstring:
    print("{}: {}:{}:{}".format(printstring, h, m, s))


def get_pg_connection():
    """
    This is a helper function which was unfortunately implemented very late since it would have been reasonable for so
    many applications. The function returns a connection to the PostgreSQL-Database.
    :return: psycopg2-connection class
    """

    return pg.connect("host=localhost port=5432 dbname=db_andi user=postgres  connect_timeout=2 password='postgres'")


class Vardata(object):
    """
    The vardata-class is used by create_consumptionfigs-function which is used for data screening of the HABE-data.
    A vardata-class extracts the data of a specified variable from the PG-database. It then translates the HABE-code into
    German, creates a pandas-dataframe and computes same basic statistics.
    """
    def __init__(self, conn, varname, ws, ivar, varrow):
        # conn = connection with the PG-database
        # varname (string) = HABE-code
        # ws = excel-worksheet which contains translation information from HABE-code to German
        # ivar = control variable for varname --> used to navigate within the excel-worksheet
        # varrow = row in the excel-worksheet where the HABE-code can be found.

        # Preparation of data extraction from PG-database:
        cur = conn.cursor()
        query = "SELECT {} FROM original_data.habe_{}".format(varname.lower(), ws.name.lower())
        cur.execute(query)

        # Extract data from PG-database and store in list
        self.sqldata = self.__get_data(cur)

        # Translate HABE-code into German:
        self.dataname = self.__get_name(ivar, varrow, ws)

        # Convert list into pandas-data-frame. The second column is needed to display violinplots and boxplots in the same fig
        self.df = pd.DataFrame({self.dataname: self.sqldata, 'pseudo': np.nan})

        # Compute some basic statistics:
        self.avg = self.df[self.dataname].mean()
        self.med = self.df[self.dataname].median()
        self.mi = self.df[self.dataname].min()
        self.ma = self.df[self.dataname].max()
        self.stdev = self.df[self.dataname].std()

        # Compute coefficient of variation:
        self.cv = self.stdev/self.avg if self.avg>0 else 0

        # Label text for figures:
        self.txt = "mean: {}\nmedian: {}\nmin: {}\nmax: {}\nstd: {}\ncv: {}".format(
            round(self.avg,1), round(self.med,1), round(self.mi,1), round(self.ma,1),
            round(self.stdev,1), round(self.cv,1))

        cur.close()

    def __get_data(self, cur):
        # Create list from PG-SQL-data
        sqldat = []
        for row in cur:
            sqldat.append(float(row[0]))
        return sqldat

    def __get_name(self, ivar, varrow, ws):
        # Look for German name of variable in excel-worksheet:
        for i in range(0, varrow):
            if ws.cell_value(ivar, i) != '':
                return ws.cell_value(ivar, i)


def createsinglefig(datatop, databottom, unit):
    """
    The createsinglefig-function creates one figure for the HABE-data-screening induced by create_consumptionfigs-fct.
    """
    # Set figure size to A4 page
    fig_width_cm = 21  # A4 page
    fig_height_cm = 29.7
    inches_per_cm = 1 / 2.58  # Convert cm to inches
    fig_width = fig_width_cm * inches_per_cm  # width in inches
    fig_height = fig_height_cm * inches_per_cm  # height in inches
    fig_size = [fig_height, fig_width]  # height and width are in the order needed for landscape

    # Create figure and set an overall-figure-title
    fig = plt.figure(figsize=fig_size)
    plt.suptitle("top: {}, bottom: {}".format(datatop.dataname, databottom.dataname), size=14)

    # First subplot (see jupyter notebook for more infos regarding histograms, violinplots and boxplots): histogram
    plt.subplot(2, 3, 1)
    sns.distplot(datatop.sqldata, kde=False, bins='auto', norm_hist=True)
    plt.title('Binning: Auto')
    plt.ylabel('{}\nnormalized frequency (-)'.format(datatop.dataname))
    plt.xlabel(unit)

    # Second subplot: histogram with different binning
    plt.subplot(2, 3, 2)
    sns.distplot(datatop.sqldata, kde=False, bins='doane', norm_hist=True)
    plt.title('Binning: Doane')
    plt.xlabel(unit)

    # Third subplot: boxplot and violinplot. The "order" is once switched in order display both plots side-by-side.
    # in the violinplot is cut=0 in order not to go beyond minimum and maximum values.
    plt.subplot(2, 3, 3)
    ax = sns.boxplot(data=datatop.df, palette='Set3', width=0.4, order=[0, 1], linewidth=0.7)
    ax = sns.violinplot(data=datatop.df, palette='Set3', width=0.4, order=[1, 0], cut=0, inner='quartile', linewidth=0.7)
    # Display basic statistics:
    ax.text(0.3, datatop.ma*0.7, datatop.txt, fontsize=12)
    plt.ylabel(unit)
    plt.ylim(datatop.mi, datatop.ma*1.1)
    # The next line overwrites the 0, 1 ticks
    plt.xticks([0, 1], ['boxplot','violinplot'])

    # The next three subplots are the same as above but for a second variable
    plt.subplot(2, 3, 4)
    sns.distplot(databottom.sqldata, kde=False, bins='auto', norm_hist=True)
    plt.title('Binning: Auto')
    plt.ylabel('{}\nnormalized frequency (-)'.format(databottom.dataname))
    plt.xlabel(unit)

    plt.subplot(2, 3, 5)
    sns.distplot(databottom.sqldata, kde=False, bins='doane', norm_hist=True)
    plt.title('Binning: Doane')
    plt.xlabel(unit)

    plt.subplot(2, 3, 6)
    ax = sns.boxplot(data=databottom.df, palette='Set3', width=0.4, order=[0, 1], linewidth=0.7)
    ax = sns.violinplot(data=databottom.df, palette='Set3', width=0.4, order=[1, 0], cut=0, inner='quartile', linewidth=0.7)
    ax.text(0.3,databottom.ma*0.7,databottom.txt, fontsize=12)
    plt.ylabel(unit)
    plt.ylim(databottom.mi, databottom.ma*1.1)
    plt.xticks([0, 1], ['boxplot','violinplot'])

    return fig


def create_consumptionfigs(wb, sheetname, varrow, unit, level='all'):
    """
    The function create_consumptionfigs is used to create histograms, boxplots, violinplots and some statistics for all
    HABE-data
    :param wb: excel-workbook with information on HABE-data (meta-data-description: translation from HABE-code to readable code)
    :param sheetname: (string) = name of the sheet with translation-table
    :param varrow: (integer) = number of row in the excel-sheet in which the HABE-codes can be found
    :param unit: (string) = label for figures
    :param level: (integer or 'all') = level of HABE-aggregation. If integer is specified, this refers to the number of digits
    in the HABE-code (level=5 means e.g. "A5111")
    """

    # connect to PG-database:
    conn = pg.connect("host=localhost port=5432 dbname=db_andi user=postgres  connect_timeout=2 password='postgres'")

    # read excel-sheet and retrieve HABE-code for variables:
    ws = wb.sheet_by_name(sheetname)
    varnames = ws.col_values(varrow)

    # create multi-page-pdf to store the figures:
    ipdf = 0  # this is a control variable to number different pdfs (putting everything into one pdf uses too much memory)
    resultpath = r"D:\froemelt\Documents\Andi\02_Doktorat\03_Projects\05_HEIA\03_Computations\pyHEIA\Consumption-Data-Screening"
    pp = PdfPages('{}\{}_{}.pdf'.format(resultpath, sheetname, ipdf))

    # if only one aggregation level should be considered, then the following if-loop replaces all variables which do not
    # correspond to the specified aggregation level
    if level!='all':
        varnames = [var if len(var) == level else '' for var in varnames]

    # the following list stores all variables which were already analyzed. This list is necessary, because we want to
    # place two variables on each A4-page in the resulting pdf.
    varschecked = ['', 'Variablenname', 'HaushaltID']

    # "core": we go through all variables of the HABE-data:
    for ivar, varname in enumerate(varnames):  # ivar is needed to have a control variable within the excel-sheet.

        # if the pdf has already 20 or more pages, we then close the pdf and open a new one ("varname not in
        # (varschecked)" was necessary in an older version).
        if pp.get_pagecount() >= 20:  # and varname not in(varschecked):
            pp.close()
            del pp
            ipdf += 1
            pp = PdfPages('{}\{}_{}.pdf'.format(resultpath, sheetname, ipdf))

        # check if variable was already analyzed:
        if varname not in(varschecked):

            # instantiation of a Vardata-class which extracts the data from the PG-database, creates a Pandas-
            # Dataframe, translates the variablename from HABE-code into German and computes some statistics
            datatop = Vardata(conn, varname, ws, ivar, varrow)
            varschecked.extend([varname])

            # since we aim to put always two variables on one A4-page, we are now looking for a next variable:
            ivarbot = ivar
            while ivarbot < len(varnames)-1 and varnames[ivarbot] in(varschecked):
                ivarbot += 1

            # the following if-loop is only necessary if the last element in the varnames-list is empty
            if varnames[ivarbot] == '':
                ivarbot = ivar

            # instantiation of a Vardata-class for the second variable --> important: if the first variable was already
            # the last entry in the variable-names-list, then this variable will be analyzed for second time.
            databottom = Vardata(conn, varnames[ivarbot], ws, ivarbot, varrow)
            varschecked.extend([varnames[ivarbot]])

            # call of the createsinglefig-function which creates the figures for one A4-page:
            fig = createsinglefig(datatop, databottom, unit)
            fig.savefig(pp, format='pdf', orientation='landscape', papertype='a4')

            # close and delete fig to release memory:
            plt.close(fig)
            del fig

        #if varname == "A5115":  # only for debugging
            #break
    pp.close()
    conn.close()

    # Finally, we want to merge all pdfs into one:
    allpdfs = [f for f in os.listdir(resultpath) if f.endswith('.pdf')]
    merger = PdfFileMerger()
    for filename in allpdfs:
        with open("{}\{}".format(resultpath, filename), 'rb') as f:
            merger.append(PdfFileReader(f))
    merger.write("{}\{}_level_{}.pdf".format(resultpath, sheetname, level))

    # Delete all intermediate pdfs:
    for f in allpdfs:
        os.remove("{}\{}".format(resultpath, f))

    # Move file to a specific new folder:
    newfolder = "{}\Level {}".format(resultpath, level)
    os.makedirs(newfolder, exist_ok=True)
    os.rename("{}\{}_level_{}.pdf".format(resultpath, sheetname, level), "{}\Level {}\{}_level_{}.pdf".format(resultpath, level, sheetname, level))


class HABE_HH(object):
    # -------------------------------------------------------------
    # ATTENTION: this class is more or less deprecated and replaced by HABE_HH_V2 see below
    # -------------------------------------------------------------
    # This class is used by habe_hh_predictor_builder(). It takes a list of dicts as an input. Thereby, the dicts
    # represent persons of a certain household with attributes defined by HABE-tables. The goal is to process these
    # attributes in a way to derive new predictors which can be easily matched with STATPOP or which shall be further
    # investigated (see also "D:\froemelt\Documents\Andi\02_Doktorat\03_Projects\05_HEIA\03_Computations\Consumption\Datascreening.xlsx"

    def __init__(self, pers, **kwargs):
        self._hh_wide_attr(**pers[0])
        self._hh_wide_geogr(**pers[0])
        self._hh_age_sex_stats(pers)
        self._hh_marit_stats(pers)
        self._hh_nat_stats(pers)

    def _hh_wide_attr(self, haushaltid, e10, primaereinkommen08, bruttoeinkommen08,
                     verfuegbareseinkommen08, sparbetrag08, anzahlselbstaendiger05,
                     anzahlunselbstaendiger05, anzahlrentner05, anzahlausbildung05,
                     anzahlandere05, mieterhaushalt05, mindestenseinauto05, mindestenseinvelo05,
                      mindestenseincomputer05, mindestenseinnatel05, mindestenseinhaustier05, **kwargs):
        # This function fills in the attributes/predictors which are readily available in the HABE-tables and which are
        # the same for all household members.
        # Attributes which cannot be derived from STATPOP (in this case: all attributes) are annotated with "prov_"
        self.haushaltid = haushaltid
        self.prov_e10 = e10     # Erwerbseinkommen
        self.prov_primaereinkommen08 = primaereinkommen08
        self.prov_bruttoeinkommen08 = bruttoeinkommen08
        self.prov_verfuegbareseinkommen08 = verfuegbareseinkommen08
        self.prov_sparbetrag08 = sparbetrag08
        self.prov_anzahlselbstaendiger05 = anzahlselbstaendiger05
        self.prov_anzahlunselbstaendiger05 = anzahlunselbstaendiger05
        self.prov_anzahlrentner05 = anzahlrentner05
        self.prov_anzahlausbildung05 = anzahlausbildung05
        self.prov_anzahlandere05 = anzahlandere05
        self.prov_mieterhaushalt05 = mieterhaushalt05
        self.prov_mindestenseinauto05 = mindestenseinauto05
        self.prov_mindestenseinvelo05 = mindestenseinvelo05
        self.prov_mindestenseincomputer05 = mindestenseincomputer05
        self.prov_mindestenseinnatel05 = mindestenseinnatel05
        self.prov_mindestenseinhaustier05 = mindestenseinhaustier05

    def _hh_wide_geogr(self, kanton08, sprachregion98, grossregion01, **kwargs):
        # This function creates dummy variables based on the geographic information from HABE. This could be derived from
        # STATPOP (partly in combination with ARE/BFS-Raumgliederungen).
        self.sprachregion98_dch = 1 if sprachregion98 == 1 else 0
        self.sprachregion98_fch = 1 if sprachregion98 == 2 else 0
        self.sprachregion98_ich = 1 if sprachregion98 == 3 else 0

        self.kanton08_zh = 1 if kanton08 == 1 else 0
        self.kanton08_be = 1 if kanton08 == 2 else 0
        self.kanton08_lu = 1 if kanton08 == 3 else 0
        self.kanton08_sg = 1 if kanton08 == 17 else 0
        self.kanton08_ag = 1 if kanton08 == 19 else 0
        self.kanton08_ti = 1 if kanton08 == 21 else 0
        self.kanton08_vd = 1 if kanton08 == 22 else 0
        self.kanton08_ge = 1 if kanton08 == 25 else 0
        self.kanton08_rest = 1 if kanton08 == 99 else 0

        self.grossregion01_ge = 1 if grossregion01 == 1 else 0
        self.grossregion01_mit = 1 if grossregion01 == 2 else 0
        self.grossregion01_nw = 1 if grossregion01 == 3 else 0
        self.grossregion01_zh = 1 if grossregion01 == 4 else 0
        self.grossregion01_ost = 1 if grossregion01 == 5 else 0
        self.grossregion01_zen = 1 if grossregion01 == 6 else 0
        self.grossregion01_ti = 1 if grossregion01 == 7 else 0

    def _hh_age_sex_stats(self, pers):
        # This function creates new predictors: it uses the same age categories as HABE, but distinguishes now also
        # between male and female persons.

        # Create a dict to count the persons of each sex/age-group:
        sex_age_dict = {}
        for sex in ['fem', 'male']:
            for age in ['0004', '0514', '1524', '2534', '3544', '4554', '5564', '6574', '7599']:
                sex_age_dict.update({(sex, age): 0})


        # Go through all HH members and create a tuple-key for the sex_age_dict
        for p in pers:
            sex = 'fem' if p['geschlecht98'] == 2 else 'male'

            if p['lebensalter98'] <= 4:
                age = '0004'
            elif 5 <= p['lebensalter98'] <= 14:
                age = '0514'
            elif 15 <= p['lebensalter98'] <= 24:
                age = '1524'
            elif 25 <= p['lebensalter98'] <= 34:
                age = '2534'
            elif 35 <= p['lebensalter98'] <= 44:
                age = '3544'
            elif 45 <= p['lebensalter98'] <= 54:
                age = '4554'
            elif 55 <= p['lebensalter98'] <= 64:
                age = '5564'
            elif 65 <= p['lebensalter98'] <= 74:
                age = '6574'
            elif 75 <= p['lebensalter98']:
                age = '7599'

            # Increase count of the age/sex-group to which the HH-member under consideration belongs:
            sex_age_dict[(sex, age)] += 1

        # Translate the sex_age_dict into attributes of the class:
        for ky in sex_age_dict.keys():
            attrname = 'anz' + ky[0] + ky[1]
            setattr(self, attrname, sex_age_dict[ky])

    def _hh_marit_stats(self, pers):
        # This function creates new predictors: it counts how many HH-members belong to which category of marital status

        # In a first step, a dict which carries the counts for each marital status is created. At the same time a dict
        # which acts as a switch/case-statement is created:
        marit_dict = {}
        switchmarit = {}
        for i, marit in enumerate(['unwed', 'married', 'wid', 'div']):
            switchmarit.update({i+1: marit})
            marit_dict.update({marit: 0})

        # Go through all HH-members and use the switch/case-dict to determine the key for the "count-dict":
        for p in pers:
            ky = switchmarit.get(p['zivilstand03'])
            marit_dict[ky] += 1

        # Translate the marital-count-dict into attributes of the class:
        for ky in marit_dict.keys():
            attrname = 'anz_' + ky
            setattr(self, attrname, marit_dict[ky])

    def _hh_nat_stats(self, pers):
        # This function creates new predictors: it counts how many HH-members are Swiss and how many foreigners

        # In a first step, a dict which carries the counts for each nationality status is created. At the same time a dict
        # which acts as a switch/case-statement is created:
        nat_dict = {}
        switchnat = {}
        for i, nat in enumerate(['ch', 'ausl']):
            switchnat.update({i+1: nat})
            nat_dict.update({nat: 0})

        # Go through all HH-members and use the switch/case-dict to determine the key for the "count-dict":
        for p in pers:
            ky = switchnat.get(p['nationalitaet01'])
            nat_dict[ky] += 1

        # Translate the nationality-count-dict into attributes of the class:
        for ky in nat_dict.keys():
            attrname = 'anz_' + ky
            setattr(self, attrname, nat_dict[ky])


def habe_hh_predictor_builder():
    # -------------------------------------------------------------
    # ATTENTION: this function is more or less deprecated and replaced by habe_hh_preparer see below
    # -------------------------------------------------------------
    # This function is used to translate the HH-characteristics provided by the HABE into meaningful predictors (see also
    # "D:\froemelt\Documents\Andi\02_Doktorat\03_Projects\05_HEIA\03_Computations\Consumption\Datascreening.xlsx"). The
    # outcome is a PG-table with predictors which are either easily matchable with STATPOP-data or which need further
    # consideration whether they need to be "pre-modelled"

    # connecting to the database:
    conn = pg.connect("host=localhost port=5432 dbname=db_andi user=postgres  connect_timeout=2 password='postgres'")

    # the following list lists the attributes which need to be retrieved from the habe-standard-table
    stdvars = [
        'haushaltid',
      'e10',
      'primaereinkommen08',
      'bruttoeinkommen08',
      'verfuegbareseinkommen08',
      'sparbetrag08',
      'anzahlselbstaendiger05',
      'anzahlunselbstaendiger05',
      'anzahlrentner05',
      'anzahlausbildung05',
      'anzahlandere05',
      'mieterhaushalt05',
      'mindestenseinauto05',
      'mindestenseinvelo05',
      'mindestenseincomputer05',
      'mindestenseinnatel05',
      'mindestenseinhaustier05',
      'grossregion01',
      'sprachregion98',
      'kanton08'
    ]

    # the following list lists the attributes which need to be retrieved from the habe-personen-table
    persvars = [
    'personid',
    'lebensalter98',
    'geschlecht98',
    'zivilstand03',
    'nationalitaet01'
    ]

    # in a first step, all household-IDs are retrieved:
    cur = conn.cursor()
    query = """
    SELECT DISTINCT haushaltid FROM original_data.habe_standard
    """
    cur.execute(query)
    hhids = []
    for row in cur:
        hhids.append(row[0])
    cur.close()

    # container for the resulting dicts which carry the new predictors:
    habehhlist = []

    # core lines: go through all households:
    for hhid in hhids:
        # retrieve relevant data for each person of a certain household with a join-statement
        cur = conn.cursor(cursor_factory=pge.RealDictCursor)
        query = """
        SELECT {}, {} FROM original_data.habe_personen pers LEFT JOIN original_data.habe_standard std
        ON pers.haushaltid = std.haushaltid WHERE pers.haushaltid = {}
        """.format(', '.join(["pers."+p for p in persvars]), ', '.join(["std."+s for s in stdvars]), hhid)
        cur.execute(query)
        perssql = cur.fetchall()

        # Instantiate a HABE_HH-class. This will automatically initialize the needed processing for the new predictors
        habehh = HABE_HH(perssql)

        # convert the class to a dict:
        habehhdict = habehh.__dict__

        # add the HH under consideration to the results-list
        habehhlist.append(habehhdict)

    # The following list contains the names of the newly created predictors which will be stored in the PG-database:
    attrlist = [
    'haushaltid bigint',
    'anz_ch int',
    'anz_ausl int',
    'anz_unwed int',
    'anz_married int',
    'anz_wid int',
    'anz_div int',
    'sprachregion98_dch int',
    'sprachregion98_fch int',
    'sprachregion98_ich int',
    'grossregion01_ge int',
    'grossregion01_mit int',
    'grossregion01_nw int',
    'grossregion01_ost int',
    'grossregion01_ti int',
    'grossregion01_zen int',
    'grossregion01_zh int',
    'kanton08_ag int',
    'kanton08_be int',
    'kanton08_ge int',
    'kanton08_lu int',
    'kanton08_sg int',
    'kanton08_ti int',
    'kanton08_vd int',
    'kanton08_zh int',
    'kanton08_rest int',
    'anzfem0004 int',
    'anzfem0514 int',
    'anzfem1524 int',
    'anzfem2534 int',
    'anzfem3544 int',
    'anzfem4554 int',
    'anzfem5564 int',
    'anzfem6574 int',
    'anzfem7599 int',
    'anzmale0004 int',
    'anzmale0514 int',
    'anzmale1524 int',
    'anzmale2534 int',
    'anzmale3544 int',
    'anzmale4554 int',
    'anzmale5564 int',
    'anzmale6574 int',
    'anzmale7599 int',
    'prov_e10 float',
    'prov_primaereinkommen08 float',
    'prov_bruttoeinkommen08 float',
    'prov_verfuegbareseinkommen08 float',
    'prov_sparbetrag08 float',
    'prov_anzahlausbildung05 int',
    'prov_anzahlselbstaendiger05 int',
    'prov_anzahlunselbstaendiger05 int',
    'prov_anzahlrentner05 int',
    'prov_anzahlandere05 int',
    'prov_mieterhaushalt05 int',
    'prov_mindestenseinauto05 int',
    'prov_mindestenseinvelo05 int',
    'prov_mindestenseincomputer05 int',
    'prov_mindestenseinnatel05 int',
    'prov_mindestenseinhaustier05 int'
    ]

    # Create the new PG-table:
    writecur = conn.cursor()
    query = """
    CREATE TABLE working_tables.habe_hh_predictors
    ({}, CONSTRAINT habepredictors_pkey PRIMARY KEY (haushaltid))
    """.format(', '.join(attrlist))
    writecur.execute(query)
    conn.commit()

    # Create a list of attributes which will be used to insert the values:
    cols = [s.split(' ')[0] for s in attrlist]

    # Insert all the values to the new PG-Table
    query = """
    INSERT INTO working_tables.habe_hh_predictors(%s)
    VALUES (%s(%s)s)
    """ % (', '.join(cols), '%', ')s, %('.join(cols))
    writecur.executemany(query, habehhlist)
    conn.commit()

    # Finally: create an index on the HH-ID
    writecur.execute("""
    CREATE INDEX habepredictors_hhid
    ON working_tables.habe_hh_predictors
    USING btree
    (haushaltid);
    """)
    conn.commit()

    writecur.close()
    cur.close()
    conn.close()


class HABE_HH_V2(object):
    """
    This class is used by habe_hhs_preparer(). The goal is to process these
    attributes in a way to derive new predictors which can be easily matched with STATPOP or which shall be further
    investigated (see also "D:\froemelt\Documents\Andi\02_Doktorat\03_Projects\05_HEIA\03_Computations\02_Consumption\Consumption_Cockpit.xlsx").
    
    Furthermore, this class also converts water and wastewater expenditures into m3, electricity expenditures into kWh,
    energy expenditures into MJ (lower heating value) and waste into number of waste bags.
    Finally it also fills data gaps of water via wastewater amounts and vice versa.
    """

    def __init__(self, pers, codeclasses, codetransl, pricedicts, hh_nk_model, **kwargs):
        """
        classical init-function for class HABE_HH_V2
        :param pers: list of dicts, Thereby, the dicts represent persons of a certain household with attributes defined by HABE-tables
        :param codeclasses: tuple of original HABE-codes for amounts, expenditures, durable goods and income
        :param codetransl: dict which translates original HABE-codes into codes which will be used in the new pg-table
        :param pricedicts: dict of dicts with all the relevant information to convert expenditures into amounts (above all price information)
        :param hh_nk_model: a dict which will be used to flag HH which have imputed water/wastewater data
        """
        self._codetransl_ = codetransl
        self._pricedicts_ = pricedicts
        self._hh_wide_attr(pers[0], codeclasses, hh_nk_model)  # compute household-wide attributes --> therefore pass just first person which has all information
        self._hh_wide_geogr(**pers[0])
        self._hh_age_sex_stats(pers)
        self._hh_marit_stats(pers)
        self._hh_nat_stats(pers)

    def _hh_wide_attr(self, values, codeclasses, hh_nk_model, **kwargs):
        """
        This function fills in the attributes which are readily available in the HABE-tables and which are
        the same for all household members. It also converts water/wastewater to m3, electricity to kWh, waste to waste
        bags, and energy to MJ (lower heating value).
        """
        self.haushaltid = values['haushaltid']

        # in a first step, we fill in all the attributes of the household and create dynamically attributes with the
        # code-name that will be used in the PG-table-DB
        for codes in codeclasses:
            for code in codes:
                try:
                    setattr(self, self._codetransl_[code], values[code])
                except:
                    setattr(self, self._codetransl_[code], 0)

        # in order to convert water, wastewater and waste expenditures into amounts, we determine the "hh-type" by the amount
        # of persons per HH
        if values['anzahlpersonen98'] == 1:
            hhtype = '1pHH'
        elif values['anzahlpersonen98'] == 2:
            hhtype = '2pHH'
        elif values['anzahlpersonen98'] == 3:
            hhtype = '3pHH'
        else:
            hhtype = '4pHH'

        # by means of the water-prices-dict, we convert the water expenditures into m3
        waterexpcode = 'a571204'
        water_prices_dict = self._pricedicts_['water']
        hh_water_price = water_prices_dict[(values['kanton08'], hhtype)]  # get the price for the HH based on canton and HH-type
        hh_water_amount = values[waterexpcode] / hh_water_price
        #self.mx571203 = hh_water_amount
        self.mx571204 = hh_water_amount
        del water_prices_dict, hh_water_amount

        # by means of the wastewater-prices-dict, we convert the water expenditures into m3
        wwexpcode = 'a571203'
        ww_prices_dict = self._pricedicts_['ww']
        hh_ww_price = ww_prices_dict[(values['kanton08'], hhtype)]  # get the price for the HH based on canton and HH-type
        hh_ww_amount = values[wwexpcode] / hh_ww_price
        self.mx571203 = hh_ww_amount
        #self.mx571204 = hh_ww_amount
        del ww_prices_dict, hh_ww_amount

        # The following lines do a simple data imputation: if we do not have information for water expenditures, but for
        # wastewater, then we assume that amount of water = amount of wastewater (which is plausible given how prices
        # are set for water/wastewater in Switzerland); Furthermore, based on the water price we also compute the expenditures
        # for water supply. Of course, the same procedure is also done vice versa. Since we assume (which is also plausible)
        # that missing expenditures are hidden in "Nebenkosten pauschal", we need to correct for the imputed data and
        # subtract the imputed water/wastewater expenditures accordingly.
        if self.mx571204 == 0 and self.mx571203 != 0:
            self.mx571204 = self.mx571203
            self.a571204 = self.mx571203 * hh_water_price
            hh_nk_model['wviaaw'] = 1  # flag the HH for which water data was imputed via wastewater
            new_nk_amount = self.a571201 - self.a571204  # compute the new "Nebenkosten pauschal"
            if new_nk_amount < 0:  # if - for some reason - the new amount would be zero, then we set the NK to zero and update all depending statistics
                self.a571201 = 0
                self.a5712 = self.a571201 + self.a571202 + self.a571203 + self.a571204 + self.a571205
                self.a571 = self.a5711 + self.a5712 + self.a5713
                self.a57 = self.a571 + self.a572 + self.a573
                self.a50 = self.a51 + self.a52 + self.a53 + self.a56 + self.a57 + self.a58 + self.a61 + self.a62 + self.a63 + self.a66 + self.a67 + self.a68
        elif self.mx571203 == 0 and self.mx571204 != 0:
            self.mx571203 = self.mx571204
            self.a571203 = self.mx571204 * hh_ww_price
            hh_nk_model['awviaw'] = 1  # flag the HH for which wastewater data was imputed via water
            new_nk_amount = self.a571201 - self.a571203  # compute the new "Nebenkosten pauschal"
            if new_nk_amount < 0:  # if - for some reason - the new amount would be zero, then we set the NK to zero and update all depending statistics
                self.a571201 = 0
                self.a5712 = self.a571201 + self.a571202 + self.a571203 + self.a571204 + self.a571205
                self.a571 = self.a5711 + self.a5712 + self.a5713
                self.a57 = self.a571 + self.a572 + self.a573
                self.a50 = self.a51 + self.a52 + self.a53 + self.a56 + self.a57 + self.a58 + self.a61 + self.a62 + self.a63 + self.a66 + self.a67 + self.a68

        # finally, we also consider the case if we have water and wastewater expenditures available but which come to different
        # amounts of water --> in this case, take the mean of the two estimates
        elif self.mx571203 != 0 and self.mx571204 != 0:
            waw_mean = (self.mx571203 + self.mx571204) / 2
            self.mx571203 = waw_mean
            self.mx571204  = waw_mean

        # by means of the waste-prices-dict, we convert the waste expenditures into amounts of waste bags
        wasteexpcode = 'a571202'
        waste_prices_dict = self._pricedicts_['waste']
        if values['kanton08'] == 25:  # this if-loop is necessary since for Geneva (canton-id 25), no waste prices are available
            hh_waste_price = 0
            hh_waste_amount = 0
        else:
            hh_waste_price = waste_prices_dict[(values['kanton08'], hhtype)]
            hh_waste_amount = values[wasteexpcode] / hh_waste_price
        self.mx571202 = hh_waste_amount
        del waste_prices_dict, hh_waste_price, hh_waste_amount

        # The following lines convert the expenditures for electricity to kWh based on the electricity-prices-dict. The
        # conversion is done based on canton, year and household-type. However, household types are not defined by no.
        # of HH-member but rather depending on the amount of electricity used. Therefore, we start with the "lowest"-consumption
        # category and estimate the amount of electricity based on this price. If this exceeds the amount of electricity
        # for the lowest category, the next category is tested --> proceed accordingly until correct HH-consumption-category is found
        elexpcode = 'a571301'
        electricity_prices_dict = self._pricedicts_['electricity']
        hh_electricity_price = electricity_prices_dict[(values['jahr08'], values['kanton08'], 'H1')]
        hh_electricity_amount = values[elexpcode] / (hh_electricity_price/100)  # division by 100 since price is given in Rp. per kWh.
        if hh_electricity_amount*12 > (2500+1600)/2:  # multiplication with 12 because we need to compare the annual electricity consumption
            hh_electricity_amount = values[elexpcode] / (electricity_prices_dict[(values['jahr08'], values['kanton08'], 'H2')]/100)
            if hh_electricity_amount*12 > (4500+2500)/2:
                hh_electricity_amount = values[elexpcode] / (((
                electricity_prices_dict[(values['jahr08'], values['kanton08'], 'H3')] +
                electricity_prices_dict[(values['jahr08'], values['kanton08'], 'H4')])/2)/ 100)  # Categories H3 and H4 show the same annual electricity consumption,
                                                                                                 # they differ in the time when electricity is used. We decided to take
                                                                                                 # the average of both categories.
                if hh_electricity_amount * 12 > (4500 + 7500) / 2:
                    hh_electricity_amount = values[elexpcode] / (
                    electricity_prices_dict[(values['jahr08'], values['kanton08'], 'H5')] / 100)
                    if hh_electricity_amount * 12 > (7500 + 13000) / 2:
                        hh_electricity_amount = values[elexpcode] / (
                            electricity_prices_dict[(values['jahr08'], values['kanton08'], 'H7')] / 100)
                        if hh_electricity_amount * 12 > (13000 + 25000) / 2:
                            hh_electricity_amount = values[elexpcode] / (
                                electricity_prices_dict[(values['jahr08'], values['kanton08'], 'H6')] / 100)
        self.mx571301 = hh_electricity_amount
        del electricity_prices_dict, hh_electricity_price, hh_electricity_amount

        # the conversion from energy expenditures to amount of energy is conducted with a separate function
        fuelexpcode = 'a571302'
        centralheatexpcode = 'a571303'
        self.mx571302 = self._compute_heat_energy(values[fuelexpcode], values['jahr08'])
        self.mx571303 = self._compute_heat_energy(values[centralheatexpcode], values['jahr08'])

    def _compute_heat_energy(self, expvalue, yr):
        """
        This function converts energy expenditures to amounts of energy used. Since we do not know which energy carrier
        is used by the household, we base our energy amount estimate on a energy mix (building an average). This means
        we convert the expenditures into amounts by assuming that the energy carrier was oil in a first step, in a second
        step, we will do the same assuming that the energy carrier was gas and finally we compute the amount of energy
        based on firewood-prices. In the end we take a weighted average of all three estimates. Thereby, weighting is performed
        based on final energy consumption by households according to the Gesamtenergiestatistik.
        """

        # Analogously to electricity, the price for oil depends on the amount of consumption. Therefore, the oil price
        # is determined again iteratively.
        oil_prices_dict = self._pricedicts_['energy']['oil_prices']
        hh_oil_price = oil_prices_dict[(yr, '800 - 1500')]
        hh_oil_amount = (expvalue / hh_oil_price) * 100  # because price in per 100 l
        if hh_oil_amount * 12 > 1500:
            hh_oil_amount = (expvalue / oil_prices_dict[(yr, '1501 - 3000')]) * 100
            if hh_oil_amount * 12 > 3000:
                hh_oil_amount = (expvalue / oil_prices_dict[(yr, '3001 - 6000')]) * 100
                if hh_oil_amount * 12 > 6000:
                    hh_oil_amount = (expvalue / oil_prices_dict[(yr, '6001 - 9000')]) * 100
                    if hh_oil_amount * 12 > 9000:
                        hh_oil_amount = (expvalue / oil_prices_dict[(yr, '9001 - 14000')]) * 100
                        if hh_oil_amount * 12 > 14000:
                            hh_oil_amount = (expvalue / oil_prices_dict[(yr, '14001 - 20000')]) * 100
                            if hh_oil_amount * 12 > 20000:
                                hh_oil_amount = (expvalue / oil_prices_dict[(yr, '>20000')]) * 100
        hh_oil_amount *= self._pricedicts_['energy']['oilenergydensity']  # convert oil amount from liters to MJ based on data from ecoinvent 3 data quality guidelines
        del hh_oil_price, oil_prices_dict

        # Just as for electricity and oil, the price for gas depends on the consumed amount.
        gas_prices_dict = self._pricedicts_['energy']['gas_prices']
        hh_gas_price = gas_prices_dict[(yr, '20000')]
        hh_gas_amount = expvalue / hh_gas_price
        if hh_gas_amount * 12 > (20000 + 50000) / 2:
            hh_gas_amount = expvalue / gas_prices_dict[(yr, '50000')]
            if hh_gas_amount * 12 > (50000 + 100000) / 2:
                hh_gas_amount = expvalue / gas_prices_dict[(yr, '100000')]
                if hh_gas_amount * 12 > (100000 + 500000) / 2:
                    hh_gas_amount = expvalue / gas_prices_dict[(yr, '500000')]

        # The amount of gas is given in kWh upper calorific value (according to a phone call to Werkbetriebe Frauenfeld and a
        # a phone call to BFE). Therefore, we need to convert the UHV to LHV and kWh to MJ in the next line:
        hh_gas_amount *= (3.6 * self._pricedicts_['energy']['lhvgasconversion'])
        del gas_prices_dict, hh_gas_price

        # by means of the firewood price, we convert the energy expenditures also to energy amounts
        wood_prices_dict = self._pricedicts_['energy']['wood_prices']
        hh_wood_price = wood_prices_dict[yr]
        hh_wood_amount = (expvalue / hh_wood_price) * 6000 # because price in per 6000 kg
        hh_wood_amount *= self._pricedicts_['energy']['lhvwood']

        energy_shares = self._pricedicts_['energy']['shares']

        # finally, we return the share-weighted average of the three energy amounts estimates
        return hh_oil_amount * energy_shares[(yr, 'oil')] + hh_gas_amount * energy_shares[(yr, 'gas')] + hh_wood_amount * energy_shares[(yr, 'wood')]  # lower calorific value in MJ

    def _hh_wide_geogr(self, kanton08, sprachregion98, grossregion01, **kwargs):
        # This function creates dummy variables based on the geographic information from HABE. This could be derived from
        # STATPOP (partly in combination with ARE/BFS-Raumgliederungen).
        self.char_langregion_dch = 1 if sprachregion98 == 1 else 0
        self.char_langregion_fch = 1 if sprachregion98 == 2 else 0
        self.char_langregion_ich = 1 if sprachregion98 == 3 else 0

        self.char_kanton_zh = 1 if kanton08 == 1 else 0
        self.char_kanton_be = 1 if kanton08 == 2 else 0
        self.char_kanton_lu = 1 if kanton08 == 3 else 0
        self.char_kanton_sg = 1 if kanton08 == 17 else 0
        self.char_kanton_ag = 1 if kanton08 == 19 else 0
        self.char_kanton_ti = 1 if kanton08 == 21 else 0
        self.char_kanton_vd = 1 if kanton08 == 22 else 0
        self.char_kanton_ge = 1 if kanton08 == 25 else 0
        self.char_kanton_rest = 1 if kanton08 == 99 else 0

        self.char_georegion_ge = 1 if grossregion01 == 1 else 0
        self.char_georegion_mit = 1 if grossregion01 == 2 else 0
        self.char_georegion_nw = 1 if grossregion01 == 3 else 0
        self.char_georegion_zh = 1 if grossregion01 == 4 else 0
        self.char_georegion_ost = 1 if grossregion01 == 5 else 0
        self.char_georegion_zen = 1 if grossregion01 == 6 else 0
        self.char_georegion_ti = 1 if grossregion01 == 7 else 0

    def _hh_age_sex_stats(self, pers):
        # This function creates new household characteristic statistics: it uses the same age categories as HABE, but distinguishes now also
        # between male and female persons.

        # Create a dict to count the persons of each sex/age-group:
        sex_age_dict = {}
        for sex in ['fem', 'male']:
            for age in ['0004', '0514', '1524', '2534', '3544', '4554', '5564', '6574', '7599']:
                sex_age_dict.update({(sex, age): 0})

        # Go through all HH members and create a tuple-key for the sex_age_dict
        for p in pers:
            sex = 'fem' if p['geschlecht98'] == 2 else 'male'

            if p['lebensalter98'] <= 4:
                age = '0004'
            elif 5 <= p['lebensalter98'] <= 14:
                age = '0514'
            elif 15 <= p['lebensalter98'] <= 24:
                age = '1524'
            elif 25 <= p['lebensalter98'] <= 34:
                age = '2534'
            elif 35 <= p['lebensalter98'] <= 44:
                age = '3544'
            elif 45 <= p['lebensalter98'] <= 54:
                age = '4554'
            elif 55 <= p['lebensalter98'] <= 64:
                age = '5564'
            elif 65 <= p['lebensalter98'] <= 74:
                age = '6574'
            elif 75 <= p['lebensalter98']:
                age = '7599'

            # Increase count of the age/sex-group to which the HH-member under consideration belongs:
            sex_age_dict[(sex, age)] += 1

        # Translate the sex_age_dict into attributes of the class:
        for ky in sex_age_dict.keys():
            attrname = 'char_no' + ky[0] + ky[1]
            setattr(self, attrname, sex_age_dict[ky])

    def _hh_marit_stats(self, pers):
        # This function creates new household characteristics statistics: it counts how many HH-members belong to which category of marital status

        # In a first step, a dict which carries the counts for each marital status is created. At the same time a dict
        # which acts as a switch/case-statement is created:
        marit_dict = {}
        switchmarit = {}
        for i, marit in enumerate(['unwed', 'married', 'wid', 'div']):
            switchmarit.update({i+1: marit})
            marit_dict.update({marit: 0})

        # Go through all HH-members and use the switch/case-dict to determine the key for the "count-dict":
        for p in pers:
            ky = switchmarit.get(p['zivilstand03'])
            marit_dict[ky] += 1

        # Translate the marital-count-dict into attributes of the class:
        for ky in marit_dict.keys():
            attrname = 'char_no' + ky
            setattr(self, attrname, marit_dict[ky])

    def _hh_nat_stats(self, pers):
        # This function creates new Household characteristics statistics: it counts how many HH-members are Swiss and how many foreigners

        # In a first step, a dict which carries the counts for each nationality status is created. At the same time a dict
        # which acts as a switch/case-statement is created:
        nat_dict = {}
        switchnat = {}
        for i, nat in enumerate(['ch', 'ausl']):
            switchnat.update({i+1: nat})
            nat_dict.update({nat: 0})

        # Go through all HH-members and use the switch/case-dict to determine the key for the "count-dict":
        for p in pers:
            ky = switchnat.get(p['nationalitaet01'])
            nat_dict[ky] += 1

        # Translate the nationality-count-dict into attributes of the class:
        for ky in nat_dict.keys():
            attrname = 'char_no' + ky
            setattr(self, attrname, nat_dict[ky])


def get_price_dicts(excel):
    """
    This is a helper function which reads the prices given by the excel-file and returns a dict of dicts containing all
    the relevant information to convert expenditures to amounts
    :param excel: dict which needs the following keys: 'path', 'overview', 'water', 'electricity', 'waste',
    'energy'. While 'path' should be clear, the other keys shall define the sheet-names in which the specific prices for
    the conversion from CHF to amounts are given, except for 'overview' which means the 'overview'-sheet
    :return: dict of dicts containing information to convert expenditures to amounts
    """

    # open excel-workbook
    wb = xlrd.open_workbook(excel['path'])

    # open worksheet with water prices information and get id-numbers of cantons
    ws_water = wb.sheet_by_name(excel['water'])
    cantonids = ws_water.col_values(0)
    rowstart = cantonids.index('py_start_read')  # find row where to start reading price information
    cantonids = cantonids[rowstart + 2:]

    # create a dict containing all the price information for water (per canton and household-type)
    water_prices_dict = {}
    for i, cid in enumerate(cantonids):
        for j, hhtype in enumerate(['1pHH', '2pHH', '3pHH', '4pHH']):
            water_prices_dict.update({(int(cid), hhtype): ws_water.cell_value(rowstart + 2 + i, 2 + j)})
    del ws_water, rowstart

    # open worksheet with wastewater prices information
    ws_ww = wb.sheet_by_name(excel['ww'])
    rowstart = ws_ww.col_values(0).index('py_start_read')  # find row where to start reading price information

    # create a dict containing all the price information for wastewater (per canton and household-type)
    ww_prices_dict = {}
    for i, cid in enumerate(cantonids):
        for j, hhtype in enumerate(['1pHH', '2pHH', '3pHH', '4pHH']):
            ww_prices_dict.update({(int(cid), hhtype): ws_ww.cell_value(rowstart + 2 + i, 2 + j)})
    del ws_ww, rowstart

    # open worksheet with waste prices information
    ws_waste = wb.sheet_by_name(excel['waste'])
    rowstart = ws_waste.col_values(0).index('py_start_read')  # find row where to start reading price information

    # create a dict containing all the price information for waste (per canton and household-type)
    waste_prices_dict = {}
    for i, cid in enumerate(cantonids):
        for j, hhtype in enumerate(['1pHH', '2pHH', '3pHH', '4pHH']):
            waste_prices_dict.update({(int(cid), hhtype): ws_waste.cell_value(rowstart + 2 + i, 2 + j)})
    del ws_waste, rowstart

    # open worksheet with electricity prices information
    ws_electricity = wb.sheet_by_name(excel['electricity'])
    indexcol = ws_electricity.col_values(0)  # get first col, which also contains years
    rowstart = indexcol.index('py_start_read')  # find row where to start reading price information
    indexcol = indexcol[rowstart:]

    # create a dict containing all the price information for electricity (per canton, household-type and year)
    electricity_prices_dict = {}
    for yr in [2009, 2010, 2011]:
        ind = indexcol.index(yr) + rowstart
        for i, cid in enumerate(cantonids):
            for j, hhtype in enumerate(['H' + str(n) for n in range(1, 8)]):
                electricity_prices_dict.update(
                    {(int(yr), int(cid), hhtype): ws_electricity.cell_value(ind + 2 + i, 2 + j)})
    del ws_electricity, rowstart, indexcol, ind

    # open worksheet with energy prices information and energy shares information
    energy_shares_dict = {}
    ws_energy = wb.sheet_by_name(excel['energy'])
    indexcol = ws_energy.col_values(0)  # get first column which contains information where to start reading for the different energy carriers
    shrstart = indexcol.index('py_start_read_shares')  # find row where to start reading energy shares information
    oilstart = indexcol.index('py_start_read_oil')  # find row where to start reading fuel oil price information
    gasstart = indexcol.index('py_start_read_gas')  # find row where to start reading gas price information
    woodstart = indexcol.index('py_start_read_wood')  # find row where to start reading wood price information

    # create a dict for the energy shares (per year and per energy carrier)
    for i, yr in enumerate([2009, 2010, 2011]):
        for j, encarrier in enumerate(['oil', 'gas', 'wood']):
            energy_shares_dict.update({(yr, encarrier): ws_energy.cell_value(shrstart + 2 + i, 1 + j)})

    # create a dict for oil prices (per year and per quantity)
    oil_prices_dict = {}
    for i, yr in enumerate([2009, 2010, 2011]):
        for j, quantity in enumerate(['800 - 1500', '1501 - 3000', '3001 - 6000', '6001 - 9000', '9001 - 14000',
                                      '14001 - 20000', '>20000']):
            oil_prices_dict.update({(yr, quantity): ws_energy.cell_value(oilstart + 3 + i, 1 + j)})

    # further information is needed for oil to convert expenditures to MJ (lower heating value), e.g. density of oil and lower heating value
    oildensity = ws_energy.cell_value(oilstart + 7, 1)
    lhv = ws_energy.cell_value(oilstart + 8, 1)  # lower heating value
    oilenergydensity = oildensity * lhv

    # the following lines check if the density/lhv correspond to what would be expected and raises a warning if not
    if oildensity != 0.86:
        print('ATTENTION: oil density is {} instead of 0.86 kg/l, are you sure?'.format(oildensity))
    if lhv != 42.6:
        print('ATTENTION: lower calorific value of oil is {} instead of 42.6 MJ/kg, are you sure?'.format(lhv))

    # create a dict for gas prices (per year and quantity)
    gas_prices_dict = {}
    for i, yr in enumerate([2009, 2010, 2011]):
        for j, quantity in enumerate(['20000', '50000', '100000', '500000']):
            gas_prices_dict.update({(yr, quantity): ws_energy.cell_value(gasstart + 2 + i, 1 + j)})

    # to convert expenditures for gas into MJ (lower heating value) we need some more information:
    uhv = ws_energy.cell_value(gasstart + 6, 1)  # upper heating value
    lhv = ws_energy.cell_value(gasstart + 7, 1)  # lower heating value
    lhvgasconversion = lhv / uhv

    # the following lines check if the uhv/lhv correspond to what would be expected and raises a warning if not
    if uhv != 50.4:
        print('ATTENTION: upper calorific value of gas is {} instead of 50.4 MJ/kg, are you sure?'.format(uhv))
    if lhv != 45.4:
        print('ATTENTION: lower calorific value of gas is {} instead of 45.4 MJ/kg, are you sure?'.format(lhv))

    # create a dict for wood prices (per year and quantity)
    wood_prices_dict = {}
    for i, yr in enumerate([2009, 2010, 2011]):
        wood_prices_dict.update({yr: ws_energy.cell_value(woodstart + 2 + i, 1)})

    # to convert expenditures for wood into MJ (lower heating value) we need some more information:
    lhvwood = ws_energy.cell_value(woodstart + 6, 1)  # lower heating value

    # the following lines check if the lhv corresponds to what would be expected and raises a warning if not
    if lhvwood != 18:
        print('ATTENTION: lower calorific value of wood is {} instead of 18 MJ/kg, are you sure?'.format(lhvwood))

    # summarize and store all energy-dicts in one big energy-dict:
    energy_dict = {'shares': energy_shares_dict, 'oil_prices': oil_prices_dict, 'gas_prices': gas_prices_dict,
                   'wood_prices': wood_prices_dict,
                   'oilenergydensity': oilenergydensity, 'lhvgasconversion': lhvgasconversion, 'lhvwood': lhvwood}

    return {'water': water_prices_dict, 'ww': ww_prices_dict, 'waste': waste_prices_dict,
                             'electricity': electricity_prices_dict, 'energy': energy_dict}


def habe_hhs_preparer(excel):
    """
    This function is used to build a large PG-table which contains all information that will be used for pattern recognition
    and consumption modelling. However, the resulting PG-table will contain too much information and needs filtering
    before use. Furthermore, some HH-characteristics provided by the HABE are translated into meaningful predictors
    (see also "D:\froemelt\Documents\Andi\02_Doktorat\03_Projects\05_HEIA\03_Computations\02_Consumption\Consumption_Cockpit.xlsx").
    This is thus a successor function of habe_hh_predictor and provides likewise characteristics which are easily
    matchable with STATPOP-data. Moreover, it will also convert some expenditures into amounts and fill some data gaps for water
    and wastewater.
    
    PLEASE BE CAREFUL: The columns "Variable-Code", "Amounts" and "pg-code 1"/"pg-code 2"  in the excel-file should not
    change names! Furthermore, there is one important location in the code (indicated with XXX) which is very specificly
    tailored to the excel file and which needs to be carefully checked if this still applies!
    
    :param excel: dict which needs the following keys: 'path', 'overview', 'water', 'electricity', 'waste',
    'energy'. While 'path' should be clear, the other keys shall define the sheet-names in which the specific prices for
    the conversion from CHF to amounts are given, except for 'overview' which means the 'overview'-sheet
    """

    # connecting to the database:
    conn = pg.connect("host=localhost port=5432 dbname=db_andi user=postgres  connect_timeout=2 password='postgres'")

    # the following list lists the attributes which need to be retrieved from the habe-standard-table
    stdvars = [
        'haushaltid',
        'e10',
        'e11',
        'e12',
        'e15',
        'primaereinkommen08',
        'e20',
        'e21',
        'e22',
        'e23',
        'e25',
        'bruttoeinkommen08',
        'verfuegbareseinkommen08',
        'e70',
        'sparbetrag08',
        'anzahlpersonen98',
        'anzahlselbstaendiger05',
        'anzahlunselbstaendiger05',
        'anzahlrentner05',
        'anzahlausbildung05',
        'anzahlandere05',
        'mieterhaushalt05',
        'fraualsreferenzperson05',
        'mindestenseinhaustier05',
        'grossregion01',
        'sprachregion98',
        'kanton08',
        'jahr08'
    ]

    # the following list lists the attributes which need to be retrieved from the habe-personen-table
    persvars = [
        'personid',
        'lebensalter98',
        'geschlecht98',
        'zivilstand03',
        'nationalitaet01'
    ]

    # open the overview-sheet of the excel-cockpit containing all the relevant information
    wb = xlrd.open_workbook(excel['path'])
    ws = wb.sheet_by_name(excel['overview'])

    # read all the original codes of HABE for amounts and expenditures
    headrow = ws.col_values(0).index('header')  # find the header row
    mengind = ws.row_values(headrow).index('Amounts')
    ausgind = ws.row_values(headrow).index('Variable-Code')
    mengcodes = ws.col_values(mengind)
    ausgcodes = ws.col_values(ausgind)

    # create a translator-dict to translate the original HABE-code into the pg-table-code which will be used in the
    # pg-table that is going to be created by this function
    codetransl = {}
    for c, code in enumerate([mengcodes, ausgcodes]):
        if c == 0:  # amount-codes
            x = ws.row_values(headrow).index('pg-code 2')
        else:  # expenditure-codes
            x = ws.row_values(headrow).index('pg-code 1')
        for i, var in enumerate(code):
            if var not in ['Variable-Code', '', '1: yes', '0: no', 'Amounts', 0.0]:
                codetransl.update({var.lower(): ws.cell_value(i, x)})

    # In the following, the amount-codes, durable-goods-codes and the expenditure-codes are separated
    # XXX: ATTENTION: the following lines are very specifically tailored to the excel file and should be checked
    # before using the function!
    mengcodes = [m.lower() for m in mengcodes[82:613] if m != 0.0 and m != '']
    konscodes = [a.lower() for a in ausgcodes[613:633]]
    ausgcodes = [a.lower() for a in ausgcodes[82:613]]

    # The following function retrieves the prices and creates a prices-dict which will be used to convert expenditures to
    # amounts
    price_dicts = get_price_dicts(excel)

    # in a first step, all household-IDs are retrieved:
    cur = conn.cursor()
    query = """
    SELECT DISTINCT haushaltid FROM original_data.habe_standard
    """
    cur.execute(query)
    hhids = []
    for row in cur:
        hhids.append(row[0])
    cur.close()

    # container for the resulting dicts which carry the new predictors:
    habehhlist = []

    # container for the ids of the household for which the water amounts/expenditures were imputed via wastewater or vice versa
    # this will be used to create a pg-table that provides an overview of all HH which have imputed dwelling extra cost data
    habenkmodellist = []

    # core lines: go through all households:
    cur = conn.cursor(cursor_factory=pge.RealDictCursor)

    for hhid in hhids:
        # retrieve relevant data for each person of a certain household with a join-statement
        query = """
        SELECT {}, {}, {}, {}, {} FROM original_data.habe_personen pers LEFT JOIN original_data.habe_standard std
        ON pers.haushaltid = std.haushaltid LEFT JOIN original_data.habe_ausgaben ausg ON pers.haushaltid = ausg.haushaltid
        LEFT JOIN original_data.habe_mengen meng ON pers.haushaltid = meng.haushaltid LEFT JOIN original_data.habe_konsumgueter kons
        ON pers.haushaltid = kons.haushaltid WHERE pers.haushaltid = {}
        """.format(', '.join(["pers." + p for p in persvars]), ', '.join(["std." + s for s in stdvars]),
                   ', '.join(["ausg." + a for a in ausgcodes]), ', '.join(["meng." + m for m in mengcodes]),
                   ', '.join(["kons." + k for k in konscodes]), hhid)
        cur.execute(query)
        perssql = cur.fetchall()

        # Create a dict which will be used to fill in a pg-table that provides an overview of the HH which contain modelled
        # dwelling extra cost values
        hh_nk_model = {'haushaltid': perssql[0]['haushaltid'], 'nkmod': 0, 'kmod': 0, 'awviaw': 0, 'awmod': 0,
                       'wviaaw': 0, 'wmod': 0, 'elmod': 0, 'enmod': 0}

        # Instantiate a HABE_HH_V2-class. This will automatically initialize the needed processing for each HH --> it will
        # create new statistics (e.g. no. of female persons in a certain age, etc.), compute binary/dummy variables, convert
        # some expenditures to amounts and impute some data gaps for water and wastewater
        habehh = HABE_HH_V2(perssql, (stdvars, ausgcodes, mengcodes, konscodes), codetransl, price_dicts, hh_nk_model)

        # convert the class to a dict:
        habehhdict = {attr: getattr(habehh, attr) for attr in dir(habehh) if
                      not attr.startswith('_') and not attr.endswith('_') and not attr == ''}

        # add the HH under consideration to the results-list
        habehhlist.append(habehhdict)
        habenkmodellist.append(hh_nk_model)

    # create a list of attribute names for the new pg-table which will contain all the consumption data:
    attrlist = ['haushaltid bigint']  # start with primary key
    for a in [attr for attr in dir(habehh) if not attr.startswith('_') and not attr.endswith('_') and not attr == '']:  # go through all attributes of a HH
        if a.lower().startswith('m') or a.lower().startswith('e') or a.lower().startswith('a'):
            a += ' float'  # specify "float" for expenditures, amounts and income
        elif a.lower().startswith('cg') or a.lower().startswith('char'):
            a += ' int'  # specify "integer" for durable goods and household characteristics
        else:
            continue
        attrlist.append(a)

    # Create the new PG-table --> if already exists, drop table; create also the pg-table that gives an overview of
    # households which have modelled values --> as this table was prepared before: truncate table
    writecur = conn.cursor()
    query = """
    DROP TABLE IF EXISTS working_tables.habe_hh_prepared
    """
    writecur.execute(query)
    conn.commit()

    query = """
    TRUNCATE TABLE working_tables.habe_nk_model
    """
    writecur.execute(query)
    conn.commit()

    query = """
    CREATE TABLE working_tables.habe_hh_prepared
    ({}, CONSTRAINT habehhprep_pkey PRIMARY KEY (haushaltid))
    """.format(', '.join(attrlist))
    writecur.execute(query)
    conn.commit()

    # Create a list of attributes which will be used to insert the values:
    cols = [s.split(' ')[0] for s in attrlist]

    # Insert all the values to the new PG-Table
    query = """
    INSERT INTO working_tables.habe_hh_prepared(%s)
    VALUES (%s(%s)s)
    """ % (', '.join(cols), '%', ')s, %('.join(cols))
    writecur.executemany(query, habehhlist)
    conn.commit()

    # Finally: create an index on the HH-ID
    writecur.execute("""
    CREATE INDEX habehhprep_hhid
    ON working_tables.habe_hh_prepared
    USING btree
    (haushaltid);
    """)
    conn.commit()

    # Create a list of attributes which will be used to insert the values to the pg-table that provides an overview of
    # the households with modelled dwelling extra cost data
    nkcols = [ky for ky in habenkmodellist[0].keys()]

    # Insert all the values to the dwelling extra cost modelling table
    query = """
        INSERT INTO working_tables.habe_nk_model (%s)
        VALUES (%s(%s)s)
        """ % (', '.join(nkcols), '%', ')s, %('.join(nkcols))
    writecur.executemany(query, habenkmodellist)
    conn.commit()

    writecur.close()
    cur.close()
    conn.close()


class Seasonal_Fig():
    """
    This class takes two attributes for which seasonal figures shall be created. It plots one variable in the upper part
    of a A4-landscape-page and the other one in the lower part. For each variable, three subplots are created:
    monthly average with data from all three years, monthly average for each year seperately, and annual averages. The error
    bars are the 95%-confidence interval of the average (determined by bootstrapping). In order to indicate, how strong
    seasonality is, the ANOVA-Test as well as the Kruskal-Wallis-Test is conducted for the monthly data from all years
    (so we investigate if the survey month is statistically significant to determine the household behavior --> remember:
    if p-value is <= alpha-level [significance level] then the Null-Hypothesis is rejected [Null-Hypothesis = "all means
    are the same" or for Kruskal-Wallis: "all medians are the same"] --> if rejected, then we have seasonality!)
    """
    def __init__(self, pp, dfs, var, varnames_dict, not_orig_table):
        """
        This init-function creates the seasonal graphs.
        :param pp: Multi-pages-PDF-object (the graphs are stored here)
        :param dfs: tuple of dataframes (a dataframe for each attribute)
        :param var: tuple/list of attributes-names (same order as dataframes in dfs is required)
        :param varnames_dict: translation dict (optional) to translate pg-codes into human readable names.
        :param not_orig_table: (optional) string with name of table from which data shall be retrieved (if data shall not come
        from original HABE-data) --> in this class, it is not used to retrieve data, but rather to indicate how to deal
        with variable-names
        """

        # Initializing the figure (A4-landscape)
        fig_width_cm = 21  # A4 page
        fig_height_cm = 29.7
        inches_per_cm = 1 / 2.58  # Convert cm to inches
        fig_width = fig_width_cm * inches_per_cm  # width in inches
        fig_height = fig_height_cm * inches_per_cm  # height in inches
        fig_size = [fig_height,
                    fig_width]  # height and width are in the order needed for landscape
        paper_rc = {'lines.linewidth': 1, 'lines.markersize': 10}  # parameters to control the linewidth and markersize
        sns.set_context("paper", rc=paper_rc)

        # if-loop to capture the case there is not translation-dict given (fake translation-dict is created)
        if not varnames_dict:
            varnames_dict = {}
            varnames_dict[var[0]] = var[0]
            varnames_dict[var[1]] = var[1]

        # In case, we work with original HABE-data, we have to do some extra translations of variable names
        if not not_orig_table:
            # The next dict shall help to convert pg-codes used in habe_hh_prepared and which might be retrieved from the Consumption_Cockpit_xlsx
            # to the original HABE-code.
            eink_dict = {'e_primaereink': 'Primaereinkommen08'.lower(), 'e_bruttoeink': 'Bruttoeinkommen08'.lower(),
                         'e_verfuegbeink': 'VerfuegbaresEinkommen08'.lower(), 'e_sparbetrag': 'Sparbetrag08'.lower()}

            # Translation of variables to the original HABE-name
            var[0] = eink_dict[var[0]] if var[0] in eink_dict.keys() else var[0]
            var[1] = eink_dict[var[1]] if var[1] in eink_dict.keys() else var[1]

        # Core: plotting the seasonal graphs
        fig = plt.figure(figsize=fig_size)
        plt.suptitle("top: {}, bottom: {}".format(varnames_dict[var[0]] if var[0] in varnames_dict.keys() else var[0],
                                                  varnames_dict[var[1]] if var[1] in varnames_dict.keys() else var[1], size=12))
        plt.subplot(2, 3, 1)
        self._seasonal_subplot_month(var[0], dfs[0])
        plt.subplot(2, 3, 2)
        self._seasonal_subplot_monthyr(var[0], dfs[0])
        plt.subplot(2, 3, 3)
        self._seasonal_subplot_yr(var[0], dfs[0])
        plt.subplot(2, 3, 4)
        self._seasonal_subplot_month(var[1], dfs[1])
        plt.subplot(2, 3, 5)
        self._seasonal_subplot_monthyr(var[1], dfs[1])
        plt.subplot(2, 3, 6)
        self._seasonal_subplot_yr(var[1], dfs[1])

        # save fig to PDF-object and close/delete figure.
        fig.savefig(pp, format='pdf', orientation='landscape', papertype='a4')
        plt.close(fig)
        del fig

    def _seasonal_subplot_month(self, var, df):
        """
        This sub-function takes all HH of a certain month (independent of year) and averages them. Furthermore, it
        performs an ANOVA- and a Kruskal-Wallis-test
        :param var: string with name of variable
        :param df: pandas-Dataframe with data of variable
        :return: pointplot
        """

        ax = sns.pointplot(x="month_name", y=var, data=df, color=sns.color_palette('viridis', 2)[0])
        if var.startswith('m') or var.startswith('mx'):
            yl = ('kg/kWh/MJ/m3')
        else:
            yl = 'CHF'
        plt.ylabel(yl)
        plt.xlabel('')
        ax.set_ylim(0, )
        plt.xticks(rotation=45)

        # Split the dataset in monthly subsets
        mths = []
        for mth in df['month_name'].unique():
            mths.append(df.loc[df['month_name'] == mth, ['month_name', var]])

        # Conduct a ANOVA- and a Kruskal-Wallis-Test with the monthly subsets
        F_val, pav_val = stats.f_oneway(*[mth[var] for mth in mths])
        H_val, pkru_val = stats.kruskal(*[mth[var] for mth in mths])

        # Plot the test results in the subplot
        at = AnchoredText("ANOVA: F = {}, p = {}\nKruskal: H = {}, p = {}".format(
            round(F_val, 2), round(pav_val, 3), round(H_val, 2), round(pkru_val, 3)), loc=8, frameon=False)  # prop=dict(size=10)
        ax.add_artist(at)

    def _seasonal_subplot_monthyr(self, var, df):
        """
        This sub-function takes all HH of a certain month and year and averages them to form a pointplot for each year.
        :param var: string with name of variable
        :param df: pandas-Dataframe with data of variable
        :return: pointplot
        """
        ax = sns.pointplot(x="month_name", y=var, data=df, hue='year', palette=sns.color_palette("viridis", 3),
                           linestyles=['-', '--', ':'])
        plt.ylabel('')
        plt.xlabel('')
        ax.set_ylim(0, )
        plt.xticks(rotation=45)

    def _seasonal_subplot_yr(self, var, df):
        """
        This sub-function takes all HH of a certain year averages them.
        :param var: string with name of variable
        :param df: pandas-Dataframe with data of variable
        :return: pointplot
        """
        ax = sns.pointplot(x="year", y=var, data=df, color=sns.color_palette('viridis', 2)[-1])
        plt.ylabel('')
        plt.xlabel('')
        ax.set_ylim(0, )
        plt.xticks(rotation=45)


def get_data_for_seasonal_graph(conn, var, not_orig_table=False):
    """
    This function retrieves the data for a specific attribute for which seasonal graphs shall be created from the PGDB
    and converts it to a Pandas-Dataframe. If not further specified with the kwarg not_orig_table, it retrieves the
    data - if possible - from original-HABE-tables. However, other tables can be chosen by passing the table-name as
    string in not_orig_table.
    :param conn: connection to the PG-DB
    :param var: name of the attribute for which data needs to be retrieved
    :param not_orig_table: (optional) string with name of table from which data shall be retrieved (if data shall not come
    from original HABE-data)
    :return: Pandas-DataFrame with data for the seasonal graphs
    """

    cur = conn.cursor(cursor_factory=pge.RealDictCursor)

    # If no other specific PG-table is specified, data is - if possible - retrieved from original HABE-tables
    if not not_orig_table:
        # The next if-statements help to create a query with the correct tables where the data shall be retrieved from (if
        # possible always from original_data, but for computed attributes (mx) from habe_hh_prepared). Furthermore it makes a
        # LEFT-JOIN with the month-table.
        if var.startswith('a'):
            query = """
            SELECT year, month_no, month_name, {} FROM original_data.habe_ausgaben ausg LEFT JOIN working_tables.habe_hh_month mth
            ON ausg.haushaltid=mth.haushaltid
            """.format(var)
        elif var.startswith('mx'):
            query = """
            SELECT year, month_no, month_name, {} FROM working_tables.habe_hh_prepared mx LEFT JOIN working_tables.habe_hh_month mth
             ON mx.haushaltid=mth.haushaltid
            """.format(var)
        elif var.startswith('m') and not var.startswith('mx'):
            query = """
            SELECT year, month_no, month_name, {} FROM original_data.habe_mengen m LEFT JOIN working_tables.habe_hh_month mth
             ON m.haushaltid=mth.haushaltid
            """.format(var)
        elif var.startswith('e') or var.lower() in ('e_primaereink', 'e_bruttoeink', 'e_verfuegbeink', 'Primaereinkommen08'.lower(),
                                            'Bruttoeinkommen08'.lower(), 'VerfuegbaresEinkommen08'.lower(),
                                            'e_sparbetrag', 'Sparbetrag08'.lower()):
            # Some income-categories need to be translated from the pg-code used in habe_hh_prepared to the pg-code in the original HABE-data
            if var == 'e_primaereink':
                var = 'Primaereinkommen08'
            elif var == 'e_bruttoeink':
                var = 'Bruttoeinkommen08'
            elif var == 'e_verfuegbeink':
                var = 'VerfuegbaresEinkommen08'
            elif var == 'e_sparbetrag':
                var = 'Sparbetrag08'

            query = """
            SELECT year, month_no, month_name, {} FROM original_data.habe_standard std LEFT JOIN working_tables.habe_hh_month mth
            ON std.haushaltid=mth.haushaltid
            """.format(var)
    else:
        # If data shall be retrieved from another tables than the original HABE-tables, then we proceed as follows
        query = """SELECT year, month_no, month_name, {var} FROM {tabl} t LEFT JOIN working_tables.habe_hh_month mth
        ON t.haushaltid=mth.haushaltid
        """.format(var=var, tabl=not_orig_table)

    # retrieve the PG-data
    cur.execute(query)
    sqldata = cur.fetchall()
    cur.close()

    # convert data to a pandas-dataframe and return
    return pd.DataFrame(data=sqldata)


def create_seasonal_figs(savepath, varlist, varnames_dict=None, not_orig_table=False):
    """
    This function creates figures which investigate different HABE-attributes for seasonality. If not further specified
    with the kwarg not_orig_table, it retrieves the data - if possible - from original-HABE-tables. However, other tables
    can be chosen by passing the table-name as string in not_orig_table.
    :param savepath: string with the path-name where the seasonal graph shall be stored
    :param varlist: list of attributes for which seasonal graphs shall be created
    :param varnames_dict: translation dict which translates PG-codes to human readable attribute names (this dict can
    for instance be created with SKLData-class)
    :param not_orig_table: (optional) string with name of table from which data shall be retrieved (if data shall not come
    from original HABE-data)
    :return: stores PDF with seasonal graphs for all attributes in varlist
    """

    # First: connect to the PG-DB
    conn = pg.connect("host=localhost port=5432 dbname=db_andi user=postgres  connect_timeout=2 password='postgres'")

    # Open a PDF-File where the seasonal graphs shall be stored in
    pp = PdfPages(savepath)

    # The next for-loop ensures that on each landscape-A4-page two attributes are included (in case of an odd number,
    # the last attribute is taken twice)
    varschecked = []  # list of attributes for which seasonal graphs have already been created.
    for i, var in enumerate(varlist):
        if var in varschecked:
            continue
        else:
            var_a = var
            try:
                var_b = varlist[i+1]
            except:
                var_b = var_a
            varschecked.append(var_a)
            varschecked.append(var_b)

            # get data for attribute a (upper part of the A4-page) and attribute b (lower part of the A4-page)
            df_a = get_data_for_seasonal_graph(conn, var_a, not_orig_table)
            df_b = get_data_for_seasonal_graph(conn, var_b, not_orig_table)

            # actually create the seasonal figure for attributes a and b
            Seasonal_Fig(pp, (df_a, df_b), [var_a, var_b], varnames_dict, not_orig_table)
    conn.close()
    pp.close()


class SKLData(object):
    """
    This class provides an interface between PGDB and sci-kit-learn. It reads the data from PGDB and converts it to a
    numpy-array. The attributes can be passed either by a list or by an excel-sheet (which however needs to be structured
    similar to "Consumption_Cockpit.xlsx". Furthermore, this class is also able to perform left-joins if data
    needs to be retrieved from several tables. IMPORTANT: the ._conn_-attribute of the class should be deleted after applying the class.
    """
    def __init__(self, conn, sqltable, attributes=None, meta=None, joining=None, excel=None, **kwargs):
        """
        classical init-function for class SKLData
        :param conn: connection to the PGDB
        :param sqltable: string with the name of the PG-table (e.g. 'working_tables.habe_hh_prepared')
        :param attributes: list of attributes for the "main" table (which might be joined with other tables)
        :param meta: string with the name of a meta-attribute (e.g. 'haushaltid') --> attribute which shall not be used for computations
        :param joining: dict with keys 'attributes' (list of lists with attributes), 'joining_tables' (list of names of
        tables to be joined), 'join_id' (list of attributes on which join shall be performed --> same order as 'joining_tables')
        :param excel: dict with keys 'path' (path to the excelfile), 'sheet_with_codes' (string with the name of the sheet
         which contains a list of attributes, e.g. 'Overview'), 'code_column_names' (tuple with the names of the columns
         in the sheet that contain the attributes, e.g. ('pg-code 1', 'pg-code 2') and optionally 'filter_column_names'
         (tuple with names of the columns that filter the attributes:0 for excluding and 1 for including, e.g.
         ('Pattern Recognition1', 'Pattern Recognition 2')
        :param kwargs: -
        """

        # if the excel-dict is passed then the attributes are taken from the specific excel-sheet.
        if excel:
            # Open the workbook:
            wb = xlrd.open_workbook(excel['path'])
            self.ws = wb.sheet_by_name(excel['sheet_with_codes'])
            # Read the workbook and get the codes:
            self.codes_ = self._get_codes(excel)
            del wb, self.ws

        else:
            # In case, no excel-dict is passed, the translation dict (which translates the PG-codes into human readable
            # names) becomes just a fake translation dict
            self.varnames_ = {}
            for a in attributes:
                self.varnames_.update({a: a})

        # The next step is the core step, because it actually retrieves data from the PGDB:
        self.data_ = self._get_data(conn, sqltable, joining, attributes)

        # Unfortunately, we need to store the DB-connection as an attribute since it will be reused in the child-class
        # SKLData_NK --> however, after applying the SKLData-class (and or the SKLData_NK-class) it is recommended to
        # delete this attribute manually.
        self._conn_ = conn

        # If we have a meta-attribute (one which shall not be part of the computations), then we store the name of it
        # in a separate attribute and we exclude the respective vector from the .data_-numpy array and store this vector
        # in a separate numpy-vector .meta_
        if meta:
            self.meta_name_ = meta
            self.meta_ = self._get_meta()

    def _get_codes(self, excel):
        """
        This function reads and filters the attributes which shall be read from the PGDB. Furthermore it creates a dict
        which is able to translate the PG-codes into human-readable names.
        :param excel: see "init"-function
        :return: a flat list of all attributes which shall be read from the PGDB.
        """

        codes = []
        # Go through the column-names which indicate the columns containing the attributes:
        for cn in excel['code_column_names']:
            # Go through all rows and columns to find the column containing the attributes:
            for r in range(self.ws.nrows):
                for c in range(self.ws.ncols):
                    if self.ws.cell(r, c).value == cn:
                        # As soon as the column is found, the attributes are added to the "attributes-container" (list of lists)
                        codes.append(self.ws.col_values(c)[r + 1:])
                        rowstart = r  # memorize at which row the list of attributes starts (might be overwritten, while
                                      # looping, but we assume that the different columns containing attributes are parallel)
                        break
                # The following is necessary to break also the outer loop (r)
                else:
                    continue
                break

        # Create a translation dict for the codes of attributes into names
        self.varnames_ = {}
        # Go through the list of lists for the attributes
        for code in codes:
            #Go through all attributes and update the translation dict (ASSUMPTION: names of the attributes are given in the 2nd column)
            for i, var in enumerate(code):
                ro = i + rowstart + 1
                if var not in ['', 0.0]:
                    self.varnames_.update({var: self.ws.cell(ro, 1).value})

        # In a next step, we try to apply a filter of the attributes (0=exclude, 1=include), but since this key in the
        # excel-dict is optional, we need a try-statement.
        try:
            # IMPORTANT: we need aas many entries of 'filter_column_names' as there are 'code_column_names'
            for j, fi in enumerate(excel['filter_column_names']):
                # Find the column with 0/1-entries which represent the filter (ASSUMPTION: are in parallel to the attributes-columns)
                filt_col = self.ws.row_values(rowstart).index(fi)
                # Get the 0/1s and memorize the indices of the rows where the filter is set to 1:
                filt = self.ws.col_values(filt_col)[rowstart + 1:]
                indices = [i for i, x in enumerate(filt) if x not in ('', -1, 0)]
                # Collapse the codes to the filtered attributes
                codes[j] = [codes[j][i] for i in indices]
        except:
            pass

        # Convert the list of lists to a flat list:
        allcodes = []
        for code in codes:
            allcodes += code

        # Just for cleaning up: exclude some "empty" code-cells
        allcodes = [c for c in allcodes if c != 0.0 and c != '']

        return allcodes

    def _get_data(self, conn, sqltable, joining, attributes):
        """
        This function actually retrieves the data from the PG-DB and converts it to a numpy-array which might be used by
        sci-kit-learn.
        :param conn: connection to the DB
        :param sqltable: name of the PGDB-table (see init-function)
        :param joining: dict of information for joining different tables (see init-function)
        :param attributes: list of attributes of the "main" table (see init-function)
        :return: numpy-array of the data retrieved from PGDB
        """

        # In a first step, we check if we already have a codes-attribute --> this is the case if an excel-dict was
        # passed to the class. Otherwise, we read the list of attributes to the .attributes_-attribute.
        try:
            self.attributes_ = self.codes_
            del self.codes_
        except:
            self.attributes_ = attributes

        # In case we want to join the "main" table with other tables, we need to add the other tables' attribute to the existing attributes-list
        if joining:
            for attribs in joining['attributes']:
                self.attributes_ += attribs

        # Start creating the SQL-query:
        query = """
        SELECT {} FROM {}
        """.format((", ".join(self.attributes_)).lower(), sqltable)

        # In case of joining different tables, we go through the joining-key 'joining_tables' which contain the names of tables to be joined
        # and we then add the attribute which is the basis for the join (to be found in the 'join_id'-key --> list with the same order as
        # 'joining_tables').
        if joining:
            for i, tbl in enumerate(joining['joining_tables']):
                query += " LEFT JOIN {} ON {}".format(tbl, "{t}.{id} = {s}.{id}".format(t=tbl,
                                                                                        id=joining['join_id'][i],
                                                                                        s=sqltable))
        # core of the whole class: retrieve data from PGDB
        cur = conn.cursor()
        cur.execute(query)
        sqldata = cur.fetchall()
        cur.close()

        # convert to numpy-array and then return data.
        return np.array(sqldata)

    def _get_meta(self):
        """
        This function separates a meta-column (attributes which shall not enter a computation) from the other data and
        returns a numpy-vector with the meta-data (often just "haushaltid").
        :return: numpy-vector with meta-data
        """

        # Find the column in which the meta-data is stored:
        i = self.attributes_.index(self.meta_name_)
        # "deep copy" the meta-data:
        metavalues = np.copy(self.data_[:, i])
        # delete meta-data from data-matrix and from attributes-list
        self.data_ = np.delete(self.data_, i, 1)
        del self.attributes_[i]
        return metavalues

    def standardize_data(self):
        """
        Standardizes the data and saves the used scaler as attribute (e.g. for reverse standardization later on)
        """
        self.data_scaled_ = preprocessing.scale(self.data_)
        self.scaler_ = preprocessing.StandardScaler().fit(self.data_)

    def reverse_standardization(self, vectors, cols):
        """
        Converts standardized data back to non-standardized data.
        :param vectors: (numpy array) pass several vectors which shall be de-standardized.
        :param cols: (list of integers) Indices of columns to which the "vectors" correspond in the .data_-matrix.
         IMPORTANT: list of integers need to have the same order as the "vectors".
        :return: (numpy array) de-standardized vectors in the same order as "vectors".
        """

        # First we creata a matrix of zeros. The number of rows correspond with the number of rows of the vectors
        # which shall be de-standardized. The number of columns correspond to the number of columns in the .data_-matrix
        # minus the number of vectors which shall be de-standardized (in the end we need a matrix with the same number
        # of columns as .data_-matrix --> we will insert the passed vectors afterwards).
        X = np.zeros([len(vectors[:, 0]), len(self.data_[0]) - len(cols)])

        # Insert the vectors which shall be de-standardized
        for i, col in enumerate(cols):
            try:  # Obviously in latest numpy, this try-statement is not necessary anymore
                X = np.insert(X, col, vectors[:, i], axis=1)
            except:  # Probably never used
                X = np.append(X, vectors[:, i], axis=1)

        # Core: inversion of scaling:
        invX = self.scaler_.inverse_transform(X)

        # Prepare result-container:
        invVectors = np.zeros([len(vectors[:, 0]), len(cols)])

        # Extract de-standardized vectors and store in result-container
        for i, col in enumerate(cols):
            invVectors[:, i] = invX[:, col]

        return invVectors


class SKLData_NK(SKLData):
    """
    This is a child-class of SKLData and adds functionalities to facilitate the imputation of missing dwelling extra cost
    information.
    """
    def __init__(self, nk_ky, *args, nk_shr=False, amount_as_target=False, add_month=False, **kwargs):
        """
        Init-function which starts with the init-function of SKLData and then adds new initializations.
        :param nk_ky: string which indicates the category of dwelling costs which needs to be treated string must be
        either 'NK', 'K', 'AW', 'W', 'U', 'El', 'EnBr' or 'EnZentr' --> however, 'U' is never needed and if 'EnBr'
        is indicated then the sum of 'EnBr' and 'EnZentr' will be chosen as a target
        :param args: to be passed to parent-class SKLData
        :param nk_shr: True/False if shares instead of absolute values should be targeted by the regression techniques
        :param amount_as_target: True/False to choose if quantities (e.g. kWh) instead of expenditures shall be targeted
        :param add_month: True/False to indicate if the survey month shall be included
        :param kwargs: -
        """

        # Call SKLData-init-function:
        super(SKLData_NK, self).__init__(*args, **kwargs)

        # Simple check if kwargs are set reasonably
        if nk_shr and amount_as_target:
            print("WARNING: computing shares of amounts doesn't make sense!")

        # Dicts to translate the nk_ky-shortcut-string into codes used in the pg-tables
        nkcats_dict = {'NK': 'a571201', 'K': 'a571202', 'AW': 'a571203', 'W': 'a571204',
                       'U': 'a571205', 'El': 'a571301', 'EnBr': 'a571302', 'EnZentr': 'a571303'}
        nkamounts_dict = {'K': 'mx571202', 'AW': 'mx571203', 'W': 'mx571204',
                          'El': 'mx571301', 'EnBr': 'mx571302', 'EnZentr': 'mx571303'}

        # In a first step, we need to know the indices of the extra costs in order to form the total of extra dwelling costs.
        # This is not always easy since for instance 'Kehricht' shall not be included here
        nk_indices = []
        ink = False  # to find out if 'Kehricht'-Key was already found
        for i, ky in enumerate(nkcats_dict.keys()):
            if ky == 'K':
                ink = True
                continue  # Kehricht shall not be added to dwelling costs

            # The next if-loop looks for the index of the target variable within nk_indices
            if ky == nk_ky and not amount_as_target:
                if ink:
                    inky = i - 1
                else:
                    inky = i
            nk_indices.append(self.attributes_.index(nkcats_dict[ky]))

        # Compute total dwelling extra costs (comprises all categories of nkcats_dict except for 'K')
        nk_tot = self.data_[:, nk_indices].sum(axis=1)

        # If 'Kehricht' is not the target or if quantities are investigated instead of expenditures, then we add the waste expenditures back to nk_indices.
        # The amount_as_target-condition is necessary if 'K' is estimated via amounts instead of expenditures.
        if (not nk_ky == 'K') or amount_as_target:
            nk_indices.append(self.attributes_.index(nkcats_dict['K']))

        # But we eliminate the target variable from the nk_indices except if the target is a quantity or the total dwelling extra costs
        if not amount_as_target and not nk_ky == 'NK':
            del nk_indices[inky]

        # Since we will not distinguish between "Brennstoffe" and "Zentralheizung", we sum up these two categories --> only EnBr needs to be entered for nk_key
        if nk_ky == 'EnBr':
            heat_tot = self.data_[:, [self.attributes_.index(nkcats_dict['EnBr']), self.attributes_.index(nkcats_dict['EnZentr'])]].sum(axis=1)
            # We replace the 'EnBr' data with the new sum of "Brennstoffe" and "Zentralheizung"
            self.data_[:, self.attributes_.index(nkcats_dict[nk_ky])] = heat_tot

        # Now we delete all expenditure categories which are part of the total dwelling extra costs plus Kehricht
        # and adjust also attributes-list accordingly
        # Important: because we only want to impute "NK pauschal" for HH which have missing dwelling categories, we need
        # to store all expenditure categories for the moment. These will be deleted in the nk_data_preparer-function.
        if not nk_ky == 'NK':
            self.data_ = np.delete(self.data_, nk_indices, 1)
            self.attributes_ = [x for i, x in enumerate(self.attributes_) if i not in nk_indices]

        # Adjust attributes-list and translation dict for the total dwelling extra costs which are computed above and will be inserted below
        self.attributes_.append('NK_tot_computed')
        self.varnames_.update({'NK_tot_computed': 'NK_tot_computed'})

        # For convenience we store the column index of the target variable in the attribute .i_target_
        if not amount_as_target:
            if not nk_ky == 'NK':
                self.i_target_ = self.attributes_.index(nkcats_dict[nk_ky])
            else:
                self.i_target_ = self.attributes_.index('NK_tot_computed')
        else:
            self.i_target_ = self.attributes_.index(nkamounts_dict[nk_ky])

        # And finally we insert total dwelling costs at the end for using it as a new predictor
        self.data_ = np.append(self.data_, nk_tot.reshape([len(nk_tot), 1]), axis=1)

        # The next if-loop transforms the target variable into share of total dwelling extra costs
        if nk_shr:
            self._transform_target_to_shares(nk_ky)

        # The next if-loop adds the month in which the households were surveyed as information (e.g. for random forest)
        if add_month:
            self._add_month_to_hhs(self._conn_)

        self._conn_.close()
        del self._conn_

    def _transform_target_to_shares(self, nk_ky):
        """
        This function transforms the target variable into shares of the dwelling extra costs
        :param nk_ky: see init-function
        """
        shares = self.data_[:, self.i_target_] / self.data_[:, -1]  # computing the shares (total dwelling extra costs were appended as last column)
        self.data_ = np.delete(self.data_, self.i_target_, 1)  # delete the target variable which are absolute amounts
        self.data_ = np.append(self.data_, shares.reshape([len(shares), 1]), axis=1)  # append the newly calculated shares as last column

        # Adjust attributes-list and translation-dict:
        del self.attributes_[self.i_target_]
        self.attributes_.append('{}_share_computed'.format(nk_ky))
        self.varnames_.update({'{}_share_computed'.format(nk_ky): '{}_share_computed'.format(nk_ky)})

        # Adjust index of target-variable (now shares)
        self.i_target_ = self.attributes_.index('{}_share_computed'.format(nk_ky))

    def _add_month_to_hhs(self, conn):
        """
        This function adds the survey month to the predictors, e.g. to be used in Random Forests
        :param conn: connection to the PGDB
        """

        # Retrieve the survey month data from the PGDB:
        query = """
        SELECT haushaltid, month_no FROM working_tables.habe_hh_month
        """
        cur = conn.cursor()
        cur.execute(query)
        sqldata = cur.fetchall()
        cur.close()

        # Putting the survey months into the correct order
        months = np.zeros(len(sqldata))
        for hh in sqldata:
            hhind = np.where(self.meta_ == hh[0])  # find the index of the respective household
            months[hhind] = hh[1]  # insert the month at the correct place

        # Append the month-information as last column to the data-matrix and adjust attributes-list and translation-dict accordingly
        self.data_ = np.append(self.data_, months.reshape([len(months), 1]), axis=1)
        self.attributes_.append('survey month')
        self.varnames_.update({'survey month': 'survey month'})


def nk_data_preparer(title, data_nk, nk_ky):
    """
    This data prepares the data for applying the regression techniques. E.g. it splits data into target and predictor data
    as well as into renters and owners, etc.
    :param title: string with a title name of the data
    :param data_nk: actual data (SKLData_NK-class)
    :param nk_ky: string indicating the target variable ('NK', 'K', 'AW', 'W', 'El', 'EnBr')
    :return: dict of split data.
    """

    def data_splitter(data_nk, nk_ky):
        """
        Subfunction of nk_data_preparer. Splits data into target and training data
        :param data_nk:
        :param nk_ky: see main function
        :return: split data
        """

        # If not "Pauschale Nebenkosten" are targeted, it is quite easy to split the data:
        if not nk_ky == 'NK':
            data_nk_target = copy.deepcopy(data_nk)  # Copy the whole dataset

            # Then replace .meta_ and .data_ with rows for which the target variable is 0
            data_nk_target.meta_ = np.copy(data_nk.meta_[data_nk.data_[:, data_nk.i_target_] == 0])
            data_nk_target.data_ = np.copy(data_nk.data_[data_nk.data_[:, data_nk.i_target_] == 0, :])

            # Now prepare the training dataset
            data_nk_training = copy.deepcopy(data_nk)

            # Replace .meta_ and .data_ with rows for which the target variable is > 0
            data_nk_training.meta_ = np.copy(data_nk.meta_[data_nk.data_[:, data_nk.i_target_] > 0])
            data_nk_training.data_ = np.copy(data_nk.data_[data_nk.data_[:, data_nk.i_target_] > 0, :])

        # In case of modelling "Pauschale Nebenkosten", things become more difficult, because we only want to impute
        # data for households having missing dwelling categories (so for the target data we look for households with
        # 'NK pauschal' = 0, but which have at the same time one of the other dwelling categories = 0).
        else:
            # Translation dict to convert target variable shortcut to codes used in PGDB
            nkcats_dict = {'NK': 'a571201', 'AW': 'a571203', 'W': 'a571204',
                           'U': 'a571205', 'El': 'a571301', 'EnBr': 'a571302', 'EnZentr': 'a571303'}

            # Computation of the total heating costs (since we consider "Brennstoffe" and "Zentralheizung" together as one variable)
            en_tot = data_nk.data_[:, [data_nk.attributes_.index(nkcats_dict['EnBr']), data_nk.attributes_.index(nkcats_dict['EnZentr'])]].sum(axis=1)

            # Add the total heating costs to the data and adjust attributes-list and translation-dict
            data_nk.data_ = np.append(data_nk.data_, en_tot.reshape([len(en_tot), 1]), axis=1)
            data_nk.attributes_.append('En_tot_computed')
            data_nk.varnames_.update({'En_tot_computed': 'En_tot_computed'})

            # We then define two indices-containers, one for the missing dwelling costs and the other for all dwelling categories
            nk_indices = []  # indices for checking if HH has categories which need to be imputed
            nk_indices2 = []  # indices of all dwelling cost categories
            for ky in nkcats_dict.keys():
                if ky in ['AW', 'W', 'El']:
                    nk_indices.append(data_nk.attributes_.index(nkcats_dict[ky]))
                nk_indices2.append(data_nk.attributes_.index(nkcats_dict[ky]))
            nk_indices.append(data_nk.attributes_.index('En_tot_computed'))
            nk_indices2.append(data_nk.attributes_.index('En_tot_computed'))
            nk_indices2.append(data_nk.attributes_.index('a571202'))  # since nk_indices2 will be used to delete columns which shall not enter the computations, we also need to add "Kehricht" manually

            # Now follows the split into target and training data:
            data_nk_target = copy.deepcopy(data_nk)  # Copy the whole dataset
            # Then replace .meta_ and .data_ with rows for which 'NK pauschal' is zero and for the same time one of the other dwelling costs needs to be imputed as well
            data_nk_target.meta_ = np.copy(data_nk.meta_[data_nk.data_[:, data_nk.attributes_.index(nkcats_dict['NK'])] == 0][np.any(data_nk.data_[data_nk.data_[:,  data_nk.attributes_.index(nkcats_dict['NK'])] == 0][:, nk_indices] == 0, axis=1)])
            data_nk_target.data_ = np.copy(data_nk.data_[data_nk.data_[:, data_nk.attributes_.index(nkcats_dict['NK'])] == 0, :][np.any(data_nk.data_[data_nk.data_[:,  data_nk.attributes_.index(nkcats_dict['NK'])] == 0, :][:, nk_indices] == 0, axis=1), :])

            # Delete columns of dwelling categories and adjust attributes-list accordingly (similar as in SKLData_NK if target is not 'NK')
            data_nk_target.data_ = np.delete(data_nk_target.data_, nk_indices2, 1)
            data_nk_target.attributes_ = [x for i, x in enumerate(data_nk_target.attributes_) if i not in nk_indices2]
            # Save the index of the target variable
            data_nk_target.i_target_ = data_nk_target.attributes_.index('NK_tot_computed')

            # Now prepare the training dataset
            data_nk_training = copy.deepcopy(data_nk)

            # In a first step, we go through all households and exclude the indices which are not already in the target dataset
            indices = [i for i, e in enumerate(data_nk.meta_) if e not in data_nk_target.meta_]

            # Reduce .meta_ and .data_ to the households which are not already in the target dataset
            data_nk_training.meta_ = np.copy(data_nk.meta_[indices])
            data_nk_training.data_ = np.copy(data_nk.data_[indices, :])

            # Delete columns of dwelling categories and adjust attributes-list accordingly (similar as in SKLData_NK if target is not 'NK')
            data_nk_training.data_ = np.delete(data_nk_training.data_, nk_indices2, 1)
            data_nk_training.attributes_ = [x for i, x in enumerate(data_nk_training.attributes_) if i not in nk_indices2]
            # Save the index of the target variable
            data_nk_training.i_target_ = data_nk_training.attributes_.index('NK_tot_computed')

            # Reset data_nk (while debugging: found out that this is necessary)
            data_nk.data_ = np.delete(data_nk.data_, data_nk.attributes_.index('En_tot_computed'), 1)
            data_nk.attributes_ = [x for i, x in enumerate(data_nk.attributes_) if i != data_nk.attributes_.index('En_tot_computed')]
            data_nk.i_target_ = data_nk.attributes_.index('NK_tot_computed')
        return data_nk_training, data_nk_target

    # In a first step, we split the whole dataset into training and target data
    data_nk_training, data_nk_target = data_splitter(data_nk, nk_ky)

    # Standardize the training and the target dataset respectively
    data_nk_training.standardize_data()
    # Good modelling practice: use the scaler of the training dataset to standardize the target dataset (instead of standardizing the target dataset by its own)
    data_nk_target.data_scaled_ = np.copy(data_nk_training.scaler_.transform(data_nk_target.data_))

    # So far, we have used the whole dataset. But since we assume that for dwelling costs it might be important if a
    # household is renting or owning its apartment, we split the whole dataset into renters and owners:
    irent = data_nk.attributes_.index('char_renter')

    # Get renter-households
    data_nk_renters = copy.deepcopy(data_nk)
    data_nk_renters.meta_ = np.copy(data_nk.meta_[data_nk.data_[:, irent] == 1])
    data_nk_renters.data_ = np.copy(data_nk.data_[data_nk.data_[:, irent] == 1, :])

    # Split renter-households into a training and a target dataset
    data_nk_renters_training, data_nk_renters_target = data_splitter(data_nk_renters, nk_ky)

    # Standardize the renter training and renter target dataset
    data_nk_renters_training.standardize_data()
    data_nk_renters_target.data_scaled_ = np.copy(
        data_nk_renters_training.scaler_.transform(data_nk_renters_target.data_))

    # Get owner-households
    data_nk_owners = copy.deepcopy(data_nk)
    data_nk_owners.meta_ = np.copy(data_nk.meta_[data_nk.data_[:, irent] == 0])
    data_nk_owners.data_ = np.copy(data_nk.data_[data_nk.data_[:, irent] == 0, :])

    # Split owner-households into a training and a target dataset
    data_nk_owners_training, data_nk_owners_target = data_splitter(data_nk_owners, nk_ky)

    # Standardize the owner training and owner target dataset
    data_nk_owners_training.standardize_data()
    data_nk_owners_target.data_scaled_ = np.copy(
        data_nk_owners_training.scaler_.transform(data_nk_owners_target.data_))

    # Put all datasets into a dict in order to return
    return {title: [data_nk_training, data_nk_target], '{}_renters'.format(title): [data_nk_renters_training, data_nk_renters_target],
            '{}_owners'.format(title): [data_nk_owners_training, data_nk_owners_target]}


def nk_model_basic_diagnostic_plot(nk_model_class, savepath, savename):
    """
    This function creates basic diagnostic plots for the regression models which are applied for imputing missing information
    on dwelling extra costs. Three graphs are created: 1. scatter plot with predicted training data vs. observed training data
    (goal would be a straight diagonal line), 2. scatter plot with predicted training data vs. residuals and 3. two boxplots
    comparing the computed target-set vs. the observed training-set (do not need to be the same, but just to check if
    reasonable results).
    :param nk_model_class: Lasso_NK, KNN_NK or RF_NK-class which carry all the information from fitting a regression model to
     impute missing dwelling extra costs.
    :param savepath: (string) path to the folder where the graph shall be stored
    :param savename: (string) name for the diagram (without file-extension)
    """

    # Create a DataFrame with the predicted training set and the observed training set
    data = pd.DataFrame(np.hstack([nk_model_class.y_pred_pred_destd_, nk_model_class.y_pred_destd_]), columns=['x', 'y'])

    # Compute the residuals
    data['r'] = data['y'] - data['x']

    # For the boxplots: create two dataframes, each of them with a "pseudo"-column which shall help to plot the boxplots in the same figure
    df_pred = pd.DataFrame(data={'Observed (Training-Set)': [x[0] for x in list(nk_model_class.y_pred_destd_)], 'pseudo': np.nan})
    df_targ = pd.DataFrame(data={'Predicted Target (Target-Set)': [x[0] for x in list(nk_model_class.y_targ_destd_)], 'pseudo': np.nan})

    # Initialize figure settings
    fig_width_cm = 21  # A4 page
    fig_height_cm = 29.7
    inches_per_cm = 1 / 2.58  # Convert cm to inches
    fig_width = fig_width_cm * inches_per_cm  # width in inches
    fig_height = fig_height_cm * inches_per_cm  # height in inches
    fig_size = [fig_height,
                fig_width]  # height and width are in the order needed for landscape
    paper_rc = {'lines.linewidth': 1, 'lines.markersize': 6}  # parameters to control the linewidth and markersize
    sns.set_context("talk", rc=paper_rc)
    sns.set(font_scale=0.97)

    # Core: create figures:
    fig = plt.figure(figsize=fig_size)
    gridspec.GridSpec(2, 2)
    plt.suptitle(savename)

    # Scatter plot: observed training set vs. predicted training set
    plt.subplot2grid((2, 2), (0, 0), colspan=1, rowspan=1)
    sns.regplot(x='x', y='y', data=data, scatter=True, fit_reg=False, color=sns.color_palette('viridis', 2)[0],
                scatter_kws={'alpha': 0.5})
    plt.xlabel('Predicted (Training-Set)')
    plt.ylabel('Observed (Training-Set)')
    # Diagonal line:
    plt.plot([min(min(data['x']), min(data['y'])), max(max(data['x']), max(data['y']))],
             [min(min(data['x']), min(data['y'])), max(max(data['x']), max(data['y']))], 'k-')

    # Scatter plot: predicted training set vs. residuals
    plt.subplot2grid((2, 2), (1, 0), colspan=1, rowspan=1)
    sns.regplot(x='x', y='r', data=data, scatter=True, fit_reg=False, color=sns.color_palette('viridis', 2)[-1],
                scatter_kws={'alpha': 0.5})
    plt.xlabel('Predicted (Training-Set)')
    plt.ylabel('Residuals')
    # Horizontal line at y=0:
    plt.plot([min(data['x']), max(data['x'])], [0, 0], 'k-')

    # Boxplots comparing observed training set vs. predicted target set:
    plt.subplot2grid((2, 2), (0, 1), colspan=1, rowspan=2)
    sns.boxplot(data=df_pred, order=[0, 1], palette='Set3')
    sns.boxplot(data=df_targ, order=[1, 0], palette='Set3')
    plt.xticks([0, 1], ['Observed (Training-Set)', 'Predicted (Target-Set)'])

    # Save figure in two file formats
    fig.savefig(os.path.join(savepath, "{}.pdf".format(savename)), format='pdf', orientation='landscape', papertype='a4')
    fig.savefig(os.path.join(savepath, "{}.png".format(savename)), format='png')
    plt.close(fig)
    del fig


class Lasso_NK(object):
    """
    This class performs LASSO-Regression to impute missing dwelling cost data. Optionally, the LASSO-Regression can
    be conducted with the PC-components resulting from a pre-PCA. The hyperparameters are tuned with 10-fold
    cross-validation. Optionally, hyperparameters can also be passed to this class. IMPORTANT: LASSO is always computed
    with standardized data.
    NOTE: Good modelling practice would be to implement scaling and PCA within cross-validation. However, the sklearn-pipeline
    does not allow for transforming y at the moment. The workaround would have been too cumbersome. Since this good
    modelling practice does only influence the performance metrics but not the fitting (and because internet research
    suggests that the difference is minimum), we decided to keep scaling and PCA outside cross-validation.
    """
    def __init__(self, training, target, meta, title, pre_pca=False, modelparams=None):
        """
        Init-function for the LASSO-Regression class designed for imputing missing dwelling costs.
        :param training: Training dataset (SKLData_NK-class, result of nk_data_preparer)
        :param target: Target dataset (SKLData_NK-class, result of nk_data_preparer)
        :param meta: Pandas-Dataframe to store performance metrics of the LASSO-Regression. Expected indices are:
        ['R2', 'MSE', 'alpha', 'pred_mean', 'pred_min', 'pred_max', 'orig_mean', 'orig_min', 'orig_max']
        :param title: string with the column-name for storing the performance metrics of the LASSO-Regression in the
        Pandas-Dataframe (see parameter meta)
        :param pre_pca: (optional) True/False to choose if a PCA should be performed to pre-condition the dataset
        :param modelparams: (optional) Dict with key 'alpha' in order to compute a LASSO-Regression with pre-set parameters
        """

        # Generate a mask to separate the target variable from the predictors
        mask = np.ones(len(training.data_scaled_[0]), dtype=bool)
        mask[training.i_target_] = False
        X_pred = training.data_scaled_[:, mask]
        y_pred = training.data_scaled_[:, training.i_target_]

        if pre_pca:
            # Optionally, a PCA can be performed to pre-condition the dataset. The PCA will decorrelate the predictors
            # and thus facilitate the work of the LASSO-Regression
            pca = decomposition.PCA(svd_solver='full')  # initializing PCA
            X_pred = pca.fit_transform(X_pred)  # transform the predictors

        # In the next step, the hyperparameters are tuned if no pre-set parameters are passed
        if not modelparams:
            # The following while-loop shall prevent to limit the fitting by a maximum number of iterations
            maxiter = 1500
            actual_iter = 15000
            while maxiter <= actual_iter:
                maxiter *= 10
                # LASSO-Parameters: cv=10 chosen according to "Applied Predictive Modelling"-Book, normalize is not
                # needed since already standardized, n_alphas=1000 is subjectively chosen, fit_intercept=False -->
                # checks were run and show no big difference if True or False. However, it decreases the chance to get
                # negative values and it is also not necessary if data is already centered (standardized).
                lassoCV = linear_model.LassoCV(fit_intercept=False, normalize=False, n_jobs=-1, n_alphas=1000,
                                               max_iter=maxiter, cv=10).fit(X_pred, y_pred)

                actual_iter = lassoCV.n_iter_

            # Store the best-performing alpha
            alfa = lassoCV.alpha_
            self.lassoCV_ = lassoCV
        else:
            # In case we pass a pre-set parameter, we use of course this value for alpha
            alfa = modelparams['alpha']

        # After tuning or setting alpha, we can definitely fit the LASSO-Regression
        lasso = linear_model.Lasso(alpha=alfa, fit_intercept=False, normalize=False)

        # Core: fit LASSO and compute performance metrics (here R2, while MSE in case a pre-set parameter for alpha is passed)
        maxiter  = 1500
        actual_iter = 15000
        while maxiter <= actual_iter:  # Again we enter a while-loop to prevent the maximum iteration parameter to stop the computations
            maxiter *= 10
            lasso.max_iter = maxiter

            r2scores = model_selection.cross_val_score(lasso, X_pred, y_pred, cv=10, scoring='r2')

            if modelparams:  # If no parameters are passed, MSE is already computed while tuning alpha
                msescores = model_selection.cross_val_score(lasso, X_pred, y_pred, cv=10, scoring='neg_mean_squared_error')

            # Actual fit:
            lasso.fit(X_pred, y_pred)

            actual_iter = lasso.n_iter_

        self.lasso_ = lasso

        if not pre_pca:
            # Prepare target dataset by removing target column
            X_targ = target.data_scaled_[:, mask]
        else:
            # If dataset is pre-conditioned with PCA, then we need to also transform the target dataset accordingly
            X_targ = pca.transform(target.data_scaled_[:, mask])

        # Core: predict target variable of target dataset
        y_targ = lasso.predict(X_targ)

        # For computing residuals and producing basic diagnostic plots, we als need the predictions of the training set
        y_pred_pred = lasso.predict(X_pred)

        # Convert predicted target vector back to non-standardized values. For this we need the training dataset since
        # we used the training-sclaer to standardize.
        y_targ_destd = training.reverse_standardization(y_targ.reshape([len(y_targ), 1]), [training.i_target_])
        self.y_targ_destd_ = y_targ_destd

        # In order to compare the predicted variables with original values, we also reverse standardization of the predictors target vector
        y_pred_destd = training.reverse_standardization(y_pred.reshape([len(y_pred), 1]), [training.i_target_])
        self.y_pred_destd_ = y_pred_destd

        # In order compute residuals and producing basic diagnostic plots, we also reverse standardization of the predicted predictors target vector
        y_pred_pred_destd = training.reverse_standardization(y_pred_pred.reshape([len(y_pred_pred), 1]), [training.i_target_])
        self.y_pred_pred_destd_ = y_pred_pred_destd

        # Finally we store the performance metrics to the "meta"-pandas-DataFrame
        if not modelparams:
            meta[title] = [r2scores.mean(), lassoCV.mse_path_.mean(axis=1)[np.where(lassoCV.alphas_ == lassoCV.alpha_)][0],
                           alfa, y_targ_destd.mean(), y_targ_destd.min(), y_targ_destd.max(), y_pred_destd.mean(), y_pred_destd.min(), y_pred_destd.max()]
        else:
            meta[title] = [r2scores.mean(), abs(msescores.mean()), alfa, y_targ_destd.mean(), y_targ_destd.min(),
                           y_targ_destd.max(), y_pred_destd.mean(), y_pred_destd.min(), y_pred_destd.max()]


class KNN_NK(object):
    """
    This class performs a K-Nearest-Neighbor-Regression to impute missing dwelling cost data. The hyperparameters are tuned with 10-fold
    cross-validation. Optionally, hyperparameters can also be passed to this class. IMPORTANT: KNN is always computed
    with standardized data.
    NOTE: Good modelling practice would be to implement scaling within cross-validation. However, the sklearn-pipeline
    does not allow for transforming y at the moment. The workaround would have been too cumbersome. Since this good
    modelling practice does only influence the performance metrics but not the fitting (and because internet research
    suggests that the difference is minimum), we decided to keep scaling outside cross-validation.
    """
    def __init__(self, training, target, meta, title, dist='euclidean', modelparams=None):
        """
        Init-function for the KNN-Regression class designed for imputing missing dwelling costs.
        :param training: Training dataset (SKLData_NK-class, result of nk_data_preparer)
        :param target: Target dataset (SKLData_NK-class, result of nk_data_preparer)
        :param meta: Pandas-Dataframe to store performance metrics of the KNN-Regression. Expected indices are:
        ['R2', 'MSE', 'k', 'pred_mean', 'pred_min', 'pred_max', 'orig_mean', 'orig_min', 'orig_max', 'k=1', 'k=3', 'k=5', 'k=10', 'k=20', 'k=50', 'k=100']
        :param title: string with the column-name for storing the performance metrics of the KNN-Regression in the
        Pandas-Dataframe (see parameter meta)
        :param dist: (optional) default is 'euclidean', but also 'mahalanobis' could be chosen to take into account correlation
        of predictors (HOWEVER: not working at the moment)
        :param modelparams: (optional) Dict with key 'k' in order to compute a KNN-Regression with pre-set parameters
        """

        # Generate a mask to separate the target variable from the predictors
        mask = np.ones(len(training.data_scaled_[0]), dtype=bool)
        mask[training.i_target_] = False
        X_pred = training.data_scaled_[:, mask]
        y_pred = training.data_scaled_[:, training.i_target_]

        # In the next step, the hyperparameters are tuned if no pre-set parameters are passed
        if not modelparams:
            if dist == 'euclidean':
                # Working with euclidean-distances just requires simple initialization of a KNN-instance
                knn = neighbors.KNeighborsRegressor(n_jobs=-1)

                # Tuning of parameter k is done by discrete neighbor-numbers. Unfortunately, the errors occur for running
                # GridSearch in parallel, but since knn-instance will be computed already in parallel, this does not matter
                # performance-wise. cv=10 chosen according to "Applied Predictive Modelling"-Book.
                knnCV = model_selection.GridSearchCV(knn, param_grid={'n_neighbors': [1, 3, 5, 10, 20, 50, 100]},
                                                 refit=True, n_jobs=1, pre_dispatch=None, cv=10,
                                                     scoring='neg_mean_squared_error'
                                                     ).fit(X_pred, y_pred)

            else:
                # Computing KNN with Mahalanobis would be great since this also takes into account correlation of predictors
                # however, this is not working at the moment (known issue at SKLEARN).
                dm = neighbors.DistanceMetric.get_metric('mahalanobis', V=np.cov(X_pred))  # initalize distance-metric by computing covariance matrix of dataset

                # Initialize KNN-instance with mahalanobis distance metric. The algorithm was chosen manually in an attempt to
                # get it run.
                knn = neighbors.KNeighborsRegressor(n_jobs=-1, metric=dm, algorithm='ball_tree')

                # Tuning of parameter k is done by discrete neighbor-numbers. Unfortunately, the errors occur for running
                # GridSearch in parallel, but since knn-instance will be computed already in parallel, this does not matter
                # performance-wise. cv=10 chosen according to "Applied Predictive Modelling"-Book.
                knnCV = model_selection.GridSearchCV(knn, param_grid={'n_neighbors': [1, 3, 5, 10, 20, 50, 100]},
                                                 refit=True, n_jobs=1, pre_dispatch=None, cv=10, scoring='neg_mean_squared_error').fit(X_pred, y_pred)

            # Retrieve best-performing parameter:
            k = knnCV.best_params_['n_neighbors']

            # Store MSE of best-performing KNN
            mseCV = abs(knnCV.best_score_)

            self.knnCV_ = knnCV
        else:
            # In case we pass a pre-set parameter, we use of course this value for the no. of neighbors
            k = modelparams['k']

        # After tuning or setting the no. of neighbors, we can definitely fit the KNN-Regression
        if dist =='euclidean':
            knn = neighbors.KNeighborsRegressor(n_jobs=-1, n_neighbors=k)
        else:
            knn = neighbors.KNeighborsRegressor(n_jobs=-1, n_neighbors=k, metric=dm, algorithm='ball_tree')

        # Computing performance metrics (here R2, while MSE was already computed while tuning k)
        r2scores = model_selection.cross_val_score(knn, X_pred, y_pred, cv=10, scoring='r2')

        # Core: fit KNN
        knn.fit(X_pred, y_pred)
        self.knn_ = knn

        # Prepare target dataset by removing target column
        X_targ = target.data_scaled_[:, mask]

        # Core: predict target variable of target dataset
        y_targ = knn.predict(X_targ)

        # For computing residuals and producing basic diagnostic plots, we als need the predictions of the training set
        y_pred_pred = knn.predict(X_pred)

        # Convert predicted target vector back to non-standardized values. For this we need the training dataset since
        # we used the training-sclaer to standardize.
        y_targ_destd = training.reverse_standardization(y_targ.reshape([len(y_targ), 1]), [training.i_target_])
        self.y_targ_destd_ = y_targ_destd

        # In order to compare the predicted variables with original values, we also reverse standardization of the predictors target vector
        y_pred_destd = training.reverse_standardization(y_pred.reshape([len(y_pred), 1]), [training.i_target_])
        self.y_pred_destd_ = y_pred_destd

        # In order compute residuals and producing basic diagnostic plots, we also reverse standardization of the predicted predictors target vector
        y_pred_pred_destd = training.reverse_standardization(y_pred_pred.reshape([len(y_pred_pred), 1]), [training.i_target_])
        self.y_pred_pred_destd_ = y_pred_pred_destd

        # Finally we store the performance metrics to the "meta"-pandas-DataFrame
        if not modelparams:
            metalist = [r2scores.mean(), mseCV, k, y_targ_destd.mean(), y_targ_destd.min(), y_targ_destd.max(), y_pred_destd.mean(), y_pred_destd.min(), y_pred_destd.max()]
            metalist += [abs(m) for m in knnCV.cv_results_['mean_test_score']]
            meta[title] = metalist
        else:
            # If model parameters were already given, then we need to compute MSE seperately (otherwise it is estimated while tuning)
            msescores = model_selection.cross_val_score(knn, X_pred, y_pred, cv=10, scoring='neg_mean_squared_error')
            metalist = [r2scores.mean(), abs(msescores.mean()), k, y_targ_destd.mean(), y_targ_destd.min(), y_targ_destd.max(),
                        y_pred_destd.mean(), y_pred_destd.min(), y_pred_destd.max()]
            meta[title] = metalist

class RF_NK(object):
    """
    This class performs a Random-Forest-Regression to impute missing dwelling cost data. The hyperparameters are tuned with 10-fold
    cross-validation. Optionally, hyperparameters can also be passed to this class.
    NOTE: Good modelling practice would be to implement scaling within cross-validation. However, the sklearn-pipeline
    does not allow for transforming y at the moment. The workaround would have been too cumbersome. Since this good
    modelling practice does only influence the performance metrics but not the fitting (and because internet research
    suggests that the difference is minimum), we decided to keep scaling outside cross-validation.
    """
    def __init__(self, training, target, meta, title, scaling=True, modelparams=None):
        """
        Init-function for the RF-Regression class designed for imputing missing dwelling costs.
        :param training: Training dataset (SKLData_NK-class, result of nk_data_preparer)
        :param target: Target dataset (SKLData_NK-class, result of nk_data_preparer)
        :param meta: Pandas-Dataframe to store performance metrics of the RF-Regression. Expected indices are:
        ['R2', 'MSE', 'r2_oob', 'noTrees', 'pred_mean', 'pred_min', 'pred_max','orig_mean', 'orig_min', 'orig_max',
        '30trees', '100trees', '300trees', '500trees']
        :param title: string with the column-name for storing the performance metrics of the RF-Regression in the
        Pandas-Dataframe (see parameter meta)
        :param scaling: (optional) True/False to indicate if RF shall be fitted with standardized or non-standardized data
        :param modelparams: (optional) Dict with key 'n_estimators' in order to compute a RF-Regression with pre-set parameters
        """

        # Generate a mask to separate the target variable from the predictors
        mask = np.ones(len(training.data_scaled_[0]), dtype=bool)
        mask[training.i_target_] = False

        # Depending on paramter scaling, either standardized data or non-standardized data is loaded
        if scaling:
            X_pred = training.data_scaled_[:, mask]
            y_pred = training.data_scaled_[:, training.i_target_]
        else:
            X_pred = training.data_[:, mask]
            y_pred = training.data_[:, training.i_target_]

        # In the next step, the hyperparameters are tuned if no pre-set parameters are passed
        if not modelparams:
            # RF-Parameters: oob_score=True since with that we have a performance metric for free. n_jobs=-2 is chosen
            # because this class is often used on the big Linux-Server where we should not use all cores.
            rf = ensemble.RandomForestRegressor(n_jobs=-2, oob_score=True)

            # Tuning the number of trees is chosen via GridSearch: Unfortunately, the errors occur for running
            # GridSearch in parallel, but since RF-instance will be computed already in parallel, this does not matter
            # performance-wise. cv=10 chosen according to "Applied Predictive Modelling"-Book. IMPORTANT: we changed the
            # scoring from R2 (default in sklearn) to MSE in order to have it comparable to the other classes (and because
            # it might even be a better indicator)
            rfCV = model_selection.GridSearchCV(rf, param_grid={'n_estimators': [30, 100, 300, 500]},
                                                refit=True, n_jobs=1,
                                                pre_dispatch=None, cv=10, scoring='neg_mean_squared_error'
                                                ).fit(X_pred, y_pred)

            # Retrieve the best RF-model and the best-performing parameter
            rfbest = rfCV.best_estimator_
            notrees = rfCV.best_params_['n_estimators']

            # Store MSE and R2 of out-of-bag of best-performing RF
            mseCV = abs(rfCV.best_score_)  # mean score of CV of best RF
            r2oob = rfbest.oob_score_  # r2 of OOB while bootstrapping the best RF

            self.rfCV_ = rfCV
        else:
            # In case we pass a pre-set parameter, we use of course this value for the no. of trees
            notrees = modelparams['n_estimators']

        # After tuning or setting the no. of trees, we can definitely fit the RF-Regression
        rf = ensemble.RandomForestRegressor(n_jobs=-2, oob_score=True, n_estimators=notrees)

        # Computing performance metrics (here R2, while MSE was already computed while tuning no. of trees)
        r2scores = model_selection.cross_val_score(rf, X_pred, y_pred, cv=10, scoring='r2')

        # Core: fit RF
        rf.fit(X_pred, y_pred)
        self.rf_ = rf

        # Different procedure whether we use scaled/standardized data or not:
        if scaling:
            # Prepare target dataset by removing target column
            X_targ = target.data_scaled_[:, mask]

            # Core: predict target variable of target dataset
            y_targ = rf.predict(X_targ)

            # For computing residuals and producing basic diagnostic plots, we als need the predictions of the training set
            y_pred_pred = rf.predict(X_pred)

            # Convert predicted target vector back to non-standardized values. For this we need the training dataset since
            # we used the training-sclaer to standardize.
            y_targ_destd = training.reverse_standardization(y_targ.reshape([len(y_targ), 1]), [training.i_target_])

            # In order to compare the predicted variables with original values, we also reverse standardization of the predictors target vector
            y_pred_destd = training.reverse_standardization(y_pred.reshape([len(y_pred), 1]), [training.i_target_])

            # In order compute residuals and producing basic diagnostic plots, we also reverse standardization of the predicted predictors target vector
            y_pred_pred_destd = training.reverse_standardization(y_pred_pred.reshape([len(y_pred_pred), 1]),[training.i_target_])
        else:
            # Prepare target dataset by removing target column
            X_targ = target.data_[:, mask]

            # Core: predict target variable of target dataset, backtransformation is not needed sinde we use non-standardized data
            y_targ = rf.predict(X_targ)
            y_targ_destd = y_targ.reshape([len(y_targ), 1])

            # For computing residuals and producing basic diagnostic plots, we als need the predictions of the training set
            y_pred_pred = rf.predict(X_pred)
            y_pred_pred_destd = y_pred_pred.reshape([len(y_pred_pred), 1])

            # Since here we did not use standardized data, we do not need to back transform
            y_pred_destd = y_pred.reshape([len(y_pred), 1])

        self.y_targ_destd_ = y_targ_destd
        self.y_pred_destd_ = y_pred_destd
        self.y_pred_pred_destd_ = y_pred_pred_destd

        # Finally we store the performance metrics to the "meta"-pandas-DataFrame
        if not modelparams:
            metalist = [r2scores.mean(), mseCV, r2oob, notrees, y_targ_destd.mean(), y_targ_destd.min(), y_targ_destd.max(), y_pred_destd.mean(), y_pred_destd.min(), y_pred_destd.max()]
            metalist += [abs(m) for m in rfCV.cv_results_['mean_test_score']]
            meta[title] = metalist
        else:
            # If model parameters were already given, then we need to compute MSE seperately (otherwise it is estimated while tuning)
            msescores = model_selection.cross_val_score(rf, X_pred, y_pred, cv=10, scoring='neg_mean_squared_error')
            metalist = [r2scores.mean(), abs(msescores.mean()), rf.oob_score_, notrees, y_targ_destd.mean(), y_targ_destd.min(),
                        y_targ_destd.max(), y_pred_destd.mean(), y_pred_destd.min(), y_pred_destd.max()]
            meta[title] = metalist


def write_nk_results_to_pg(target_data, predictions, nk_key, amounts_as_targets=False, scaling=True):
    """
    This function needs to be applied after having imputed missing dwelling cost information. It writes the final results
    of the selected model for imputing a certain dwelling category to the PGDB. ATTENTION: this function is written straight
    forward and even though functional, not very generic (many things are specifically solved for the present case).
    :param target_data: List of SKLData_NK-classes of the target dataset (often only one is used, but if we want to write
    the split of renters and owners to PGDB, then we need to pass more than one class).
    :param predictions: List of numpy-vectors with predictions (again: often one is enough, but using the renter/owner-
    split involves using a list)
    :param nk_key: string indicating the target variable ('NK', 'K', 'AW', 'W', 'El', 'EnBr')
    :param amounts_as_targets: (optional) False/True to indicate whether the target variable was a quantity or an expenditure/share
    :param scaling: (optional) False/True to indicate if we worked with standardized data or not. DEPRECATED! AFTER DEBUGGING
    WE GOT RID OF THIS PARAMETER!
    """

    # Connect to the database
    conn = pg.connect("host=localhost port=5432 dbname=db_andi user=postgres  connect_timeout=2 password='postgres'")

    # Create the correct translation-dict to translate the target-variable abbreviation into pg-code depending on if we
    # targeted quantities or expenditures
    if not amounts_as_targets:
        nkcats_dict = {'NK': 'a571201', 'K': 'a571202', 'AW': 'a571203', 'W': 'a571204',
                   'U': 'a571205', 'El': 'a571301', 'EnBr': 'a571302', 'EnZentr': 'a571303'}
    else:
        nkcats_dict = {'K': 'mx571202', 'AW': 'mx571203', 'W': 'mx571204',
                       'El': 'mx571301', 'EnBr': 'mx571302', 'EnZentr': 'mx571303'}

    # The following dict is used for the habe_nk_model-table which logs which household needed computed dwelling cost data
    nkmodels_dict = {'NK': 'nkmod', 'K': 'kmod', 'AW': 'awmod', 'W': 'wmod', 'El': 'elmod', 'EnBr': 'enmod'}

    # Now we go through the target datasets --> datasets which contain households for which data needed to be imputed,
    # in order to have in the end a list of dicts indicating the household-id and the imputed value for this HH.
    predict_list = []
    for ids, ds in enumerate(target_data):
        for ihh, hh in enumerate(ds.meta_):  # go through all household-ids
            if not nk_key in ('AW', 'W', 'EnBr'):  # for 'NK', 'K' and 'El'
                predict_list.append({'haushaltid': hh, nkcats_dict[nk_key]: predictions[ids][ihh][0]})
            else:  # for 'AW', 'W' and 'EnBr': we also want to retrieve the total dwelling extra costs since we need this to compute the shares
                nktot = ds.data_[:, ds.attributes_.index('NK_tot_computed')]

                predict_list.append({'haushaltid': hh, nkcats_dict[nk_key]: predictions[ids][ihh][0], 'NKtot': nktot[ihh]})

    if nk_key == 'NK':
        # In the following lines we retrieve already existing dwelling costs in order to subtract from the imputed total
        # 'NK pauschal'.

        # Get the column-names as well as the household-IDs for the PGDB
        nkcats_dict2 = nkcats_dict.copy()
        del nkcats_dict2['K']
        del nkcats_dict2['NK']  # In principle this is not necessary, since NK should be zero (otherwise it would have be modeled)
        collist = [nkcats_dict2[ky] for ky in nkcats_dict2.keys()]
        hhids = [str(hh['haushaltid']) for hh in predict_list]

        # Extract the dwelling costs from the PGDB --> we already could sum them within the query. Although not checked,
        # I am not sure if this would have been faster.
        cur = conn.cursor(cursor_factory=pge.RealDictCursor)
        query = """
                SELECT haushaltid, {col} FROM working_tables.habe_hh_prepared WHERE haushaltid IN ({hhids})
                """.format(col=', '.join(collist), hhids=', '.join(hhids))
        cur.execute(query)
        nk_vals_sql = cur.fetchall()

        for hh in predict_list:
            y_pred = hh[nkcats_dict[nk_key]]

            # Find data from PGDB where householdid matches:
            hh_sql = [x for x in nk_vals_sql if x['haushaltid'] == hh['haushaltid']][0]

            # Go through the households-data-dict and sum all existing dwelling costs
            nk_sql = 0
            for cat in hh_sql.keys():
                if cat != 'haushaltid':
                    nk_sql += hh_sql[cat]

            # Overwrite estimated 'NK pauschal' with 'NK pauschal' with subtracted existing dwelling costs.
            # If 'NK pauschal' becomes negative, replace with 0.
            hh[nkcats_dict[nk_key]] = max(y_pred - nk_sql, 0)
        cur.close()

    if nk_key in ('EnBr', 'AW', 'W', 'K', 'El'):
        # In this IF-loop, the categories for which the results should be stored in a temporary table are written to PGDB.

        # In a first step, the respective column in the existing temporary table habe_nk_temptable which temporary
        # stores the computed values, will be set to NULL
        cur = conn.cursor()
        query = """
        UPDATE working_tables.habe_nk_temptable SET {} = NULL
        """.format(nk_key.lower())
        cur.execute(query)
        conn.commit()
        cur.close()

        # In the following, we actually write the results to PGDB
        writecur = conn.cursor()
        for hh in predict_list:  # Go through all HH for which dwelling data was imputed
            # First we try to retrieve the household-id under consideration. This is a test query to check if we need to
            # "INSERT" the values or to "UPDATE" the values.
            query = """
            SELECT haushaltid FROM working_tables.habe_nk_temptable WHERE haushaltid = {hhid}
            """.format(hhid=hh['haushaltid'])
            writecur.execute(query)
            test = writecur.fetchall()

            # Get the imputed value
            if nk_key in ('EnBr', 'W', 'AW'):
                value = hh[nkcats_dict[nk_key]] * hh['NKtot']
            else:  # for K and El
                value = hh[nkcats_dict[nk_key]]

            # Write the value to PGDB:
            if test:
                query = """
                UPDATE working_tables.habe_nk_temptable SET {nkky} = {val} WHERE haushaltid = {hhid}
                """.format(nkky=nk_key.lower(), val=value, hhid=hh['haushaltid'])
                writecur.execute(query)
                conn.commit()
            else:
                query = """
                INSERT INTO working_tables.habe_nk_temptable (haushaltid, {nkky}) VALUES ({hhid}, {val})
                """.format(nkky=nk_key.lower(), val=value, hhid=hh['haushaltid'])
                writecur.execute(query)
                conn.commit()

            # And finally, we also insert an entry into the habe_nk_model in order to track which households needed
            # imputed data and which not.
            query = """
            UPDATE working_tables.habe_nk_model SET {col} = 1 WHERE haushaltid = {haushaltid}
            """.format(col=nkmodels_dict[nk_key], haushaltid=hh['haushaltid'])
            writecur.execute(query)
            conn.commit()

        writecur.close()
        conn.close()
    else:  # This else-loop now only considers NK --> while developing the code this was not the case, therefore, the code could look more efficient nowadays
        # In case imputing data was not done via shares, the following procedure applies to write results to PGDB:
        writecur = conn.cursor()

        # Go through all HHs
        for hh in predict_list:
            query = """
            UPDATE working_tables.habe_hh_prepared SET {col} = {val} WHERE haushaltid = {haushaltid}
            """.format(col=nkcats_dict[nk_key], val=hh[nkcats_dict[nk_key]], haushaltid=hh['haushaltid'])
            writecur.execute(query)
            conn.commit()

            # In case we write NK to the PGDB, we also need to update depending variables
            if nk_key == 'NK':
                query = """
                UPDATE working_tables.habe_hh_prepared SET a5712 = a571201+a571202+a571203+a571204+a571205 WHERE haushaltid = {}
                """.format(hh['haushaltid'])
                writecur.execute(query)
                conn.commit()

                query = """
                UPDATE working_tables.habe_hh_prepared SET a571 = a5711+a5712+a5713 WHERE haushaltid = {}
                """.format(hh['haushaltid'])
                writecur.execute(query)
                conn.commit()

                query = """
                UPDATE working_tables.habe_hh_prepared SET a57 = a571+a572+a573 WHERE haushaltid = {}
                """.format(hh['haushaltid'])
                writecur.execute(query)
                conn.commit()

                query = """
                UPDATE working_tables.habe_hh_prepared SET a50 = a51+a52+a53+a56+a57+a58+a61+a62+a63+a66+a67+a68 WHERE haushaltid = {}
                """.format(hh['haushaltid'])
                writecur.execute(query)
                conn.commit()

            # And finally, we also insert an entry into the habe_nk_model in order to track which households needed
            # imputed data and which not.
            query = """
                    UPDATE working_tables.habe_nk_model SET {col} = 1 WHERE haushaltid = {haushaltid}
                    """.format(col=nkmodels_dict[nk_key], haushaltid=hh['haushaltid'])
            writecur.execute(query)
            conn.commit()

        writecur.close()
        conn.close()


def clean_up_nk_pg(excel):
    """
    This function needs to be applied after imputing dwelling extra cost data in order to write the results from the PGDB-
    table habe_nk_temptable to habe_hh_prepared_imputed. This function converts also estimated expenditures to quantities
    and vice versa. Furthermore, it updates depending, cumulated categories and ensures that no expenditure becomes negative.
    ATTENTION: this function is written straight forward and even though functional, not very generic
    (many things are specifically solved for the present case).
    :param excel: dict which needs the following keys: 'path', 'overview', 'water', 'ww', 'electricity', 'waste',
    'energy'. While 'path' should be clear, the other keys shall define the sheet-names in which the specific prices for
    the conversion from CHF to amounts are given, except for 'overview' which means the 'overview'-sheet.
    """

    # In a first step, we extract the prices from the excel-sheet and produce dicts:
    price_dicts = get_price_dicts(excel)

    water_prices_dict = price_dicts['water']
    ww_prices_dict = price_dicts['ww']
    electricity_prices_dict = price_dicts['electricity']
    oil_prices_dict = price_dicts['energy']['oil_prices']
    gas_prices_dict = price_dicts['energy']['gas_prices']
    wood_prices_dict = price_dicts['energy']['wood_prices']
    energy_dict = price_dicts['energy']
    waste_prices_dict = price_dicts['waste']

    # Connect to PG-DB
    conn = pg.connect("host=localhost port=5432 dbname=db_andi user=postgres  connect_timeout=2 password='postgres'")
    cur = conn.cursor()

    # Create the the habe_hh_prepared_imputed-table (for the moment: exact copy of habe_hh_prepared)
    query = """
    DROP TABLE IF EXISTS working_tables.habe_hh_prepared_imputed
    """
    cur.execute(query)
    conn.commit()

    query = """
    CREATE TABLE working_tables.habe_hh_prepared_imputed AS TABLE working_tables.habe_hh_prepared;
    ALTER TABLE working_tables.habe_hh_prepared_imputed ADD PRIMARY KEY (haushaltid);
    CREATE INDEX habehhprepimp_hhid ON working_tables.habe_hh_prepared_imputed USING btree (haushaltid);
    """
    cur.execute(query)
    conn.commit()
    cur.close()

    conn2 = pg.connect("host=localhost port=5432 dbname=db_andi user=postgres  connect_timeout=2 password='postgres'")
    old_isolation_level = conn2.isolation_level
    conn2.set_isolation_level(0)
    cur2 = conn2.cursor()
    query = """
    VACUUM ANALYZE working_tables.habe_hh_prepared_imputed;
    """
    cur2.execute(query)
    conn2.commit()
    cur2.close()
    conn2.set_isolation_level(old_isolation_level)
    conn2.close()

    # Retrieve all the modelled dwelling cost data as well as additional information which is needed to properly insert
    # the data to habe_hh_prepared_imputed.
    cur = conn.cursor(cursor_factory=pge.RealDictCursor)
    query = """
    SELECT hh.haushaltid, a571201, k, aw, w, el, enbr, char_nopers, char_canton, jahr08 FROM
    working_tables.habe_nk_temptable nk LEFT JOIN working_tables.habe_hh_prepared hh ON hh.haushaltid=nk.haushaltid
    LEFT JOIN original_data.habe_standard std ON nk.haushaltid=std.haushaltid
    """
    cur.execute(query)
    hhs = cur.fetchall()

    # Go through all households for which data was imputed:
    for hh in hhs:

        # In a first step we determine the household-type which is needed to extract the correct prices for waste,
        # wastewater and water (only based on number of persons)
        if hh['char_nopers'] == 1:
            hhtype = '1pHH'
        elif hh['char_nopers'] == 2:
            hhtype = '2pHH'
        elif hh['char_nopers'] == 3:
            hhtype = '3pHH'
        else:
            hhtype = '4pHH'

        # In the following we go through all categories and treat the imputed data differently.

        hhelexp = 0.  # This is meant for HHs for which electricity was not imputed.
        if hh['el']:
            # ATTENTION: this function assumes that electricity was estimated as quantity (and neither as share nor as expenditure)
            # Therefore, we need to determine the "household-type" according to the electricity consumption
            if hh['el'] * 12. <= (1600. + 2500.) / 2.:
                hheltype = 'H1'
            elif (1600.+2500.)/2. < hh['el'] * 12. <= (2500.+4500.)/2.:
                hheltype = 'H2'
            elif (2500.+4500.)/2. < hh['el'] * 12. <= (4500.+7500.)/2.:
                hheltype = 'H3'
            elif (4500.+7500.)/2. < hh['el'] * 12. <= (7500.+13000.)/2.:
                hheltype = 'H5'
            elif (7500.+13000.)/2. < hh['el'] * 12. <= (13000.+25000.)/2.:
                hheltype = 'H7'
            elif hh['el'] * 12. > (13000.+25000.)/2.:
                hheltype = 'H6'

            # Get the correct electricity price (division by 100 because price data is in Rp. per kWh)
            hh_el_price = electricity_prices_dict[(int(hh['jahr08']), int(hh['char_canton']), hheltype)] / 100.
            if hheltype == 'H3':
                hh_el_price2 = electricity_prices_dict[(int(hh['jahr08']), int(hh['char_canton']), 'H4')] / 100.
                hh_el_price = (hh_el_price + hh_el_price2) / 2.

            # Computation of the expenditure for electricity
            hhelexp = hh_el_price * hh['el']

            # Fill in the estimated amount and the computed expenditure for electricity
            query = """
            UPDATE working_tables.habe_hh_prepared_imputed SET (a571301, mx571301) = ({valexp}, {valam}) WHERE haushaltid = {haushaltid}
            """.format(valexp=hhelexp, valam=hh['el'], haushaltid=hh['haushaltid'])
            cur.execute(query)
            conn.commit()

        # The next if-loop should not be necessary, but we keep it to prevent mistakes. In principle, there should not be
        # any household for which only one amount h['aw'] and h['w'] is None, because we already imputed data for wastewater
        # via water (if only water data was available) and vice versa. Consequently, the modelling should have only be
        # applied to HHs for which both categories were lacking.
        if hh['aw']: assert hh['w']

        if hh['w']:
            assert hh['aw']  # again, this should not be necessary

            # Extract the water price and compute the amount of water in m3
            hh_water_price = water_prices_dict[(hh['char_canton'], hhtype)]
            hh_water_amount = hh['w'] / hh_water_price

            # Extract the wastewater price and compute the amount of wastewater in m3
            hh_ww_price = ww_prices_dict[(hh['char_canton'], hhtype)]
            hh_ww_amount = hh['aw'] / hh_ww_price

            # To be consistent with the approach applied in habe_hhs_preparer()/HABE_HH_V2 we take the average of both
            # water amount estimates
            hh_w_amount = (hh_ww_amount + hh_water_amount) / 2.

            # Insert the estimated expenditure and the computed amount for water and wastewater
            query = """
            UPDATE working_tables.habe_hh_prepared_imputed SET (a571203, mx571203, a571204, mx571204) = ({valexpww}, {valam}, {valexpw}, {valam}) WHERE haushaltid = {haushaltid}
            """.format(valexpww=hh['aw'], valam=hh_w_amount, valexpw=hh['w'], haushaltid=hh['haushaltid'])
            cur.execute(query)
            conn.commit()

        if hh['enbr']:
            # Juste as in HABE_HH_V2: the price for oil depends on the amount of consumption. Therefore, the oil price
            # is determined again iteratively.
            hh_oil_price = oil_prices_dict[(int(hh['jahr08']), '800 - 1500')]
            hh_oil_amount = (hh['enbr'] / hh_oil_price) * 100.  # because price in per 100 l
            if hh_oil_amount * 12. > 1500.:
                hh_oil_amount = (hh['enbr'] / oil_prices_dict[(int(hh['jahr08']), '1501 - 3000')]) * 100.
                if hh_oil_amount * 12. > 3000.:
                    hh_oil_amount = (hh['enbr'] / oil_prices_dict[(int(hh['jahr08']), '3001 - 6000')]) * 100.
                    if hh_oil_amount * 12. > 6000.:
                        hh_oil_amount = (hh['enbr'] / oil_prices_dict[(int(hh['jahr08']), '6001 - 9000')]) * 100.
                        if hh_oil_amount * 12. > 9000.:
                            hh_oil_amount = (hh['enbr'] / oil_prices_dict[(int(hh['jahr08']), '9001 - 14000')]) * 100.
                            if hh_oil_amount * 12. > 14000.:
                                hh_oil_amount = (hh['enbr'] / oil_prices_dict[(int(hh['jahr08']), '14001 - 20000')]) * 100.
                                if hh_oil_amount * 12. > 20000.:
                                    hh_oil_amount = (hh['enbr'] / oil_prices_dict[(int(hh['jahr08']), '>20000')]) * 100.

            hh_oil_amount *= energy_dict['oilenergydensity']  # convert oil amount from liters to MJ based on data from ecoinvent 3 data quality guidelines

            # Just as for oil, the price for gas depends on the consumed amount.
            hh_gas_price = gas_prices_dict[(int(hh['jahr08']), '20000')]
            hh_gas_amount = hh['enbr'] / hh_gas_price
            if hh_gas_amount * 12. > (20000. + 50000.) / 2.:
                hh_gas_amount = hh['enbr'] / gas_prices_dict[(int(hh['jahr08']), '50000')]
                if hh_gas_amount * 12. > (50000. + 100000.) / 2.:
                    hh_gas_amount = hh['enbr'] / gas_prices_dict[(int(hh['jahr08']), '100000')]
                    if hh_gas_amount * 12. > (100000. + 500000.) / 2.:
                        hh_gas_amount = hh['enbr'] / gas_prices_dict[(int(hh['jahr08']), '500000')]

            # The amount of gas is given in kWh upper calorific value (according to a phone call to Werkbetriebe Frauenfeld and a
            # a phone call to BFE). Therefore, we need to convert the UHV to LHV and kWh to MJ in the next line:
            hh_gas_amount *= (3.6 * energy_dict['lhvgasconversion'])

            # by means of the firewood price, we convert the energy expenditures also to energy amounts
            hh_wood_price = wood_prices_dict[int(hh['jahr08'])]
            hh_wood_amount = (hh['enbr'] / hh_wood_price) * 6000.  # because price in per 6000 kg
            hh_wood_amount *= energy_dict['lhvwood']

            # Finally, we compute the share-weighted average of the three energy amounts estimates
            energy_shares = energy_dict['shares']
            enamount = hh_oil_amount * energy_shares[(int(hh['jahr08']), 'oil')] + hh_gas_amount * energy_shares[
                (int(hh['jahr08']), 'gas')] + hh_wood_amount * energy_shares[(int(hh['jahr08']), 'wood')]  # lower calorific value in MJ

            # Insert the estimated expenditure and the computed amount for energy
            query = """
            UPDATE working_tables.habe_hh_prepared_imputed SET (a571302, mx571302) = ({valexp}, {valam}) WHERE haushaltid = {haushaltid}
            """.format(valexp=hh['enbr'], valam=enamount, haushaltid=hh['haushaltid'])
            cur.execute(query)
            conn.commit()

        if hh['k']:
            if not hh['char_canton'] == 25:
                hh_waste_price = waste_prices_dict[(hh['char_canton'], hhtype)]
                wasteexp = hh['k'] * hh_waste_price

                # Insert the estimated expenditure and the computed amount for waste
                query = """
                UPDATE working_tables.habe_hh_prepared_imputed SET (a571202, mx571202) = ({valexp}, {valam}) WHERE haushaltid = {haushaltid}
                """.format(valexp=wasteexp, valam=hh['k'], haushaltid=hh['haushaltid'])
                cur.execute(query)
                conn.commit()
            else:  # Special treatment of canton of GE because they do not have waste bag fees

                # Only insert the estimated amount for waste
                query = """
                UPDATE working_tables.habe_hh_prepared_imputed SET mx571202 = {valam} WHERE haushaltid = {haushaltid}
                """.format(valam=hh['k'], haushaltid=hh['haushaltid'])
                cur.execute(query)
                conn.commit()

        # Since we assume that the estimated extra costs were hidden in the "NK pauschal", we need to subtract the
        # computed/estimated expenditures from the "NK pauschal". Of course, we set "NK pauschal" to 0 if it became
        # negative.
        nknew = max(hh['a571201'] - hhelexp  - (hh['enbr'] or 0.) - (hh['aw'] or 0.) - (hh['w'] or 0.), 0)

        # And we finally insert the new "NK pauschal" in the PGDB
        query = """
        UPDATE working_tables.habe_hh_prepared_imputed SET a571201 = {val} WHERE haushaltid = {haushaltid}
        """.format(val=nknew, haushaltid=hh['haushaltid'])
        cur.execute(query)
        conn.commit()

    # Since we decided not to distinguish between EnBr and EnZentr (neither with regard to imputing data nor with regard
    # to LCA), we copy the EnZentr entries to EnBr and set EnZentr to 0 (by doing so, we bring all HHs on an even footing)
    query = """
    UPDATE working_tables.habe_hh_prepared_imputed SET a571302 = a571302+a571303"""
    cur.execute(query)
    conn.commit()

    query = """
    UPDATE working_tables.habe_hh_prepared_imputed SET mx571302 = mx571302+mx571303"""
    cur.execute(query)
    conn.commit()

    query = """
    UPDATE working_tables.habe_hh_prepared_imputed SET a571303 = 0"""
    cur.execute(query)
    conn.commit()

    query = """
    UPDATE working_tables.habe_hh_prepared_imputed SET mx571303 = 0"""
    cur.execute(query)
    conn.commit()

    # Finally, we update all depending statistics (cumulated categories)
    query = """
    UPDATE working_tables.habe_hh_prepared_imputed SET a5712 = a571201+a571202+a571203+a571204+a571205"""
    cur.execute(query)
    conn.commit()

    query = """
    UPDATE working_tables.habe_hh_prepared_imputed SET a5713 = a571301+a571302+a571303"""
    cur.execute(query)
    conn.commit()

    query = """
    UPDATE working_tables.habe_hh_prepared_imputed SET a571 = a5711+a5712+a5713"""
    cur.execute(query)
    conn.commit()

    query = """UPDATE working_tables.habe_hh_prepared_imputed SET a57 = a571+a572+a573"""
    cur.execute(query)
    conn.commit()

    query = """UPDATE working_tables.habe_hh_prepared_imputed SET a50 = a51+a52+a53+a56+a57+a58+a61+a62+a63+a66+a67+a68"""
    cur.execute(query)
    conn.commit()
    cur.close()
    conn.close()

    conn2 = pg.connect("host=localhost port=5432 dbname=db_andi user=postgres  connect_timeout=2 password='postgres'")
    old_isolation_level = conn2.isolation_level
    conn2.set_isolation_level(0)
    cur2 = conn2.cursor()
    query = """
    VACUUM ANALYZE working_tables.habe_hh_prepared_imputed;
    """
    cur2.execute(query)
    conn2.commit()
    cur2.close()
    conn2.set_isolation_level(old_isolation_level)
    conn2.close()


class SKLData_SOM(SKLData):
    """
    This is a child-class of SKLData and adds functionalities to facilitate data preparation for self-organizing maps.
    At the moment, a function to de-seasonlize HABE data is added.
    """
    def deseasonalize(self, attrs, plotpath=None, method='quantiles'):
        """
        This function removes seasonality which emerges from the fact that households are surveyed in different months.
        The function allows for two different ways to "de-seasonlize": either replace the single values with its quantile
        in its specific month or divide each value with a monthly weight that is deduced from the monthly average.
        Ex. for the first method: HH spends 6.5 CHF on oranges in January --> 6.5 CHF is the 42.3%-quantile in the month
        of January (42.3% of all households which were surveyed in January had lower expenditures) --> now replace 6.5%
        by 0.423.
        Ex. for the second method: Average expenditure on oranges of all households surveyed in January is 10 CHF and the
        overall average is 15 CHF --> Divide 6.5 by 10CHF/15CHF.
        Both methods are "approved" by the "Statistischer Dienst ETH", however, they say that the quantiles-method is
        more robust.
        PLEASE NOTE: the de-seasonalizing is performed "in place": the original data will be replaced by de-seasonlized
        data within this class.
        :param attrs: list of strings with the codes (PG-code) of the attributes which shall be de-seasonalized.
        :param plotpath: (optional) string with the path were a seasonlity plot shall be stored (shows on a a4-page the
        original data on top and the de-seasonlized data on the bottom). If this kwarg is None, then no plot will be
        created.
        :param method: (optional) choose the de-seasonlizing method ('quantiles' or 'averages'). Default is 'quantiles'
        :return: replaces values within the class by de-seasonlized data
        """

        # Retrieve the survey month data from the PGDB in order to derive monthly statistics later on:
        query = """
        SELECT haushaltid, month_no, year, month_name FROM working_tables.habe_hh_month
        """
        cur = self._conn_.cursor()
        cur.execute(query)
        sqldata = cur.fetchall()
        cur.close()

        # We will keep track of all the data which will be de-seasonlized in the 'seasonality_corrected_'-attribute:
        if not hasattr(self, 'seasonality_corrected_'):
            self.seasonality_corrected_ = []

        # Check if chosen de-seasonlizing method is implemented:
        assert method in ('quantiles', 'averages')

        # If a de-seasonlizing plot is indicated to be created, then we open a pdf-object:
        if plotpath:
            pp = PdfPages(plotpath)

        # In the next if-loop, we perform the actual de-seasonlizing with the chosen method:
        if method == 'quantiles':
            # This part conducts de-seasonalizing by replacing values by its monthly quantiles.

            # Putting the survey months into the correct order
            months = np.zeros(len(sqldata))
            hhindx = []  # list of indices indicating where to find the sql-data-HH in the data-matrix
            for hh in sqldata:
                hhind = np.where(self.meta_ == hh[0])  # find the index of the respective household
                months[hhind] = hh[1]  # insert the month at the correct place
                hhindx.append(hhind)

            # Append the month-information as last column to the data-matrix and adjust attributes-list and translation-dict accordingly
            self.data_ = np.append(self.data_, months.reshape([len(months), 1]), axis=1)
            self.attributes_.append('survey month')
            self.varnames_.update({'survey month': 'survey month'})

            # Then go through the list of attributes-codes and de-seasonalize each attribute
            for var in attrs:

                # Create a container for the de-seasonlized data:
                newvals = np.zeros(len(self.data_[:, 0]))

                # Find the index of the data-column of the attribute under consideration:
                varidx = self.attributes_.index(var)

                # Extract the data of the attributes and the survey month for further manipulation:
                varmthvals = np.copy(self.data_[:, [varidx, self.attributes_.index('survey month')]])

                # Go through all values:
                for idx, val in enumerate(varmthvals[:, 0]):
                    # Compute the quantile of the value within its survey month and store it in the newvals-container
                    mth = varmthvals[idx, 1]  # extract the survey month
                    valperc = stats.percentileofscore(varmthvals[varmthvals[:, 1] == mth, 0], val) / 100  # compute the quantile of the value within its month
                    newvals[idx] = valperc

                # Keep track of the de-seasonlized attributes and add the attribute to the list of attributes which were
                # corrected for seasonality
                self.seasonality_corrected_.append(var)

                # Replace the original data by the de-seasonlized data
                self.data_[:, varidx] = newvals

                # If indicated, create a de-seasonlity plot:
                if plotpath:
                    # Create the dataframes which are needed for the Seasonal_Fig-class
                    df_orig = pd.DataFrame(data=sqldata, columns=['haushaltid', 'month_no', 'year', 'month_name'])
                    df_orig[var] = varmthvals[hhindx, 0].flatten()
                    df_corr = pd.DataFrame(data=sqldata, columns=['haushaltid', 'month_no', 'year', 'month_name'])
                    df_corr[var] = newvals[hhindx].flatten()

                    # Call the Seasonal_Fig-class
                    Seasonal_Fig(pp, (df_orig, df_corr), [var, var], self.varnames_, not_orig_table=True)

            # Clean up the class by deleting the information of the survey month since this shall not enter the SOM-computations
            self.data_ = np.delete(self.data_, self.attributes_.index('survey month'), 1)
            del self.attributes_[self.attributes_.index('survey month')]

        elif method == 'averages':
            # This part conducts de-seasonalizing by replacing values via weights based on monthly averages. The implementation
            # of this method is based on pandas-DF (since this was the straight-forward way while developing), but we
            # expect increase in performance if it was based on numpy.

            # Then go through the list of attributes-codes and de-seasonalize each attribute.
            for var in attrs:
                # Create a pandas-DF with the information on the survey month
                var_df = pd.DataFrame(data=sqldata, columns=['haushaltid', 'month_no', 'year', 'month_name'])

                # To be able to transfer data from the pandas-DF to the data-matrix and vice versa, we need index-lists
                hhindx = []  # this list shows the index of the household in the data-matrix
                for hh in sqldata:
                    hhind = np.where(self.meta_ == hh[0])  # find the index of the respective household
                    hhindx.append(hhind)
                idxlist = [int(hh) for hh in self.meta_]  # this is a list of household-ids to reorder the dataframe later on

                # Add the original data to the pandas-dataframe (in the correct order of households):
                var_df[var] = self.data_[hhindx, self.attributes_.index(var)].flatten()

                # Compute the monthly and annual weights for deriving monthly weights:
                normalizers = var_df.groupby('month_name')[var].mean()
                yr_avg = var_df[var].mean()

                # Go through all values:
                for idx in var_df.index:
                    oldval = var_df.loc[idx, var]
                    normalizer = normalizers[var_df.loc[idx, 'month_name']] / yr_avg  # compute the monthly weight by dividing the monthly average by the annual average
                    newval = oldval / normalizer  # compute the de-seasonlized value by dividing the value under consideration by its weight
                    var_df.loc[idx, 'newvals'] = newval  # store the new value

                # Re-set the dataframe-index to household-ID and bring it it in the same order as the data-matrix
                newvals = var_df.set_index('haushaltid')
                newvals = newvals.reindex(idxlist)

                # Keep track of the de-seasonlized attributes and add the attribute to the list of attributes which were
                # corrected for seasonality
                self.seasonality_corrected_.append(var)

                # Replace the original data by the de-seasonlized data
                self.data_[:, self.attributes_.index(var)] = newvals['newvals']

                # If indicated, create a de-seasonlity plot:
                if plotpath:
                    # Create the dataframes which are needed for the Seasonal_Fig-class
                    df_orig = var_df.loc[:, ['month_name', 'year', var]]
                    df_corr = var_df.loc[:, ['month_name', 'year', 'newvals']]
                    df_corr.rename(columns={'newvals': var}, inplace=True)

                    # Call the Seasonal_Fig-class
                    Seasonal_Fig(pp, (df_orig, df_corr), [var, var], self.varnames_, not_orig_table=True)

        if plotpath:
            pp.close()

class Lasso_FI(object):
    """
    This class performs LASSO-regression and optionally also a randomized LASSO-regression with the goal of evaluating
    features according to their importance. There are two modes for running this class: either all features are evaluated
    against all features or only a subset of predictors is evaluated against a subset of targets (in this mode also
    the importance to a single attribute can be investigated). PLEASE NOTE that this class takes a SKLData-dataset
    (or one of the sub-classes of SKLData) as an input. NOTE that this class exclusively works with standardized data.
    """
    def __init__(self, data, title=None, rlasso=None, predictors=None, targets=None):
        """
        Init-function for the LASSO-feature importance class.
        :param data: SKLData-class or one of its sub-classes (SKLData_SOM or SKLData_NK).
        :param title: (optional) string to identify the data and to be used for saving to excel
        :param rlasso: (optional) True/False to indicate if a randomized LASSO shall be performed
        :param predictors: (optional) List of strings with pg-codes indicating the features that shall be investigated.
        If this parameter is turned to None, all features will be considered.
        :param targets: (optional) List of strings with pg-codes indicating the target variables that shall be investigated.
        If this parameter is left to None, all attributes will be regarded as target variables.
        """

        # In a first step we store the used data in the class. If the class is pickled, this will help to reproduce
        # the computations.
        self.data_ = data
        self.title_ = title
        self.predictors_ = predictors
        self.targets_ = targets

        # Call the core-function for performing feature importance analysis
        self.lasso_scores_, self.meta_, self.rlasso_scores_ = self._do_lasso_fi(rlasso, predictors, targets)

        # clean up results if Randomized lasso is not chosen:
        if not rlasso:
            del self.rlasso_scores_

    def _do_lasso_fi(self, rlasso, predictors, targets):
        """
        This function actually performs the feature importance computations based on LASSO-regression. For more
        information see init-function.
        :param rlasso: see init-function
        :param predictors: see init-function
        :param targets: see init-function
        """

        # Just to keep changing old code at a minimum:
        data = self.data_

        # First we need to retrieve the indices for predictors and target variables in the data. Like this, we can
        # identify which columns in the data-matrix should be retrieved. If no pg-codes for target variables and/or
        # predictor variables are passed, we assume that all attributes of the data-matrix shall be investigated.
        if targets:
            ind_t = [i for i, t in enumerate(data.attributes_) if t in targets]
        else:
            ind_t = [i for i, t in enumerate(data.attributes_)]
            targets = data.attributes_
        if predictors:
            ind_p = [i for i, p in enumerate(data.attributes_) if p in predictors]
        else:
            ind_p = [i for i, p in enumerate(data.attributes_)]
            predictors = data.attributes_

        # Since we will work exclusively with standardized data, we need to check if standardization was already done
        # before or not:
        if not hasattr(data, 'data_scaled_'):
            data.standardize_data()

        # Extract the respective data-columns for the target and the predictor data:
        target_data = data.data_scaled_[:, ind_t]
        predictor_data = data.data_scaled_[:, ind_p]

        # Prepare results-container and a container for model performance metrics. In this case here we use
        # pandas-Dataframes. Please note that we do not translate the
        # pg-codes into better readable category names in order to facilitate post-processing.
        res = pd.DataFrame(data=np.zeros((len(predictors), len(targets)), dtype=float),
                           columns=[a for a in targets],
                           index=[a for a in predictors])

        meta = pd.DataFrame(data=np.zeros((3, len(targets)), dtype=float),
                            columns=targets,
                            index=['R2', 'MSE', 'alpha'])

        # The following dataframe shall carry the randomized LASSO results. If no randomized LASSO is performed,
        # it will stay empty and be removed afterwards.
        randres = pd.DataFrame(data=np.zeros((len(predictors), len(targets)), dtype=float),
                               columns=[a for a in targets],
                               index=[a for a in predictors])

        # Core: go through all target variables and compute the feature importance for the predictors
        for i, trgt in enumerate(targets):
            # For the case that there are predictors which are at the same time target variables, we need to remove
            # these attributes from the predictors-list (otherwise it would bias the result since the attribute would
            # become the most important attribute for itself). Therefore, we define a mask:
            mask = np.ones(len(predictors), dtype=bool)
            if trgt in predictors:
                mask[predictors.index(trgt)] = False

            # The next if-loops also exclude the respective quantities if expenditures are investigated and vice versa:
            # excludes expenditures from consideration if quantities are considered.
            trgt2 = None
            if trgt.startswith('a'):
                trgt2 = trgt.replace('a','m')
                if trgt2 in predictors:
                    mask[predictors.index(trgt2)] = False
                else:
                    trgt2 += 'a'
                    if trgt2 in predictors:
                        mask[predictors.index(trgt2)] = False
            if trgt.startswith('m'):
                trgt2 = trgt.replace('a','').replace('m', 'a')
                if trgt2 in predictors:
                    mask[predictors.index(trgt2)] = False

            # Split data into predictor and target data by applying the mask
            X_red = predictor_data[:, mask]
            y = target_data[:, i]

            # The following while-loop tunes the hyperparameters and shall prevent to limit the fitting by a maximum number of iterations
            maxiter = 1500
            actual_iter = 15000
            while maxiter <= actual_iter:
                maxiter *= 10
                # LASSO-Parameters: cv=10 chosen according to "Applied Predictive Modelling"-Book, normalize is not
                # needed since already standardized, n_alphas=1000 is subjectively chosen, fit_intercept=False -->
                # checks were run and show no big difference if True or False. However, it decreases the chance to get
                # negative values and it is also not necessary if data is already centered (standardized).
                lassoCV = linear_model.LassoCV(fit_intercept=False, normalize=False, n_jobs=-1, n_alphas=1000,
                                               max_iter=maxiter, cv=10).fit(X_red, y)

                actual_iter = lassoCV.n_iter_

            # Retrieve the alpha and store the lasso-result in the class
            alfa = lassoCV.alpha_
            setattr(self, "lassoCV_{}_".format(trgt), copy.deepcopy(lassoCV))

            # In the following, we are just re-doing the LASSO-computations to compute R2
            lasso = linear_model.Lasso(alpha=alfa, fit_intercept=False, normalize=False)

            # Core: fit LASSO and compute performance metrics (here R2)
            maxiter = 1500
            actual_iter = 15000
            while maxiter <= actual_iter:  # Again we enter a while-loop to prevent the maximum iteration parameter to stop the computations
                maxiter *= 10
                lasso.max_iter = maxiter

                # Actual computation of R2
                r2scores = model_selection.cross_val_score(lasso, X_red, y, cv=10, scoring='r2')

                # Actual fit:
                lasso.fit(X_red, y)

                actual_iter = lasso.n_iter_

            # Store the lasso-performance metrics in the respective pandas-dataframe:
            meta[trgt] = [r2scores.mean(), lassoCV.mse_path_.mean(axis=1)[np.where(lassoCV.alphas_==lassoCV.alpha_)][0], alfa]

            # Copy the lasso-coefficients which are a measure of feature importance and store the lasso-class for documentation
            coeff = np.copy(lasso.coef_)
            setattr(self, "lasso_{}_".format(trgt), copy.deepcopy(lasso))
            # in case of a variable that is a predictor and target at the same time, we need to insert a NaN at the right place:
            # (if there are also quantities which directly relate to the expenditure under consideration (or vice
            # versa), we also excluded the quantity in our computation and need to insert a NaN at its place -->
            # in order to insert at the correct place, the smaller index needs to be treated first)
            if trgt2 in predictors:
                ind1 = min(predictors.index(trgt), predictors.index(trgt2))
                ind2 = max(predictors.index(trgt), predictors.index(trgt2))
                coeff = np.insert(coeff, ind1, np.nan)
                coeff = np.insert(coeff, ind2, np.nan)
            elif trgt in predictors:
                coeff = np.insert(coeff, predictors.index(trgt), np.nan)


            # Store the feature importance results in the DF-container
            res[trgt] = coeff

            # In the following we conduct a randomized LASSO approach if the option rlasso is set to True
            if rlasso:
                # Since there were sometimes error due to the randomization, we introduce a while-loop until it succeeds
                rcoeff = None
                while rcoeff is None:
                    try:
                        maxiter = max(lasso.n_iter_, 15000)
                        # Parameters for randomized LASSO: we assumed that the previously determined alpha by LASSO-CV is
                        # also a good alpha for this class. ASSUMPTION: almost all parameters were left to default values!
                        randlasso = linear_model.RandomizedLasso(alpha=alfa, fit_intercept=False, normalize=False,
                                                                     max_iter=maxiter, n_jobs=1, pre_dispatch=None
                                                                     ).fit(X_red, y)

                        # Copy the randomized lasso feature importances
                        rcoeff = np.copy(randlasso.scores_)
                        # in case of a variable that is a predictor and target at the same time, we need to insert a NaN at the right place:
                        # (if there are also quantities which directly relate to the expenditure under consideration (or vice
                        # versa), we also excluded the quantity in our computation and need to insert a NaN at its place -->
                        # in order to insert at the correct place, the smaller index needs to be treated first)
                        if trgt2 in predictors:
                            ind1 = min(predictors.index(trgt), predictors.index(trgt2))
                            ind2 = max(predictors.index(trgt), predictors.index(trgt2))
                            rcoeff = np.insert(rcoeff, ind1, np.nan)
                            rcoeff = np.insert(rcoeff, ind2, np.nan)
                        elif trgt in predictors:
                            rcoeff = np.insert(rcoeff, predictors.index(trgt), np.nan)
                    except:
                        pass

                # Store the feature importance results in the DF-container
                randres[trgt] = rcoeff

                # Store the randomized-LASSO-class for documentation:
                setattr(self, "randlasso_{}_".format(trgt), copy.deepcopy(randlasso))

        # Finally we clean up the LASSO-results by applying the same threshold as sklearn.SelectFromModel would
        # apply for LASSO in order to turn some small values to zero
        res[abs(res)<=1e-5] = 0

        return res, meta, randres

    def save2excel(self, savepath):
        """
        This function converts the results and meta-data into an excel file. Furthermore, the pg-codes are translated to
        better readable names.
        :param savepath: string with path to location where excel file shall be stored
        :return: Stores excel file in dedicated location (savepath)
        """

        # Convert the pg-code into human readable names
        lasso_df = self.lasso_scores_.copy(deep=True)
        meta_df = self.meta_.copy(deep=True)

        collist = []
        # Go through all columns and translate the code --> for quantities: a Q is inserted at the beginning of the name
        for col in lasso_df.columns:
            if col.startswith('m'):
                colnm = "Q {}".format(self.data_.varnames_[col])
            else:
                colnm = self.data_.varnames_[col]
            collist.append(colnm)
        lasso_df.columns = collist
        meta_df.columns = collist  # since the columns are exactly the same

        idxlist = []
        # Go through all indices and translate the code --> for quantities: a Q is inserted at the beginning of the name
        for idx in lasso_df.index:
            if idx.startswith('m'):
                idxnm = "Q {}".format(self.data_.varnames_[idx])
            else:
                idxnm = self.data_.varnames_[idx]
            idxlist.append(idxnm)
        lasso_df.index = idxlist

        # Check if also randomized LASSO was performed:
        if hasattr(self, 'rlasso_scores_'):
            rlasso_df = self.rlasso_scores_.copy(deep=True)
            rlasso_df.columns = collist  # since the columns are exactly the same
            rlasso_df.index = idxlist  # since the indices are exactly the same

        # Test if excel-file exists already --> if so, then the excel-file will not be overwritten but rather new sheets
        # are created in the existing file
        if not os.path.exists(savepath):
            # First save the LASSO-results to excel
            nm = self.title_ or 'Sheet1'
            lasso_df.to_excel(savepath, sheet_name=nm, na_rep='-')

            # Then re-load the workbook in order to insert also meta-data and randomized LASSO-results
            workbook = load_workbook(savepath)
            writer = pd.ExcelWriter(savepath, engine='openpyxl')
            writer.book = workbook
            nm2 = nm + '_meta'
            meta_df.to_excel(writer, nm2[0:30], na_rep='-')  # The name needs to be restricted to 30 characters for excel
            writer.save()

            # Finally, if also randomized LASSO was performed, write also these results to excel:
            if hasattr(self, 'rlasso_scores_'):
                nm += '_rlasso'
                rlasso_df.to_excel(writer, nm[0:30], na_rep='-')
                writer.save()

        # If the excelfile already exist: insert new sheets, but do not overwrite:
        else:
            nm = self.title_ or 'Sheet1'
            workbook = load_workbook(savepath)
            writer = pd.ExcelWriter(savepath, engine='openpyxl')
            writer.book = workbook
            lasso_df.to_excel(writer, nm, na_rep='-')
            writer.save()
            nm2 = nm + '_meta'
            meta_df.to_excel(writer, nm2[0:30], na_rep='-')
            writer.save()
            if hasattr(self, 'rlasso_scores_'):
                nm += '_rlasso'
                rlasso_df.to_excel(writer, nm[0:30], na_rep='-')
                writer.save()


class RF_FI(object):
    """
    This class performs Random Forest-regression with the goal of evaluating features according to their importance.
    The evaluation of feature importance is computed with the sklearn-integrated method as well as optionally also with
    the more robust method which investigates how permuting a feature decreases the overall mean squared error.
    There are two modes for running this class: either all features are evaluated against all features or only a subset
    of predictors is evaluated against a subset of targets (in this mode also
    the importance to a single attribute can be investigated). PLEASE NOTE that this class takes a SKLData-dataset
    (or one of the sub-classes of SKLData) as an input. NOTE that this class exclusively works with standardized data.
    """

    def __init__(self, data, title=None, rfmse=None, predictors=None, targets=None):
        """
        Init-function for the RF-feature importance class.
        :param data: SKLData-class or one of its sub-classes (SKLData_SOM or SKLData_NK).
        :param title: (optional) string to identify the data and to be used for saving to excel
        :param rfmse: (optional) True/False to indicate if the more robust feature importance technique which investigates
        the decrease in mean squared error shall be applied in addition or not
        :param predictors: (optional) List of strings with pg-codes indicating the features that shall be investigated.
        If this parameter is turned to None, all features will be considered.
        :param targets: (optional) List of strings with pg-codes indicating the target variables that shall be investigated.
        If this parameter is left to None, all attributes will be regarded as target variables.
        """

        # In a first step we store the used data in the class. If the class is pickled, this will help to reproduce
        # the computations.
        self.data_ = data
        self.title_ = title
        self.predictors_ = predictors
        self.targets_ = targets

        # Call the core-function for performing feature importance anaylsis
        self.rf_scores_, self.meta_, self.rfmse_scores_ = self._do_rf_fi(rfmse, predictors, targets)

        # clean up results if MSE-decrease-technique is not chosen:
        if not rfmse:
            del self.rfmse_scores_

    def _do_rf_fi(self, rfmse, predictors, targets):
        """
        This function actually performs the feature importance computations based on Random Forest-regression. For more
        information see init-function.
        :param rfmse: see init-function
        :param predictors: see init-function
        :param targets: see init-function
        """

        # Just to keep changing old code at a minimum:
        data = self.data_

        # First we need to retrieve the indices for predictors and target variables in the data. Like this, we can
        # identify which columns in the data-matrix should be retrieved. If no pg-codes for target variables and/or
        # predictor variables are passed, we assume that all attributes of the data-matrix shall be investigated.
        if targets:
            ind_t = [i for i, t in enumerate(data.attributes_) if t in targets]
        else:
            ind_t = [i for i, t in enumerate(data.attributes_)]
            targets = data.attributes_
        if predictors:
            ind_p = [i for i, p in enumerate(data.attributes_) if p in predictors]
        else:
            ind_p = [i for i, p in enumerate(data.attributes_)]
            predictors = data.attributes_

        # Since we will work exclusively with standardized data, we need to check if standardization was already done
        # before or not:
        if not hasattr(data, 'data_scaled_'):
            data.standardize_data()

        # Extract the respective data-columns for the target and the predictor data:
        target_data = data.data_scaled_[:, ind_t]
        predictor_data = data.data_scaled_[:, ind_p]

        # Prepare results-container and a container for model performance metrics. In this case here we use
        # pandas-Dataframes. Please note that we do not translate the
        # pg-codes into better readable category names in order to facilitate post-processing.
        res = pd.DataFrame(data=np.zeros((len(predictors), len(targets)), dtype=float),
                           columns=[a for a in targets],
                           index=[a for a in predictors])

        meta = pd.DataFrame(data=np.zeros((7, len(targets)), dtype=float),
                            columns=targets,
                            index=['R2', 'MSE', 'R2 OOB', 'no_Trees', '100trees', '200trees', '300trees'])


        # The following dataframe shall carry the MSE-decrease results. If no MSE-decrease-analysis is performed,
        # it will stay empty and be removed afterwards.
        rfmseres = pd.DataFrame(data=np.zeros((len(predictors), len(targets)), dtype=float),
                               columns=[a for a in targets],
                               index=[a for a in predictors])

        # Core: go through all target variables and compute the feature importance for the predictors
        for i, trgt in enumerate(targets):
            # For the case that there are predictors which are at the same time target variables, we need to remove
            # these attributes from the predictors-list (otherwise it would bias the result since the attribute would
            # become the most important attribute for itself). Therefore, we define a mask:
            mask = np.ones(len(predictors), dtype=bool)
            if trgt in predictors:
                mask[predictors.index(trgt)] = False

            # The next if-loops also exclude the respective quantities if expenditures are investigated and vice versa:
            # excludes expenditures from consideration if quantities are considered.
            trgt2 = None
            if trgt.startswith('a'):
                trgt2 = trgt.replace('a', 'm')
                if trgt2 in predictors:
                    mask[predictors.index(trgt2)] = False
                else:
                    trgt2 += 'a'
                    if trgt2 in predictors:
                        mask[predictors.index(trgt2)] = False
            if trgt.startswith('m'):
                trgt2 = trgt.replace('a', '').replace('m', 'a')
                if trgt2 in predictors:
                    mask[predictors.index(trgt2)] = False

            # Split data into predictor and target data by applying the mask
            X_red = predictor_data[:, mask]
            y = target_data[:, i]

            # RF-Parameters: oob_score=True since with that we have a performance metric for free. n_jobs=-2 is chosen
            # because this class is often used on the big Linux-Server where we should not use all cores.
            rf = ensemble.RandomForestRegressor(n_jobs=-2, oob_score=True)

            # Tuning the number of trees is chosen via GridSearch: Unfortunately, the errors occur for running
            # GridSearch in parallel, but since RF-instance will be computed already in parallel, this does not matter
            # performance-wise. cv=10 chosen according to "Applied Predictive Modelling"-Book. IMPORTANT: we changed the
            # scoring from R2 (default in sklearn) to MSE in order to have it comparable to the other classes (and because
            # it might even be a better indicator) --> ATTENTION: due to a large computational burden and due to the
            # reason that we are not attempting at finding a regression model but rather try to investigate feature
            # importances we reduced the parameter grid from [30, 100, 300, 500] to [100, 200, 300]
            rfCV = model_selection.GridSearchCV(rf, param_grid={'n_estimators': [100, 200, 300]},
                                                refit=True, n_jobs=1,
                                                pre_dispatch=None, cv=10, scoring='neg_mean_squared_error'
                                                ).fit(X_red, y)

            # Retrieve the best RF-model and the best-performing parameter
            rfbest = rfCV.best_estimator_
            notrees = rfCV.best_params_['n_estimators']

            # Store MSE and R2 of out-of-bag of best-performing RF
            mseCV = abs(rfCV.best_score_)  # mean score of CV of best RF
            r2oob = rfbest.oob_score_  # r2 of OOB while bootstrapping the best RF

            setattr(self, "rfCV_{}_".format(trgt), copy.deepcopy(rfCV))

            # After tuning the no. of trees, we can definitely fit the RF-Regression
            rf = ensemble.RandomForestRegressor(n_jobs=-2, oob_score=True, n_estimators=notrees)

            # Computing performance metrics (here R2, while MSE was already computed while tuning no. of trees)
            r2scores = model_selection.cross_val_score(rf, X_red, y, cv=10, scoring='r2')

            # Core: fit RF (somewhat redundant, but consistent with the LASSO_FI):
            rf.fit(X_red, y)

            # Store the rf-performance metrics in the respective pandas-dataframe:
            metalist = [r2scores.mean(), mseCV, r2oob, notrees]
            metalist += [abs(m) for m in rfCV.cv_results_['mean_test_score']]
            meta[trgt] = metalist

            # Copy the sklearn-integrated coefficients which are a measure of feature importance and store the rf-class for documentation
            coeff = np.copy(rf.feature_importances_)
            setattr(self, "rf_{}_".format(trgt), copy.deepcopy(rf))

            # in case of a variable that is a predictor and target at the same time, we need to insert a NaN at the right place:
            # (if there are also quantities which directly relate to the expenditure under consideration (or vice
            # versa), we also excluded the quantity in our computation and need to insert a NaN at its place -->
            # in order to insert at the correct place, the smaller index needs to be treated first)
            if trgt2 in predictors:
                ind1 = min(predictors.index(trgt), predictors.index(trgt2))
                ind2 = max(predictors.index(trgt), predictors.index(trgt2))
                coeff = np.insert(coeff, ind1, np.nan)
                coeff = np.insert(coeff, ind2, np.nan)
            elif trgt in predictors:
                coeff = np.insert(coeff, predictors.index(trgt), np.nan)

            # Store the feature importance results in the DF-container
            res[trgt] = coeff

            # In the following we conduct a robust MSE-decrease approach if the option rfmse is set to True. ATTENTION:
            # we follow the suggested approach by http://blog.datadive.net/selecting-good-features-part-iii-random-forests/,
            # except for: we do not look at R2 but at MSE!
            if rfmse:
                # Define a container for the scores:
                spl = 10  # these are the number of splits which will also enter the shuffle-split
                rfmsecoeffall = np.zeros((spl, len(X_red[0, :])))

                # Initiate a Shuffle-Split for entering cross-validation "by hand" --> Please note that the split-
                # parameters were set according to the suggested procedure in http://blog.datadive.net/selecting-good-features-part-iii-random-forests/
                # ATTENTION: because of large computational burden, the n_splits were reduced from 100 to 10.
                rs = model_selection.ShuffleSplit(n_splits=spl, test_size=.3)
                isp = 0
                for train_idx, test_idx in rs.split(X_red):
                    # Split the data:
                    X_train, X_test = X_red[train_idx], X_red[test_idx]
                    Y_train, Y_test = y[train_idx], y[test_idx]

                    # We reuse the already prepared RF-Regressor. This means, we also reuse the number of trees which
                    # was found during cross-validation above.
                    rf.fit(X_train, Y_train)

                    # compute the MSE (the blog would go for R2, but as far as we know, R would also use MSE)
                    acc = mean_squared_error(Y_test, rf.predict(X_test))

                    # Then go through all predictors and permute/shuffle the values: "Clearly, for unimportant variables,
                    # the permutation should have little to no effect on model accuracy, while permuting important
                    # variables should significantly decrease it"
                    for ip in range(X_red.shape[1]):
                        X_t = X_test.copy()  # copy --> otherwise, the shuffled X would enter the next for-iteration
                        np.random.shuffle(X_t[:, ip])

                        # Compute the new/shuffled MSE:
                        shuff_acc = mean_squared_error(Y_test, rf.predict(X_t))

                        # Compute the score for the new situation
                        score = (shuff_acc -  acc) / acc  # If we used R2, we would need to put the eq. in another way: score = (acc - shuff_acc) / acc

                        # store the score:
                        rfmsecoeffall[isp, ip] = score
                    isp += 1

                # The overall score (over all splits) is computed as the mean --> mean decrease in MSE:
                rfmsecoeff = rfmsecoeffall.mean(axis=0)

                # in case of a variable that is a predictor and target at the same time, we need to insert a NaN at the right place:
                # (if there are also quantities which directly relate to the expenditure under consideration (or vice
                # versa), we also excluded the quantity in our computation and need to insert a NaN at its place -->
                # in order to insert at the correct place, the smaller index needs to be treated first)
                if trgt2 in predictors:
                    ind1 = min(predictors.index(trgt), predictors.index(trgt2))
                    ind2 = max(predictors.index(trgt), predictors.index(trgt2))
                    rfmsecoeff = np.insert(rfmsecoeff, ind1, np.nan)
                    rfmsecoeff = np.insert(rfmsecoeff, ind2, np.nan)
                elif trgt in predictors:
                    rfmsecoeff = np.insert(rfmsecoeff, predictors.index(trgt), np.nan)

                # Clean up the scores by turning negative scores to zero:
                rfmsecoeff[rfmsecoeff<0] = 0

                # Store the feature importance results in the DF-container
                rfmseres[trgt] = rfmsecoeff

                # Store the individual scores for documentation:
                setattr(self, "rfmsecoeffall_{}_".format(trgt), copy.deepcopy(rfmsecoeffall))

        return res, meta, rfmseres

    def save2excel(self, savepath):
        """
        This function converts the results and meta-data into an excel file. Furthermore, the pg-codes are translated to
        better readable names.
        :param savepath: string with path to location where excel file shall be stored
        :return: Stores excel file in dedicated location (savepath)
        """

        # Convert the pg-code into human readable names
        rf_df = self.rf_scores_.copy(deep=True)
        meta_df = self.meta_.copy(deep=True)

        collist = []
        # Go through all columns and translate the code --> for quantities: a Q is inserted at the beginning of the name
        for col in rf_df.columns:
            if col.startswith('m'):
                colnm = "Q {}".format(self.data_.varnames_[col])
            else:
                colnm = self.data_.varnames_[col]
            collist.append(colnm)
        rf_df.columns = collist
        meta_df.columns = collist  # since the columns are exactly the same

        idxlist = []
        # Go through all indices and translate the code --> for quantities: a Q is inserted at the beginning of the name
        for idx in rf_df.index:
            if idx.startswith('m'):
                idxnm = "Q {}".format(self.data_.varnames_[idx])
            else:
                idxnm = self.data_.varnames_[idx]
            idxlist.append(idxnm)
        rf_df.index = idxlist

        # Check if also a robust MSE-decrease-analysis was performed:
        if hasattr(self, 'rfmse_scores_'):
            rfmse_df = self.rfmse_scores_.copy(deep=True)
            rfmse_df.columns = collist  # since the columns are exactly the same
            rfmse_df.index = idxlist  # since the indices are exactly the same

        # Test if excel-file exists already --> if so, then the excel-file will not be overwritten but rather new sheets
        # are created in the existing file
        if not os.path.exists(savepath):
            # First save the LASSO-results to excel
            nm = self.title_ or 'Sheet1'
            rf_df.to_excel(savepath, sheet_name=nm, na_rep='-')

            # Then re-load the workbook in order to insert also meta-data and randomized LASSO-results
            workbook = load_workbook(savepath)
            writer = pd.ExcelWriter(savepath, engine='openpyxl')
            writer.book = workbook
            nm2 = nm + '_meta'
            meta_df.to_excel(writer, nm2[0:30],
                             na_rep='-')  # The name needs to be restricted to 30 characters for excel
            writer.save()

            # Finally, if also randomized LASSO was performed, write also these results to excel:
            if hasattr(self, 'rfmse_scores_'):
                nm += '_rfmse'
                rfmse_df.to_excel(writer, nm[0:30], na_rep='-')
                writer.save()

        # If the excelfile already exist: insert new sheets, but do not overwrite:
        else:
            nm = self.title_ or 'Sheet1'
            workbook = load_workbook(savepath)
            writer = pd.ExcelWriter(savepath, engine='openpyxl')
            writer.book = workbook
            rf_df.to_excel(writer, nm, na_rep='-')
            writer.save()
            nm2 = nm + '_meta'
            meta_df.to_excel(writer, nm2[0:30], na_rep='-')
            writer.save()
            if hasattr(self, 'rfmse_scores_'):
                nm += '_rfmse'
                rfmse_df.to_excel(writer, nm[0:30], na_rep='-')
                writer.save()


class FI_Analyzer(object):
    """
    This class takes feature-importance classes as input, aggregates the results of each class by weighting the target
    variables according to their share in the total consumption (respectively in the sum of the considered expenditure
    categories), brings them on an even footing by scaling, which allows for comparison, and finally computes a total
    score over all applied FI-techniques. Furthermore, plots and excel-files are created for further analyses.
    NOTE: all input-classes need to have the same predictors and targets! ATTENTION: some parts are tailored to the
    current situation and need to be checked before the class is run.
    """
    def __init__(self, conn, listof_fi_classes, title=None):
        """
        Init-function for the feature importance analyzer class.
        :param conn: connection to the PG-DB
        :param listof_fi_classes: list of feature-importance classes (either Lasso_FI or RF_FI). IMPORTANT: the different
        classes need to have corresponding predictors and targets!
        :param title: (optional) string to identify the analysis
        """

        # store some inputs to class-attributes for documentation purposes:
        self.title_ = title
        self.listof_fi_classes_ = listof_fi_classes

        # Create a DataFrame as a container for the aggregated results (sum of weighted scores for each FI-technique)
        predictors = listof_fi_classes[0].predictors_
        fi_scores = pd.DataFrame(data=np.zeros((len(predictors), 4), dtype=float),
                           columns=['lasso_score', 'rlasso_score', 'rf_score', 'rfmse_score'],
                           index=[a for a in predictors])

        # In the following, we will extract the expenditure sums for the target-categories. These sums will be needed
        # to weight the FI-scores for the different targets.

        # Before we extract the sums, we also need to define which expenditure categories correspond to the considered
        # durable goods --> ASSUMPTION: the expenditure category will directly be assigned to a durable good statistics,
        # independent of how many durable goods there are (counts of durable goods do not enter weighting) and we also
        # do not split the expenditure among different durable good statistics (e.g.: a661100 is used for several
        # durable goods statistics and will not be distributed among them).
        # ATTENTION: this part is very specific and needs to be checked before running the class.
        cg2exp_dict = {
            'cg_nonewcars': 'a621101',
            'cg_nousedcars': 'a621102',
            'cg_nomotorbikes': 'a6212',
            'cg_nobicycles': 'a6213',
            'cg_nofreezers': 'a583101',
            'cg_nodishwashers': 'a583101',
            'cg_nowashmachines': 'a583101',
            'cg_nodriers': 'a583101',
            'cg_nocrttvs': 'a661100',
            'cg_nolcdtvs': 'a661100',
            'cg_nosat': 'a661100',
            'cg_nocams': 'a661200',
            'cg_novideorecs': 'a661100',
            'cg_novieogames': 'a661100',
            'cg_nodesktoppcs': 'a661301',
            'cg_nolaptops': 'a661301',
            'cg_noprinters': 'a661302',
            'cg_nomobilephones': 'a632100',
            'cg_nomp3players': 'a661100',
            'cg_nogps': 'a661302'
        }

        # We then extract the target variables...
        targets = listof_fi_classes[0].targets_

        #...and exclude the quantities-categories
        targetsshort = [t for t in targets if not t.startswith('m')]

        # To compute the total sum of the considered targets, we need to know the lengths of the attributes, because:
        # there is the possibility of directly coupled/correlated attributes (one sub-category being part of an
        # aggregated category) --> therefore we only include high level/aggregated categories and exclude subsequent
        # subcategories.
        targetlens = [len(cg2exp_dict[t]) if t in cg2exp_dict.keys() else len(t) for t in targetsshort]  # here we already need the dict which translates durable goods into corresponding expenditure category

        # combine attribute name with length of name (the smaller the name the more aggregated it is) and sort by length:
        zipped = zip(targetsshort, targetlens)
        zipped = list(sorted(zipped, key=lambda t: t[1]))

        # In the following we create a list of attributes which will be used to compute the total sum of considered
        # expenditure classes --> do not include categories which are sub-categories of high level categories
        checked = []
        totsumlist = []
        for a, _ in zipped:
            a = cg2exp_dict[a] if a in cg2exp_dict.keys() else a
            test = [1 for c in checked if c in a]  # check if an attribute's name which is already in the list is part of the name of the attribute under consideration
            if sum(test) == 0:  # this test is true if there is no aggregated attribute considered "above" the attribute under consideration
                totsumlist.append(a)
                checked.append(a)

        # Create the query and get the sums of the expenditures (please note that the query will have some "duplicates",
        # but they will be filtered out by postgres):
        query = """
        SELECT {}, sum({}) FROM working_tables.habe_hh_prepared_imputed
        """.format(', '.join(['sum({}) as {}'.format(a, a) for a in
                              [cg2exp_dict[t] if t in cg2exp_dict.keys() else t for t in targetsshort]]),
                   '+'.join([a for a in totsumlist]))
        cur = conn.cursor(cursor_factory=pge.RealDictCursor)
        cur.execute(query)
        sql_weights = cur.fetchall()
        cur.close()

        # Finally, we compute a dict with weights for the different expenditure categories:
        weights_dict = {ky: sql_weights[0][ky] / sql_weights[0]['sum'] for ky in sql_weights[0].keys() if ky != 'sum'}

        # This dict then needs to be updated for the quantities-categories:
        for m in list(set(targets) - set(targetsshort)):  # the difference between targets and targetsshort are the quantities
            # change the attributes name accordingly
            q = m
            a = q.replace('a', '').replace('m', 'a')

            # Assign the same weight to the quantity-category as for the expenditure category --> this is kind of
            # consistent with the durable goods
            weights_dict[m] = weights_dict[a]

        # Store the weights-dict for documentation:
        self.weights_dict_ = weights_dict

        # Then go through all feature importance classes and compute the aggregated and weighted total score:
        for fi_class in listof_fi_classes:

            # Go through all predictors and collect the scores of the different FI-techniques in a dict
            for p in fi_class.predictors_:
                totscores = {'lasso': 0, 'rlasso':0, 'rf': 0, 'rfmse': 0}

                # For each predictor: go through all targets and extract the score, weigh it and sum the weighted scores
                for t in targets:

                    # rename the target name to extract the correct weight
                    tr = t
                    tr = cg2exp_dict[tr] if tr in cg2exp_dict.keys() else tr

                    for met in totscores.keys():

                        # we go through all four FI-techniques and check if the corresponding attribute (dataframe with
                        # corresponding scores) is available
                        if hasattr(fi_class, '{}_scores_'.format(met)):

                            # Retrieve the score-dataframe for certain technique
                            fi_pd = getattr(fi_class, '{}_scores_'.format(met))

                            # Get the score of the predictor for a certain target:
                            score = float(fi_pd.loc[p, t])

                            # If the score is not null, we add the weighted score to the total sum
                            if not np.isnan(score):
                                weighted_score = abs(score) * weights_dict[tr]
                                totscores[met] += weighted_score

                # Finally we fill in the weighted total scores in the dataframe-results container
                for met in totscores.keys():
                    if fi_scores.loc[p, "{}_score".format(met)] == 0:  # this if-check shall ensure that previously inserted scores are not overwritten
                        fi_scores.loc[p, "{}_score".format(met)] = totscores[met]

        # Store the results as a class-attribute
        self.fi_scores_ = fi_scores

        # Bring the scores of the different FI-techniques on an even footing by scaling and compute an overall total score
        self.fi_scores_scaled_ = self._scale_scores()

    def _scale_scores(self):
        """
        This function brings the scores of the different FI-techniques on an even footing by scaling and compute an
        overall total score.
        :return: Scaled feature importance scores and a total score
        """

        # Make a copy to keep the original scores:
        fi_scores_scaled = self.fi_scores_.copy(deep=True)

        # Go through all columns (=FI-techniques) to scale each the scores for each FI-technique separately
        for col in self.fi_scores_.columns:
            # For the scaling, we use a minmax-scaler of sci-kit learn
            minmax = preprocessing.MinMaxScaler()
            ranks = minmax.fit_transform(np.array(self.fi_scores_[col]).reshape(-1,1))
            fi_scores_scaled[col] = ranks

        # Finally, the total score is just the sum over all FI-techniques
        fi_scores_scaled['total_score'] = fi_scores_scaled.sum(axis=1)

        return fi_scores_scaled

    def create_plots(self, savepath):
        """
        This function creates a line plot which shall visualize the different score-rankings of the feature importances
        as well as a pairgrid-plot which gives information on how well the different FI-techniques correspond with each
        other.
        :param savepath: string with the path where the generated PDF shall be saved to.
        """
        def corrfunc(x, y, **kws):
            """
            This is a helper-function for the pairgrid-plot. It computes different correlation measures and adds them
            to the lower part of the PairGrid-Plot.
            """

            # Computation of the correlation measures which will be plotted in the PairGrid_plot
            pearson, _ = stats.pearsonr(x, y)
            kendall, _ = stats.kendalltau(x, y)
            spearman, _ = stats.spearmanr(x, y)

            # In the following, we add the correlation statistics to an AnchoredText and scale the size of the text
            # proportional to the spearman-correlation measure (but not above 18pts and not below 6pts).
            ax = plt.gca()
            txt = AnchoredText("pearson: {}\nkendall: {}\nspearman: {}".format(round(pearson, 2), round(kendall, 2),
                                                                               round(spearman, 2)),
                               prop=dict(size=max(abs(spearman) * 18, 6)), loc=10, frameon=False)
            ax.add_artist(txt)

        # The following for- and if-loop shall ensure that for the plots only FI-techniques are treated for which
        # effectively scores were computed
        cols = []
        for col in self.fi_scores_.columns:
            # We go through all columns and if the sum of the column is not zero, we add it to the list of columns or
            # list of FI-techniques respectively, which shall be considered for plotting
            if self.fi_scores_.sum()[col]:
                cols.append(col)

        # Make copies of the parts (=columns=FI-techniques) of the FI-scores dataframes which shall be plotted
        fi_scores = self.fi_scores_.loc[:, cols].copy(deep=True)
        cols.append('total_score')
        fi_scores_scaled = self.fi_scores_scaled_.loc[:, cols].copy(deep=True)

        # Initializing the figure (A4-landscape)
        pp = PdfPages(savepath)

        fig_width_cm = 21  # A4 page
        fig_height_cm = 29.7
        inches_per_cm = 1 / 2.58  # Convert cm to inches
        fig_width = fig_width_cm * inches_per_cm  # width in inches
        fig_height = fig_height_cm * inches_per_cm  # height in inches
        fig_size = [fig_width,
                    fig_height]  # height and width are in the order needed for portrait
        paper_rc = {'lines.linewidth': 2, 'lines.markersize': 6}  # parameters to control the linewidth and markersize
        sns.set_context("talk", rc=paper_rc)
        fig = plt.figure(figsize=fig_size)

        # First, we create two sub-figures showing the sorted scores for each FI-technique separately and for the total
        # score respectively
        gridspec.GridSpec(2, 1)
        plt.suptitle(self.title_)

        # Go through all FI-techniques and plot the respective scores in a sorted way
        plt.subplot2grid((2, 1), (0, 0), colspan=1, rowspan=1)
        plt.title("Scaled Feature Importance")
        for i, met in enumerate(fi_scores_scaled.columns):
            if met == 'total_score':
                continue
            plt.plot(sorted(fi_scores_scaled[met], reverse=True),color=cm.viridis(i*100), label=met)
        plt.xlabel('Attributes')
        plt.ylabel('Scaled Feature Importance Scores')
        plt.legend(loc='best')

        # Plot in a second subplot the total score of all techniques:
        plt.subplot2grid((2, 1), (1, 0), colspan=1, rowspan=1)
        plt.title("Total Score")
        plt.plot(sorted(fi_scores_scaled['total_score'], reverse=True), color=cm.viridis(150))
        plt.xlabel('Attributes')
        plt.ylabel('Total score')

        # Save and close figure
        fig.savefig(pp, format='pdf', papertype='a4')
        plt.close(fig)
        del fig

        # The following two lines shall compute the size for the PairGrid-Plot to scale the whole plot such that its
        # width is nearly the size of a A4-paper
        facets = len(cols[:-1])
        sz = round(fig_width / facets, 0)

        # Compute the PairGrid-plot
        g = sns.PairGrid(fi_scores_scaled.loc[:,cols[:-1]], size=sz, palette=["blue_d"])
        g.map_diag(sns.distplot, kde=False, norm_hist=True)
        g.map_upper(plt.scatter, alpha=0.5)  # sns.kdeplot in combination with cmap="Blues_d"  is also nice
        g.map_lower(corrfunc)

        # Save and close the PairGrid-Plot
        g.savefig(pp, format='pdf', papertype='a4')
        del g

        pp.close()


    def save2excel(self, savepath):
        """
        This function converts the weighted aggregated total scores (scaled and unscaled) into an excel file.
        Furthermore, the pg-codes are translated to better readable names.
        :param savepath: string with path to location where excel file shall be stored
        :return: Stores excel file in dedicated location (savepath)
        """

        # Convert the pg-code into human readable names
        fi_scores_df = self.fi_scores_.copy(deep=True)
        fi_scores_scl_df = self.fi_scores_scaled_.copy(deep=True)

        idxlist = []
        # Go through all indices and translate the code --> for quantities: a Q is inserted at the beginning of the name
        for idx in fi_scores_df.index:
            if idx.startswith('m'):
                idxnm = "Q {}".format(self.listof_fi_classes_[0].data_.varnames_[idx])
            else:
                idxnm = self.listof_fi_classes_[0].data_.varnames_[idx]
            idxlist.append(idxnm)
        fi_scores_df.index = idxlist
        fi_scores_scl_df.index = idxlist

        # Test if excel-file exists already --> if so, then the excel-file will not be overwritten but rather new sheets
        # are created in the existing file
        if not os.path.exists(savepath):
            # First save the unscaled scores to excel
            nm = self.title_ or 'Sheet1'
            fi_scores_df.to_excel(savepath, sheet_name=nm, na_rep='-')

            # Then re-load the workbook in order to insert also scaled scores
            workbook = load_workbook(savepath)
            writer = pd.ExcelWriter(savepath, engine='openpyxl')
            writer.book = workbook
            nm2 = nm + '_scaled_scores'
            fi_scores_scl_df.to_excel(writer, nm2[0:30], na_rep='-')  # The name needs to be restricted to 30 characters for excel
            writer.save()

        # If the excelfile already exist: insert new sheets, but do not overwrite:
        else:
            nm = self.title_ or 'Sheet1'
            workbook = load_workbook(savepath)
            writer = pd.ExcelWriter(savepath, engine='openpyxl')
            writer.book = workbook
            fi_scores_df.to_excel(writer, nm, na_rep='-')
            writer.save()
            nm2 = nm + '_scaled_scores'
            fi_scores_scl_df.to_excel(writer, nm2[0:30], na_rep='-')
            writer.save()


class Clustering_Analyzer(object):
    """
    This class not only applies different clustering algorithms  but also provides different methods to analyze
    SKLEARN-clustering-algorithms (at the moment implemented: Agglomerative clustering, K-Means, DBSCAN). The class
    takes a SKLEARN-clustering-class as an input (this means, the clustering class needs to be
    instantiated outside the present class). Note also, that as data-input a SKLData-Class or a SOMPY-SOM can be passed.
    """
    def __init__(self, clusterer, data, tuning_params, scaling=True,
                 saving={'title': 'clustering_analysis', 'savepath': r"D:\froemelt\Desktop"}):
        """
        Init-function for the clustering analyzer class.
        :param clusterer: SKLEARN-clustering-class instance (at the moment: K-Means, DBSCAN or Agglomerative Clustering)
        :param data: Either a SKLData-class or a SOMPY-SOM-class
        :param tuning_params: dict to pass different tuning parameters. For K-Means or agglomerative
        clustering, one can pass by the key 'n_clusters' a list of cluster-numbers. For DBSCAN one can pass
        list of 'min_samples' or 'eps' by these keys.
        :param scaling: True/False (optional): indicates if the data that needs to be clustered shall be scaled.
        NOTE: for clustering it is actually always recommended to scale.
        :param saving: optional: dict with keys: 'title' and 'savepath' to give the option to save the clustering analysis results
        """

        # Get and optionally prepare the data that shall be clustered:
        self.data_ = data
        if scaling and type(data)==SKLData:
            try:
                self.clustering_data_ = data.data_scaled_
            except:
                data.standardize_data()
                self.clustering_data_ = data.data_scaled_
        elif type(data)==SKLData:
            self.clustering_data_ = data.data_
        elif scaling and type(data)==smp.sompy.SOM:
            self.clustering_data_ = data.codebook.matrix
        elif type(data)==smp.sompy.SOM:
            self.clustering_data_ = data._normalizer.denormalize_by(data.data_raw, data.codebook.matrix)

        # Prepare a results container for the scores (the final container will be different depending on clustering)
        self.scores_dict_ = {}

        # Depending on the SKLEARN-clustering-instance which is passed, we apply different methods:
        if type(clusterer)==AgglomerativeClustering:
            self._hierarchical_clustering(clusterer, tuning_params, saving)
        elif type(clusterer)==KMeans:
            self._kmeans_clustering(clusterer, tuning_params, saving)
        elif type(clusterer)==DBSCAN:
            self._dbscan_clustering(clusterer, tuning_params, saving)

        # If a "saving-dict" is passed, we pickle the whole clustering-analyzer class
        if saving:
            with open(os.path.join(saving['savepath'], "{}.pickle".format(saving['title'])), 'wb') as f:
                pickle.dump(self, f)

    def _hierarchical_clustering(self, clusterer, tuning_params, saving):
        """
        Application and analysis of agglomerative clustering
        :param clusterer: SKLEARN-clustering instance
        :param tuning_params: dict of tuning parameters (see init-function)
        :param saving: saving-dict (see init-function)
        :return: Analyses of agglomerative clustering
        """

        # The silhouette plots of all versions that shall be computed (according to tuning-params) shall be stored in a pdf
        if saving:
            pp = PdfPages(os.path.join(saving['savepath'],'all_silhouette_plots_{}.pdf'.format(saving['title'])))

        # Prepare the scores-dict
        self.scores_dict_.update({'n_clusters': [], 'silhouette_avg': [], 'calinski_harabasz_score': [], 'largest_dendro_gap': []})

        # Start a pyprind-Bar to observe the progress
        bar = pyprind.ProgBar(len(tuning_params['n_clusters']), monitor=True)

        # Now we go through all the tuning_params (in this case the no. of clusters) and apply clustering
        for nc in tuning_params['n_clusters']:

            # Set the number of clusters to investigate (actually, when we perform agglomerative clustering, the algorithm
            # basically computes all clustering-possibilities. However, it is easier to analyze if we go iteratively through
            # the different numbers of clusters):
            clusterer.n_clusters = nc

            # Core: do clustering
            clusterer.fit(self.clustering_data_)

            # save the trained clusterer as attribute of the class
            nm = "hierarchical_clustering_{}_clusters_".format(nc)
            setattr(self, nm, copy.deepcopy(clusterer))

            # Compute the average silhouette-coefficient and the silhouette-values for each sample:
            silhouette_avg, sample_silhouette_values = self._do_silhouette_analysis(nm, clusterer.labels_)

            # Store the Silhouette-coefficients to the scores-dict:
            self.scores_dict_['n_clusters'].append(nc)
            self.scores_dict_['silhouette_avg'].append(silhouette_avg)

            # Compute the Calinski-Harabasz-Score and append it also to the scores-dict
            self.scores_dict_['calinski_harabasz_score'].append(calinski_harabaz_score(self.clustering_data_, clusterer.labels_))

            if saving:
                # Create the Silhouette-plot for the no. of clusters under consideration and add it to the PDF
                fig = self._silhouette_plot(clusterer, silhouette_avg, sample_silhouette_values, "{} clusters".format(nc))
                fig.savefig(pp, format='pdf', orientation='landscape', papertype='a4')
                plt.close(fig)
                del fig

            bar.update()

        # Compute the largest gap in the dendrogram (as far as I can judge, we can just take the last instance of clusterer,
        # since the dendrogram should be the same for all clusterings, btw: for more information on linkage-matrix see
        # the function for deep hierarchical analysis):
        linkmat = linkage_matrix(clusterer)

        # Next, we retrieve the distances in the dendrogram from the linkage-matrix
        dists = linkmat[:, 2]

        # Then we take the first order difference
        gaps = np.diff(dists, 1)
        gaps_rev = gaps[::-1]
        k = gaps_rev.argmax() + 2

        # Add this largest gap to the scores-dict (it is the same for all clusterings):
        ks = list(k * np.ones(len(self.scores_dict_['n_clusters'])))
        self.scores_dict_['largest_dendro_gap'] = ks

        if saving:
            pp.close()

            # Finally, we also plot all average-silhouette-coefficients as well as all calinski-harabasz-scores in one plot
            fig = self._tot_silhouette_calinski_plot(tuning_params, saving)
            fig.savefig(os.path.join(saving['savepath'],'silhouette_calinski_overview_plot_{}.pdf'.format(saving['title'])),
                        format='pdf', orientation='landscape', papertype='a4')
            plt.close(fig)
            del fig

    def _kmeans_clustering(self, clusterer, tuning_params, saving):
        """
        Application and analysis of K-Means clustering
        :param clusterer: SKLEARN-clustering instance
        :param tuning_params: dict of tuning parameters (see init-function)
        :param saving: saving-dict (see init-function)
        :return: analyses of K-Means clustering
        """

        # The silhouette plots of all versions that shall be computed (according to tuning-params) shall be stored in a pdf
        if saving:
            pp = PdfPages(os.path.join(saving['savepath'],'all_silhouette_plots_{}.pdf'.format(saving['title'])))

        # Prepare the scores-dict
        self.scores_dict_.update({'n_clusters': [], 'silhouette_avg': [], 'calinski_harabasz_score': [], 'inertia': []})

        # Start a pyprind-Bar to observe the progress
        bar = pyprind.ProgBar(len(tuning_params['n_clusters']), monitor=True)

        # Now we go through all the tuning_params (in this case the no. of clusters) and apply clustering
        for nc in tuning_params['n_clusters']:

            # Set the number of clusters to investigate
            clusterer.n_clusters = nc

            # Core: do clustering. Since we do not want to be limited by the number of maximum iterations, we enter
            # a while-loop to dynamically set the maximum iterations.
            maxiter = 30
            actual_iter = 300
            while maxiter <= actual_iter:
                maxiter *= 10
                clusterer.max_iter = maxiter
                clusterer.fit(self.clustering_data_)
                actual_iter = clusterer.n_iter_

            # save the trained clusterer as attribute of the class
            nm = "kmeans_{}_clusters_".format(nc)
            setattr(self, nm, copy.deepcopy(clusterer))

            # Compute the average silhouette-coefficient and the silhouette-values for each sample:
            silhouette_avg, sample_silhouette_values = self._do_silhouette_analysis(nm, clusterer.labels_)

            # Store the Silhouette-coefficients to the scores-dict:
            self.scores_dict_['n_clusters'].append(nc)
            self.scores_dict_['silhouette_avg'].append(silhouette_avg)

            # Compute the Calinski-Harabasz-Score and append it also to the scores-dict
            self.scores_dict_['calinski_harabasz_score'].append(calinski_harabaz_score(self.clustering_data_, clusterer.labels_))

            # Store also the inertia in the scores-dict:
            self.scores_dict_['inertia'].append(clusterer.inertia_)

            if saving:
                # Create the Silhouette-plot for the no. of clusters under consideration and add it to the PDF
                fig = self._silhouette_plot(clusterer, silhouette_avg, sample_silhouette_values, "{} clusters".format(nc))
                fig.savefig(pp, format='pdf', orientation='landscape', papertype='a4')
                plt.close(fig)
                del fig

            bar.update()

        if saving:
            pp.close()

            # In the case of k-means, we do not only plot a silhouette- and calinski-overview-plot, but also an inertia-
            # elbow-plot:
            pp = PdfPages(os.path.join(saving['savepath'], 'silhouette_calinski_overview_and_inertia_elbow_plot_{}.pdf'.format(saving['title'])))

            # First, we plot all average-silhouette-coefficients as well as all calinski-harabasz-scores in one plot
            fig = self._tot_silhouette_calinski_plot(tuning_params, saving)
            fig.savefig(pp, format='pdf', orientation='landscape', papertype='a4')
            plt.close(fig)
            del fig

            # And then we add the inertia-elbow-plot:
            fig = self._inertia_elbow_plot(saving)
            fig.savefig(pp, format='pdf', orientation='landscape', papertype='a4')
            plt.close(fig)
            del fig

            pp.close()


    def _dbscan_clustering(self, clusterer, tuning_params, saving):
        """
        Application and analysis of DBSCAN clustering.
        :param clusterer: SKLEARN-clustering instance
        :param tuning_params: dict of tuning parameters (see init-function)
        :param saving: saving-dict (see init-function)
        :return: analyses of DBSCAN clustering
        """

        # The silhouette plots of all versions that shall be computed (according to tuning-params) shall be stored in a pdf
        if saving:
            pp = PdfPages(os.path.join(saving['savepath'],'all_silhouette_plots_{}.pdf'.format(saving['title'])))

        # Prepare the scores-dict
        self.scores_dict_.update({'eps': [], 'min_samples': [], 'n_clusters': [], 'silhouette_avg': [], 'calinski_harabasz_score': [], 'outlier_share': []})

        # Start a pyprind-Bar to observe the progress
        bar = pyprind.ProgBar(len(tuning_params['min_samples']), monitor=True)

        # Now we will enter two for-loops in order to compute all combinations of min-samples and eps-parameters:
        for s in tuning_params['min_samples']:
            # Set min_samples:
            clusterer.min_samples = s
            for e in tuning_params['eps']:
                # Set eps:
                clusterer.eps = e

                # Core: do clustering.
                clusterer.fit(self.clustering_data_)

                # save the trained clusterer as attribute of the class
                nm = "dbscan_eps_{:3f}_minsamples_{}_".format(e, s)
                setattr(self, nm, copy.deepcopy(clusterer))

                # Compute the average silhouette-coefficient and the silhouette-values for each sample as well as the
                # Calinski-Harabasz-Score. In the case of
                # DBSCAN we have to use a try-statement since there is a chance that no cluster is found.
                try:
                    silhouette_avg, sample_silhouette_values = self._do_silhouette_analysis(nm, clusterer.labels_, dbscan=True)
                    calinski_harabaz = calinski_harabaz_score(self.clustering_data_[clusterer.labels_ != -1, :], clusterer.labels_[clusterer.labels_ != -1])
                except:
                    silhouette_avg = np.nan
                    calinski_harabaz = np.nan

                # Store the Silhouette-coefficients as well as the Calinski-Harabasz-Coefficient to the scores-dict
                self.scores_dict_['eps'].append(e)
                self.scores_dict_['min_samples'].append(s)
                self.scores_dict_['n_clusters'].append(
                    len(set(clusterer.labels_)) - (1 if -1 in clusterer.labels_ else 0))
                self.scores_dict_['silhouette_avg'].append(silhouette_avg)
                self.scores_dict_['calinski_harabasz_score'].append(calinski_harabaz)
                self.scores_dict_['outlier_share'].append(len(clusterer.labels_[clusterer.labels_ == -1])/len(clusterer.labels_))

                if saving and not np.isnan(silhouette_avg):
                    # Create the Silhouette-plot for the no. of clusters under consideration and add it to the PDF
                    fig = self._silhouette_plot(clusterer, silhouette_avg, sample_silhouette_values, "eps: {}; minsamples:{}".format(e, s))
                    fig.savefig(pp, format='pdf', orientation='landscape', papertype='a4')
                    plt.close(fig)
                    del fig

            bar.update()

        if saving:
            pp.close()
            # Finally, we also plot all average-silhouette-coefficients as well as all calinski-harabasz-scores in one plot
            fig = self._tot_silhouette_calinski_plot(tuning_params, saving)
            fig.savefig(os.path.join(saving['savepath'],'silhouette_calinski_overview_plot_{}.pdf'.format(saving['title'])),
                        format='pdf', orientation='landscape', papertype='a4')
            plt.close(fig)
            del fig

    def _do_silhouette_analysis(self, nm, labels, dbscan=False):
        """
        This function computes the sample-silhouette (silhouette coefficient for each sample) as well as the average
        silhouette coefficient
        :param nm: string with ID-name of the analysis
        :param labels: np.array of cluster-labels for each sample (can be obtained by SKLEARN-attribute .labels_)
        :param dbscan: True/False to indicate if DBSCAN was applied.
        :return: Average silhouette coefficient as well as all silhouette coefficients for all samples
        """

        # Since we should not take the outlier-label into account when performing Silhouette Analysis, we exclude the
        # outlier-label of DBSCAn in the following:
        if dbscan:
            clustering_data = self.clustering_data_[labels != -1, :]
            labels = labels[labels != -1]
        else:
            clustering_data = self.clustering_data_

        # Compute average silhouette-score:
        silhouette_avg = silhouette_score(clustering_data, labels)

        # Compute silhouette-score for each sample
        sample_silhouette_values = silhouette_samples(clustering_data, labels)

        # Construct two attribute-names and store the above silhouette-computations as class-attributes
        nm3 = nm + "silhouette_avg_"
        nm4 = nm + "sample_silhouette_values_"
        setattr(self, nm3, silhouette_avg)
        setattr(self, nm4, sample_silhouette_values)

        return silhouette_avg, sample_silhouette_values


    def _silhouette_plot(self, clusterer, silhouette_avg, sample_silhouette_values, variant):
        """
        This function constructs a Silhouette plot (similar to the example of SKLEARN).
        :param clusterer: SKLEARN-clusterer instance (see init-function)
        :param silhouette_avg: Average Silhouette coefficient
        :param sample_silhouette_values: np.array of all Silhouette coefficients of all samples
        :param variant: String for identifying the plot (e.g. "5 clusters")
        :return: returns a figure with the Silhouette plot
        """

        # Setting up the figure:
        fig_width_cm = 21  # A4 page
        fig_height_cm = 29.7
        inches_per_cm = 1 / 2.58  # Convert cm to inches
        fig_width = fig_width_cm * inches_per_cm  # width in inches
        fig_height = fig_height_cm * inches_per_cm  # height in inches
        fig_size = [fig_height, fig_width]  # height and width are in the order needed for landscape

        fig = plt.figure(figsize=fig_size)

        # y_lower is a parameter used for setting the distance between the clusters in the plot
        y_lower = 10

        # The following try-statement checks if n_cluster is available. This is e.g. not the case for DBSCAN. For the
        # latter, one has to "count" the number of labels (but we exclude identified outliers (-1))
        try:
            nclusters = clusterer.n_clusters
        except:
            nclusters = len(set(clusterer.labels_)) - (1 if -1 in clusterer.labels_ else 0)
            lbls = clusterer.labels_.copy()
            clusterer.labels_ = clusterer.labels_[clusterer.labels_ != -1]

        # Go through all clusters:
        #for i in range(nclusters):
        for i in set(clusterer.labels_):  # this was inserted later on --> it allows for more general situations, however, the try-statement above "collapses" to only checking if DBSCAN was applied
            # Compute the average Silhouette-coefficient for the cluster under consideration (ith cluster)
            silhouette_i_avg = round(np.mean(sample_silhouette_values[clusterer.labels_ == i]), 2)

            # Extract all Silhouette-coefficients for all samples of cluster i and sort them:
            ith_cluster_silhouette_values = sample_silhouette_values[clusterer.labels_ == i]
            ith_cluster_silhouette_values.sort()

            # Define the upper y via the cluster size (the area between lower y and upper y shall be colored)
            size_cluster_i = ith_cluster_silhouette_values.shape[0]
            y_upper = y_lower + size_cluster_i

            # Extract a color and fill the area between y_lower and y_upper (and in x-direction between 0 and the
            # Sample-Silhouettes)
            color = sns.color_palette('viridis', nclusters)[i]  # cubehelix would also be nice
            plt.fill_betweenx(np.arange(y_lower, y_upper), 0, ith_cluster_silhouette_values,
                              facecolor=color, edgecolor=color, alpha=0.7)

            # Label the silhouette plots with their cluster numbers at the middle; the try-statement was inserted later on and allows for a more general application
            try:
                plt.text(0.05, y_lower + 0.5 * size_cluster_i, "Cluster: {}; Silh.avg: {}".format(clusterer.names_dict_[i], silhouette_i_avg))
            except:
                plt.text(0.05, y_lower + 0.5 * size_cluster_i, "Cluster: {}; Silh.avg: {}".format(i, silhouette_i_avg))

            # Compute the new y_lower for next plot
            y_lower = y_upper + 10  # 10 for the 0 samples

        # Add title and labels:
        plt.title("Silhouette plot for {}; Overall Silhouette score: {:.2f}".format(variant, silhouette_avg))
        plt.xlabel("The silhouette coefficient values")
        plt.ylabel("Clusters")

        # The vertical line for average silhouette score of all the values
        plt.axvline(x=silhouette_avg, color="red", linestyle="--")

        # Vertical line for showing x=0
        plt.axvline(x=0, color="black", linestyle="-", linewidth=0.5)

        plt.yticks([])  # Clear the yaxis labels / ticks

        # Make nice x-ticks:
        plt.xticks(np.arange(round(sample_silhouette_values.min() * 10) / 10,
                             round(sample_silhouette_values.max() * 10) / 10, 0.1))

        # In case of DBSCAN, we need to set the labels back to original state
        try:
            clusterer.labels_ = lbls
        except:
            pass

        return fig

    def _tot_silhouette_calinski_plot(self, tuning_params, saving):
        """
        This function returns a plot of all computed average Silhouette-coefficients as well as all Calinski-Harabasz-values.
        It produces a subplot for each tuning parameter. PLEASE NOTE: in principle the
        idea is that the corresponding plots of the different parameters should show the same curve of the metric scores,
        but the labels of the x-axis should be different (according to the parameters. However, it could be that this
        does not work properly.
        :param tuning_params: dict to pass different tuning parameters (see init-function)
        :param saving: dict with information where to store the data (see init-function)
        :return: figure with overview of all average Silhouette-coefficients and all Calinski-Harabasz-Values
        """

        # First we set up the amount of parameters for which the plot shall be performed (for K-Means and agglomerative
        # clustering it is only one parameter, though for DBSCAN there are two)
        params = [ky for ky in tuning_params.keys()]

        # Set up the figure:
        fig_width_cm = 21  # A4 page
        fig_height_cm = 29.7
        inches_per_cm = 1 / 2.58  # Convert cm to inches
        fig_width = fig_width_cm * inches_per_cm  # width in inches
        fig_height = fig_height_cm * inches_per_cm  # height in inches
        fig_size = [fig_height, fig_width]  # height and width are in the order needed for landscape

        fig = plt.figure(figsize=fig_size)

        # Then we go through all parameters which are to be tuned:
        for i, par in enumerate(params):
            # Add a subplot for each tuning parameter:
            ax1 = plt.subplot(len(params), 1, i+1)

            # Insert a second axis:
            ax2 = ax1.twinx()

            # The following retrieves the score-coefficients and produces a line plot (equal distances between x-ticks),
            # we also exclude NaNs from being displayed (this only concerns DBSCAN --> we assume that if there is a NaN
            # for Silhouette, there will also be one for Calinski-Harabasz --> this should be true if we think of the
            # try-statement in which both coefficients are computed in the DBSCAN-clustering
            sils = [x for x in self.scores_dict_['silhouette_avg'] if ~np.isnan(x)]
            chs = [x for i, x in enumerate(self.scores_dict_['calinski_harabasz_score']) if ~np.isnan(self.scores_dict_['silhouette_avg'][i])]
            idx = [x for i, x in enumerate(self.scores_dict_[par]) if ~np.isnan(self.scores_dict_['silhouette_avg'][i])]

            plt.xticks(range(len(idx)), idx)
            lns1 = ax1.plot(range(len(idx)), sils, color=sns.color_palette('viridis', 2)[0], label='Silhouette Avg.')
            plt.xticks(range(len(idx)), idx)
            lns2 = ax2.plot(range(len(idx)), chs, color=sns.color_palette('viridis', 2)[1], label='Calinski-Harabasz')

            # label the y-axes
            ax1.set_ylabel('Silhouette Coeff. (-)')
            ax2.set_ylabel('Calinski-Harabasz (-)')
            ax2.grid(False)
            ax1.set_xlabel('{}'.format(par))

            # Construct the legend:
            lns = lns1 + lns2
            labs = [l.get_label() for l in lns]
            ax1.legend(lns, labs, loc=0)

        plt.suptitle("{} - Clustering-Performance".format(saving['title']))

        return fig

    def _inertia_elbow_plot(self, saving):
        """
        This function returns an elbow plot of the inertia of k-means clustering.
        :param saving: dict with information where to store the data (see init-function)
        :return: figure with an elbow plot of the inertia of k-means clustering
        """

        # Set up the figure:
        fig_width_cm = 21  # A4 page
        fig_height_cm = 29.7
        inches_per_cm = 1 / 2.58  # Convert cm to inches
        fig_width = fig_width_cm * inches_per_cm  # width in inches
        fig_height = fig_height_cm * inches_per_cm  # height in inches
        fig_size = [fig_height, fig_width]  # height and width are in the order needed for landscape

        fig = plt.figure(figsize=fig_size)

        # Plot the inertia
        lns1 = plt.plot(self.scores_dict_['n_clusters'], self.scores_dict_['inertia'], color=sns.color_palette('viridis', 1)[0], label='Inertia')

        # label the y-axes
        plt.ylabel('Inertia')
        plt.xlabel('no. of clusters')

        # Construct the legend:
        plt.legend([l.get_label() for l in lns1], loc=0)

        plt.title("{} - Inertia-Elbow-Plot".format(saving['title']))

        return fig

    def make_score_dataframe(self):
        """
        This function converts the scores_dict to a pandas-dataframe
        :return: pandas dataframe of the scores-dict
        """
        return pd.DataFrame(data=self.scores_dict_)

    def hierarchical_deep_analysis(self, clusterer,
                                   dendro_params={'truncate_mode':'lastp', 'p':100, 'leaf_rotation':90.,
                                   'leaf_font_size':12., 'show_contracted':True, 'annotate_above':10, 'max_d':20},
                                   saving={'title': 'clustering_analysis', 'savepath': r"D:\froemelt\Desktop"}):
        """
        This function performs a more detailed analysis of hierarchical clustering. It produces a dendrogram as well as
        an acceleration plot. The ideas of the plots are taken from https://joernhees.de/blog/2015/08/26/scipy-hierarchical-clustering-and-dendrogram-tutorial/
        while the hacks to do the dendrogram with SKLEARN instead of scipy are coming from http://stackoverflow.com/questions/26851553/sklearn-agglomerative-clustering-linkage-matrix.
        The corresponding functions (SKLEARN-hacks) and dendrogram can be found in the module "Clustering_Tools".
        :param clusterer:  SKLEARN-clusterer-instance (see init-function)
        :param dendro_params: (optional), dict with parameters which shall be passed to the fancy_dendrogram-function
        :param saving: dict with information where to save the created plots
        """

        # Set up the PDF where to store the data
        pp = PdfPages(os.path.join(saving['savepath'], 'Deep_Analysis_Hierarchical-Clustering_{}.pdf'.format(saving['title'])))

        # Compute the linkage_matrix which is needed by scipy-dendrogram. The function linkage_matrix() can be found in
        # the "clustering-hack"-module "Clustering_Tools.py"
        # The linkage matrix has the following form: [idx1, idx2, dist, sample_count] --> indices from 0 to n_samples
        # correspond to the original samples, while above n_samples refers to idx - len(X); sample_count means the number of
        # total samples in the cluster
        linkmat = linkage_matrix(clusterer)

        # Plot a dendrogram based on https://joernhees.de/blog/2015/08/26/scipy-hierarchical-clustering-and-dendrogram-tutorial/.
        # The fancy_dendrogram-function can be found in the "clustering-hack"-module "Clustering_Tools.py"
        fig = fancy_dendrogram(linkmat, **dendro_params)

        # Save the dendrogram to the PDF
        fig.savefig(pp, format='pdf', orientation='landscape', papertype='a4')
        plt.close(fig)
        del fig

        # Set up the figure for the acceleration plot
        fig_width_cm = 21  # A4 page
        fig_height_cm = 29.7
        inches_per_cm = 1 / 2.58  # Convert cm to inches
        fig_width = fig_width_cm * inches_per_cm  # width in inches
        fig_height = fig_height_cm * inches_per_cm  # height in inches
        fig_size = [fig_height, fig_width]  # height and width are in the order needed for landscape

        fig = plt.figure(figsize=fig_size)
        ax = plt.gca()

        # Compute the cophenetic correlation which basically compares (correlates) the actual
        # pairwise distances of all your samples to those implied by the hierarchical clustering. The closer the value
        # is to 1, the better the clustering preserves the original distances.
        c, _ = cophenet(linkmat, pdist(self.clustering_data_))

        # Compute the inconsistency --> this is one option to retrieve the number of clusters in hierarchical clustering
        # However the blog of Jörn Hees does not recommend using it.
        incons = inconsistent(linkmat, d=5)
        incons_rev = incons[::-1]
        inc = incons_rev[:,3].argmax() + 2

        # Next, we retrieve the distances in the dendrogram from the linkage-matrix
        dists = linkmat[:, 2]
        dists_rev = dists[::-1]
        idxs = np.arange(1, len(dists) + 1)

        # And we plot the distances in reverse order
        plt.plot(idxs, dists_rev, color=sns.color_palette('viridis', 3)[0], label='Distances')

        # Then we take the second derivation (discrete difference) of the differences --> I think this is what Jörn
        # Hees suggests as acceleration (however, if I am not mistaken he does not really recommend using it) I can see
        # the idea behind, but I think that if we want to find the largest gap, we need to take the first order difference
        acceleration = np.diff(dists, 2)
        acceleration_rev = acceleration[::-1]

        # Plot the acceleration as well --> be careful with the indices
        plt.plot(idxs[:-2] + 1, acceleration_rev, color=sns.color_palette('viridis', 3)[1], label='Acceleration')
        k = acceleration_rev.argmax() + 2  # I think + 2 because we start index at 0 and we lose one index by differencing

        # In the following we also compute the gap according to our own ideas: Take the first order difference
        gaps = np.diff(dists, 1)
        gaps_rev = gaps[::-1]
        k2 = gaps_rev.argmax() + 2

        # Plot this "gaps" as well --> be careful with the indices
        plt.plot(idxs[:-1] + 1, gaps_rev, color=sns.color_palette('viridis', 3)[2], label='Gaps')

        # Give also some output to the console:
        txt = "cophenetic correlation coeff. {:.2f}\nno of clusters acc. to acceleration: {}\nno of clusters acc. to inconsistency: {}\nno of clusters acc. to gaps: {}".format(c,k, inc, k2)
        print(txt)

        # And add the same text to the plot
        at = AnchoredText(txt, loc=5, frameon=False,  prop=dict(size=10))
        ax.add_artist(at)
        plt.legend()

        # Save and close the figure
        fig.savefig(pp, format='pdf', orientation='landscape', papertype='a4')
        plt.close(fig)
        del fig
        pp.close()


class UMatrix_impr(smp.umatrix.UMatrixView):
    """
    This class is a hacked and enhanced version of SOMPY's UMatrixView. It is not only able to display the U-Matrix with
    different colormaps but also able to draw cluster-boundaries if a SKLEARN-clusterer is passed. PLEASE NOTE that this
    class assumes that wished cluster labels are already assigned to the SOM-class (som.cluster_labels).
    """
    def show_impr(self, som, distance2=1, row_normalized=False, show_data=False, contooor=True,
                  labels=False, cmap='coolwarm_r', savepath=None, show=False, figr=None,
                  clusters=None):
        """
        Compared to SOMPY's original UMatrixView, this function not only allows for showing the U-Matrix in different
        colormaps, but also allows for drawing the borders of clusters onto the U-Matrix.
        :param som: A trained SOMPY's SOM-class
        :param distance2: see SOMPY --> used to build the U-Matrix
        :param row_normalized: see SOMPY --> probably to compute U-Matrix without normalization
        :param show_data: if the original data shall be shown as scatter
        :param contooor: if contour lines shall be drawn ("Höhenlinien")
        :param labels: tested, but couldn't see a difference
        :param cmap: colormap to be used
        :param savepath: Where to save the U-Matrix --> please note that the file-extension determines the file-format.
        Instead of a path one can also pass a PdfPages-class
        :param show: True/False to indicate if the figure shall be shown
        :param figr: This is thought to pass an open figure --> e.g. if the U-Matrix shall be part of a subplot (see
        e.g. SOM_Clustering_Evaluator-class)
        :param clusters: dict with two keys: 'n_clusters': to indicate how many clusters need to be drawn and
        'clusterer': to pass the SKLEARN clusterer (e.g. as attribute from a Clustering_Analyzer-class instance)
        """
        umat = self.build_u_matrix(som, distance=distance2,
                                   row_normalized=row_normalized)
        msz = som.codebook.mapsize
        proj = som.project_data(som.data_raw)
        coord = som.bmu_ind_to_xy(proj)

        # If a figure is passed, we don't want to create new subplots
        if not figr:
            fig, ax = plt.subplots(1, 1)

        plt.imshow(umat, cmap=plt.cm.get_cmap(cmap), alpha=1)
        plt.axis('off')

        if contooor:
            mn = np.min(umat.flatten())
            mx = np.max(umat.flatten())
            std = np.std(umat.flatten())
            md = np.median(umat.flatten())
            mx = md + 0 * std
            plt.contour(umat, np.linspace(mn, mx, 15), linewidths=0.7,
                        cmap=plt.cm.get_cmap('Greys'))

        if show_data:
            plt.scatter(coord[:, 1], coord[:, 0], s=2, alpha=1., c='Gray',
                        marker='o', cmap='jet', linewidths=3, edgecolor='Gray')

        if labels:
            if labels is True:
                labels = som.build_data_labels()
            for label, x, y in zip(labels, coord[:, 1], coord[:, 0]):
                plt.annotate(str(label), xy=(x, y),
                             horizontalalignment='center',
                             verticalalignment='center')

        # The following lines are part of the original code, but were removed since we want to have different figure size

        #ratio = float(msz[0]) / (msz[0] + msz[1])
        #fig.set_size_inches((1 - ratio) * 15, ratio * 15)
        #plt.tight_layout()

        # Setting up the figure:
        if not figr:
            fig_width_cm = 21  # A4 page
            fig_height_cm = 29.7
            inches_per_cm = 1 / 2.58  # Convert cm to inches
            fig_width = fig_width_cm * inches_per_cm  # width in inches
            fig_height = fig_height_cm * inches_per_cm  # height in inches
            fig_size = [fig_height, fig_width]  # height and width are in the order needed for landscape
            fig.set_size_inches(fig_size[0], fig_size[1])

        plt.subplots_adjust(hspace=.00, wspace=.000)
        sel_points = list()

        # The following allows for plotting clusters in the U-Matrix. The code is partially copy/pasted from a SKLEARN-
        # Racoon-Face-Example

        if clusters:
            n_clusters = clusters['n_clusters']
            plt.title('Clusters: {}'.format(n_clusters))
            som_map = som.codebook.matrix[:, 0].reshape(som.codebook.mapsize[0], som.codebook.mapsize[1])
            label = np.reshape(clusters['clusterer'].labels_, som_map.shape)
            for l in set(clusters['clusterer'].labels_):
                plt.contour(label == l, contours=1, colors=[plt.cm.viridis(l / float(n_clusters)), ])

        # If the file-extension of the "savepath"-string determines the file-format. If "savepath" is a PdfPages-instance,
        # then the U-Matrix is added to the open PDF-File
        if savepath:
            if type(savepath)==PdfPages:
                fig.savefig(savepath, format='pdf', orientation='landscape', papertype='a4')
                plt.close(fig)
                del fig
            else:
                fig.savefig(savepath, format=savepath.split('.')[-1])

        # The U-Matrix is plotted to the screen if "show" is True
        elif show:
            plt.show()

        return sel_points, umat


class HitMapView_impr(smp.visualization.hitmap.HitMapView):
    """
    This class is a hacked and enhanced version of SOMPY's HitMapView. It is able to display the clusters with
    different colormaps and facilitates the saving to a file. This class is used to plot the labeled clusters on top
    of the SOM. PLEASE NOTE that this class assumes that wished cluster labels are already assigned to the SOM-class
    (som.cluster_labels).
    """

    def _set_labels(self, cents, ax, labels, lblsize):
        """
        Only one change compared to original code: lblsize (= size of labels) is introduced
        """
        for i, txt in enumerate(labels):
            ax.annotate(txt, (cents[i, 1], cents[i, 0]), size=lblsize, va="center", ha='center')

    def show(self, som, data=None, cmap='viridis', savepath=None, figr=None, show=False):
        """
        Compared to SOMPY's orginal HitMapView, this function not only allows for showing the clustering on top of the
        SOM in different colormaps, but also facilitates the saving
        :param som: A trained SOMPY's SOM-class
        :param data: see SOMPY
        :param cmap: colormap to be used
        :param savepath: Where to save the SOM-clustering-plot --> please note that the file-extension determines the file-format.
        :param figr: This is thought to pass an open figure --> e.g. if the SOM-clustering shall be part of a subplot (see
        e.g. SOM_Clustering_Evaluator-class)
        :param show: True/False to indicate if the figure shall be shown
        """

        try:
            codebook = getattr(som, 'cluster_labels')
        except:
            raise ValueError("Assign cluster labels first!")
            #codebook = som.cluster()

        # codebook = getattr(som, 'cluster_labels', som.cluster())
        msz = som.codebook.mapsize

        # If a figure is passed, we don't want to create new subplots
        if not figr:
            self.prepare()
            ax = self._fig.add_subplot(111)
        else:
            ax = plt.gca()

        if data:
            proj = som.project_data(data)
            cents = som.bmu_ind_to_xy(proj)
            self._set_labels(cents, ax, codebook[proj])

        else:
            cents = som.bmu_ind_to_xy(np.arange(0, msz[0] * msz[1]))
            self._set_labels(cents, ax, codebook, 6)

        plt.imshow(codebook.reshape(msz[0], msz[1])[::], alpha=0.5, cmap=cmap)
        plt.axis('off')

        if savepath:
            self._fig.savefig(savepath, format=savepath.split('.')[-1])

        if show:
            plt.show()

        return cents


class BmuHitsView_impr(BmuHitsView):
    """
    This class is a hacked and enhanced version of SOMPY's BmuHitsView. It is able to display the hits in the clusters
    with different colormaps and facilitates the saving to a file.
    PLEASE NOTE that this class assumes that wished cluster labels are already assigned to
    the SOM-class (som.cluster_labels).
    """
    def _set_labels_new(self, cents, ax, labels, onlyzeros, fontsize):
        """
        Actually, only "aesthetical" changes
        """
        for i, txt in enumerate(labels):
            if onlyzeros == True:
                if txt > 0:
                    txt = ""
            ax.annotate(txt, (cents[i, 1], cents[i, 0]), va="center", ha='center', size=fontsize)


    def show_impr(self, som, anotate=True, onlyzeros=False, labelsize=7, cmap="jet", logaritmic=False, savepath=None, show=False):
        """
        Compared to SOMPY, this function is able to plot the hits within the cluster-boundaries.
        :param som: A trained SOMPY's SOM-class
        :param anotate: see SOMPY
        :param onlyzeros: see SOMPY
        :param labelsize: see SOMPY
        :param cmap: colormap to be used
        :param logaritmic: see SOMPY
        :param savepath: Where to save the Hits-Map --> please note that the file-extension determines the file-format.
        Instead of a path one can also pass a PdfPages-class
        :param show: True/False to indicate if the figure shall be shown
        """
        (self.width, self.height, indtoshow, no_row_in_plot, no_col_in_plot,
         axis_num) = self._calculate_figure_params(som, 1, 1)

        codebook = getattr(som, 'cluster_labels')

        self.prepare()
        ax = plt.gca()
        counts = Counter(som._bmu[0])
        counts = [counts.get(x, 0) for x in range(som.codebook.mapsize[0] * som.codebook.mapsize[1])]
        mp = np.array(counts).reshape(som.codebook.mapsize[0],
                                      som.codebook.mapsize[1])

        #Adjusting the figure size
        fig_width_cm = 21  # A4 page
        fig_height_cm = 29.7
        inches_per_cm = 1 / 2.58  # Convert cm to inches
        fig_width = fig_width_cm * inches_per_cm  # width in inches
        fig_height = fig_height_cm * inches_per_cm  # height in inches
        fig_size = [fig_height, fig_width]  # height and width are in the order needed for landscape
        self._fig.set_size_inches(fig_size[0], fig_size[1])

        if not logaritmic:
            norm = matplotlib.colors.Normalize(
                vmin=0,
                vmax=np.max(mp.flatten()),
                clip=True)
        else:
            norm = matplotlib.colors.LogNorm(
                vmin=1,
                vmax=np.max(mp.flatten()))

        msz = som.codebook.mapsize

        cents = som.bmu_ind_to_xy(np.arange(0, msz[0] * msz[1]))

        if anotate:
            self._set_labels_new(cents, ax, counts, onlyzeros, labelsize)

        plt.imshow(codebook.reshape(msz[0], msz[1])[::], alpha=0.7, cmap=cmap)

        # pl = plt.pcolor(mp[::-1], norm=norm, cmap=cmap)

        # plt.axis([0, som.codebook.mapsize[1], 0, som.codebook.mapsize[0]])
        ax.set_yticklabels([])
        ax.set_xticklabels([])
        # plt.colorbar(pl)

        # If the file-extension of the "savepath"-string determines the file-format. If "savepath" is a PdfPages-instance,
        # then the Hits-Map is added to the open PDF-File
        if savepath:
            if type(savepath) == PdfPages:
                self._fig.savefig(savepath, format='pdf', orientation='landscape', papertype='a4')
                plt.close(self._fig)
            else:
                self._fig.savefig(savepath, format=savepath.split('.')[-1])

        # The Hit-Map is plotted to the screen if "show" is True
        elif show:
            plt.show()

    def show(self, som, anotate=True, onlyzeros=False, labelsize=7, cmap="jet", logaritmic = False, savepath=None, show=False):
        """
        There are only minor changes to the original SOMPY-function: it shall facilitate saving
        :param som: A trained SOMPY's SOM-class
        :param anotate: see SOMPY
        :param onlyzeros: see SOMPY
        :param labelsize: see SOMPY
        :param cmap: colormap to be used
        :param logaritmic: see SOMPY
        :param savepath: Where to save the Hits-Map --> please note that the file-extension determines the file-format.
        Instead of a path one can also pass a PdfPages-class
        :param show: True/False to indicate if the figure shall be shown
        """
        (self.width, self.height, indtoshow, no_row_in_plot, no_col_in_plot,
         axis_num) = self._calculate_figure_params(som, 1, 1)

        self.prepare()
        ax = plt.gca()
        counts = Counter(som._bmu[0])
        counts = [counts.get(x, 0) for x in range(som.codebook.mapsize[0] * som.codebook.mapsize[1])]
        mp = np.array(counts).reshape(som.codebook.mapsize[0],
                                      som.codebook.mapsize[1])


        if not logaritmic:
            norm = matplotlib.colors.Normalize(
                vmin=0,
                vmax=np.max(mp.flatten()),
                clip=True)
        else:
            norm = matplotlib.colors.LogNorm(
                vmin=1,
                vmax=np.max(mp.flatten()))

        msz = som.codebook.mapsize

        cents = som.bmu_ind_to_xy(np.arange(0, msz[0] * msz[1]))

        if anotate:
            self._set_labels(cents, ax, counts, onlyzeros, labelsize)


        pl = plt.pcolor(mp[::-1], norm=norm, cmap=cmap)

        plt.axis([0, som.codebook.mapsize[1], 0, som.codebook.mapsize[0]])
        ax.set_yticklabels([])
        ax.set_xticklabels([])
        plt.colorbar(pl)

        # If the file-extension of the "savepath"-string determines the file-format. If "savepath" is a PdfPages-instance,
        # then the Hits-Map is added to the open PDF-File
        if savepath:
            if type(savepath) == PdfPages:
                self._fig.savefig(savepath, format='pdf', orientation='landscape', papertype='a4')
                plt.close(self._fig)
            else:
                self._fig.savefig(savepath, format=savepath.split('.')[-1])
                plt.close(self._fig)

        # The Hit-Map is plotted to the screen if "show" is True
        elif show:
            plt.show()


class SOM_Clustering_Evaluator(object):
    """
    This class shall help to evaluate clustering which was applied on top of a Self-Organizing Map. It computes different
    statistics and contains methods to save the statistics to EXCEL or to create PDFs with a comprehensive summary of the
    evaluation.
    """

    def __init__(self, som, clusterer, hh_data, conn):
        """
        Init-function for the SOM_Clustering_Evaluator-class.
        :param som: A trained SOMPY's SOM-class
        :param clusterer: SKLEARN-clustering-class instance (e.g. can be passed from an attribute of a Clustering_Analyzer-instance)
        :param hh_data: SKLData_SOM-class with the data that was used to create the SOM (see parameter "som")
        :param conn: connection to the Postgres-DB
        """

        self.som_ = som
        self.clusterer_ = clusterer
        self.hh_data_ = hh_data

        # Find best matching units for all HHs
        self._bmus_ = som.find_bmu(som._data, nth=1)[0]

        # Add the cluster labels to the neurons (indicating which neuron belongs to which cluster) and also create
        # a list out of the labels
        self.som_.cluster_labels = clusterer.labels_
        self._cluster_list_ = list(clusterer.labels_)

        # Retrieve the survey month of the households from the PGDB
        cur = conn.cursor(cursor_factory=pge.RealDictCursor)
        self.hh_months_dict_ = self._get_hh_months(cur)

        # Compute the statistics for the clusters and create a dataframe
        self.clusterstats_dict_ = self._get_clusterstats()
        self.clusterstats_df_ = pd.DataFrame.from_dict(self.clusterstats_dict_, orient='index')

        # Do ANOVA-analysis and create a dataframe
        self.anova_dict_ = self._do_anova()
        self.anova_df_ = pd.DataFrame.from_dict(self.anova_dict_, orient='index')

        # To pickle the class, the conn-object needs to be removed.
        try:
            del self.hh_data_._conn_
        except:
            pass


    def _get_hh_months(self, cur):
        """
        This function retrieves the survey month for the households
        :param cur: RealDict-Cursor
        :return: dict with household-ids as keys and month-number as values
        """

        query = """
        SELECT haushaltid, month_no FROM working_tables.habe_hh_month
        """
        cur.execute(query)
        hh_months = cur.fetchall()
        cur.close()
        del cur

        return {hh['haushaltid']: hh['month_no'] for hh in hh_months}


    def _get_clusterstats(self):
        """
        This function conducts statistics of the clusters under consideration
        :return: dict with clustering statistics
        """

        clusterstats = {}

        # Create a results-container first:
        for cl in set(self._cluster_list_):
            clusterstats.update({cl: {'counts': 0, 'months': set(), 'month_count': 0}})

        # Go through all households
        for hhidx, hhid in enumerate(self.hh_data_.meta_):

            # With the help of household-index, we can find the BMU (=neuron) and via BMU the cluster the household belongs to
            hhclust = self._cluster_list_[int(self._bmus_[hhidx])]
            clusterstats[hhclust]['counts'] += 1

            # Add the survey month of the household to the set of months
            clusterstats[hhclust]['months'].update({self.hh_months_dict_[hhid]})

        # Finally we iterate again over all clusters and determine the number of different survey months included
        # within a cluster
        for cl in clusterstats.keys():
            clusterstats[cl]['month_count'] = len(clusterstats[cl]['months'])

        return clusterstats


    def _do_anova(self):
        """
        This function performs ANOVA for each attribute/feature that was part of the clustering
        :return: dict with ANOVA-results
        """

        anova_dict = {}

        # In a first step we construct a list indicating for each household to which cluster it belongs to.
        hh_clusters = [self._cluster_list_[int(self._bmus_[hhidx])] for hhidx, _ in enumerate(self.hh_data_.meta_)]

        # Then we iterate over all attributes which were used to build the clusters:
        for ia, a in enumerate(self.hh_data_.attributes_):
            print("Do ANOVA for {} of {}".format(ia + 1, len(self.hh_data_.attributes_)), end='\r')  # interesting: the end=\r allows for overwriting the print

            # We first construct a name of the attribute under consideration. Please note that quantities/amounts are
            # start with a "Q"
            nm = "Q: {}".format(self.hh_data_.varnames_[a]) if 'm' in a and not a.startswith('char') else self.hh_data_.varnames_[a]

            cls_raw = []
            cls_scl = []
            cls_bmu = []

            # Then we go through all clusters retrieve the values of the attribute under consideration for each cluster
            # and store them array-wise in a list. PLEASE NOTE: we are doing three different ANOVAs: we look at the
            # raw data (not scaled, household level), the scaled data (used to perform the SOM, household level) and
            # the neurons (scaled data, neuron level, used to perform the clustering)
            for cl in set(self._cluster_list_):
                cls_raw.append(self.hh_data_.data_[hh_clusters == cl, ia])
                cls_scl.append(self.som_._data[hh_clusters == cl, ia])
                cls_bmu.append(self.som_.codebook.matrix[self.som_.cluster_labels==cl,ia])

            # Core: perform ANOVA
            Fr_val, pr_val = stats.f_oneway(*cls_raw)
            Fs_val, ps_val = stats.f_oneway(*cls_scl)
            Fb_val, pb_val = stats.f_oneway(*cls_bmu)

            anova_dict[nm] = {'Fr': Fr_val, 'pr': pr_val, 'Fs': Fs_val, 'ps': ps_val, 'Fb': Fb_val, 'pb': pb_val}

        return anova_dict


    def save2excel(self, savepath):
        """
        This function stores the full results of ANOVA and the full clustering statistics to an excel-file specified by
        "savepath"
        :param savepath: string: where to store the excel-file
        """

        self.clusterstats_df_.to_excel(savepath, sheet_name='Cluster_Statistics', na_rep='-')

        # In order to insert a second sheet, we need to reload the workbook
        workbook = load_workbook(savepath)
        writer = pd.ExcelWriter(savepath, engine='openpyxl')
        writer.book = workbook

        self.anova_df_.to_excel(writer, sheet_name='ANOVA', na_rep='-')
        writer.save()


    def create_plot(self, savepath, avg_silhouette=None):
        """
        This function creates a visual overview of the clustering evaluation. The plot consists of four elements: 1. U-
        Matrix with cluster-borders, 2. Clustering-Map with labels where to find which cluster, 3. Histogram for counts-
        statistics (how many households per cluster), 4. Table with most important statistics about clustering and ANOVA.
        :param savepath: string: where to store the plot. Instead of a path one can also pass a PdfPages-class.
        :param avg_silhouette: (optional) total average silhouette score in order to include in the PDF.
        """

        # Set up the figure
        fig_width_cm = 21  # A4 page
        fig_height_cm = 29.7
        inches_per_cm = 1 / 2.58  # Convert cm to inches
        fig_width = fig_width_cm * inches_per_cm  # width in inches
        fig_height = fig_height_cm * inches_per_cm  # height in inches
        fig_size = [fig_width, fig_height]  # height and width are in the order needed for portrait

        fig = plt.figure(figsize=fig_size)
        gridspec.GridSpec(3, 2)
        plt.suptitle("No. of clusters: {}, Silhouette: {:.2f}".format(len(set(self._cluster_list_)), avg_silhouette or '-'))

        # First subplot contains the U-Matrix with the cluster-borders
        plt.subplot2grid((3, 2), (0, 0), colspan=2, rowspan=1)
        u = UMatrix_impr(50, 50, 'umatrix', show_axis=False, text_size=8, show_text=False)
        _ = u.show_impr(self.som_, distance2=1, row_normalized=False, show_data=False, contooor=True, labels=False, figr=fig,
                        cmap='coolwarm_r', clusters={'n_clusters': len(set(self._cluster_list_)), 'clusterer': self.clusterer_})

        # Second subplot shows the labeled clusters on top of the SOM
        plt.subplot2grid((3, 2), (1, 0), colspan=2, rowspan=1)
        hits = HitMapView_impr(50, 50, " ", text_size=8)
        _ = hits.show(self.som_, cmap='nipy_spectral', figr=fig)

        # Third subplot depicts a histogram with the no. of households per cluster
        plt.subplot2grid((3, 2), (2, 0), colspan=1, rowspan=1)
        sns.distplot(self.clusterstats_df_['counts'])
        plt.xlim(0, max(self.clusterstats_df_['counts']))

        # Finally, we also plot an overview-table with the most important stats about ANOVA and clustering:
        plt.subplot2grid((3, 2), (2, 1), colspan=1, rowspan=1)

        # In the following we retrieve some "helper"-dataframes: clusters with less than all 12 survey months, attributes
        # with p-values above 5% and this not only for the raw data but also for the neurons.
        clusters_wo_all_months = self.clusterstats_df_.loc[self.clusterstats_df_['month_count']<12, 'month_count']
        big_praw = self.anova_df_.loc[self.anova_df_['pr'] > 0.05, 'pr']
        big_pbmu = self.anova_df_.loc[self.anova_df_['pb'] > 0.05, 'pb']

        # The stats-dict computes all the important statistics which shall be displayed. We especially would like to
        # mention that we also compute how many (absolute and percent) clusters have less than 12 survey months included
        # and how many % of the attributes have p-values > 5%
        stats_dict = {
            'Min': {
                'No of HH': self.clusterstats_df_['counts'].min(),
                'F_raw': round(self.anova_df_['Fr'].min(), 2),
                'p_raw': round(self.anova_df_['pr'].min(), 2),
                'F_bmu': round(self.anova_df_['Fb'].min(), 2),
                'p_bmu': round(self.anova_df_['pb'].min(), 2)},
            'Max': {
                'No of HH': self.clusterstats_df_['counts'].max(),
                'F_raw': round(self.anova_df_['Fr'].max(), 2),
                'p_raw': round(self.anova_df_['pr'].max(), 2),
                'F_bmu': round(self.anova_df_['Fb'].max(), 2),
                'p_bmu': round(self.anova_df_['pb'].max(), 2)},
            'Mean': {
                'No of HH': round(self.clusterstats_df_['counts'].mean(),2),
                'F_raw': round(self.anova_df_['Fr'].mean(), 2),
                'p_raw': round(self.anova_df_['pr'].mean(), 2),
                'F_bmu': round(self.anova_df_['Fb'].mean(), 2),
                'p_bmu': round(self.anova_df_['pb'].mean(), 2)},
            'Median': {
                'No of HH': self.clusterstats_df_['counts'].median(),
                'F_raw': round(self.anova_df_['Fr'].median(), 2),
                'p_raw': round(self.anova_df_['pr'].median(), 2),
                'F_bmu': round(self.anova_df_['Fb'].median(), 2),
                'p_bmu': round(self.anova_df_['pb'].median(), 2)},
            'No./% Cl\n< 12 months': {
                'No of HH': "{} / {}".format(len(clusters_wo_all_months), round(len(clusters_wo_all_months)/len(set(self._cluster_list_)) * 100, 2)),
                'F_raw': "-",
                'p_raw': "-",
                'F_bmu': "-",
                'p_bmu': "-"},
            'Min months': {
                'No of HH': clusters_wo_all_months.min(),
                'F_raw': "-",
                'p_raw': "-",
                'F_bmu': "-",
                'p_bmu': "-"},
            'Perc.\np > 0.05': {
                'No of HH': "-",
                'F_raw': "-",
                'p_raw': len(big_praw) / len(self.anova_df_['pr']) * 100,
                'F_bmu': "-",
                'p_bmu': len(big_pbmu) / len(self.anova_df_['pb']) * 100,}
        }

        stats_df = pd.DataFrame.from_dict(stats_dict, orient='index')

        # create the table
        plt.axis('off')
        tbl = plt.table(cellText=stats_df.values, colWidths=[0.25] * len(stats_df.columns),
                  rowLabels=stats_df.index,
                  colLabels=stats_df.columns,
                  cellLoc='center', rowLoc='center',
                  loc='center', bbox=[0.3, 0, 0.9, 1])
        tbl.set_fontsize(8)
        # cellDict = tbl.get_celld()
        # for i in range(len(stats_df.index)):
        #     cellDict[(i+1, -1)].set_width(2.)
        #tbl.scale(1.5, 1.5)

        fig.savefig(savepath, format='pdf', orientation='portrait', papertype='a4')
        plt.close(fig)
        del fig


def get_microcensus_data(excel):
    """
    This function reads the microcensus data from the consumption-cockpit-xlsx and returns a comprehensive dict. We will
    estimate pt/bike demand based on household size, income, age of the members and gender.
    :param excel: dict (keys: 'path' and 'sheet') with information where the microcensus data can be found
    :return: dict containing the relevant information for estimating pt/bike-demand
    """

    # We first open the excel-sheet which contains the microcensus data:
    wb = xlrd.open_workbook(excel['path'])
    ws = wb.sheet_by_name(excel['sheet'])

    # List of the transport modes:
    transportmodes = ['train', 'coach', 'tram', 'bike']

    # Look for the cell where we want to start reading the information:
    inicol = ws.col_values(0)
    headerrow = ws.row_values(inicol.index('py_start_read'))  # ATTENTION: we need to have a cell where 'py_start_read' is contained

    # We then read all the columns that contain the information about the driven kilometers by pt/bike
    traincol = ws.col_values(headerrow.index('Tagesdistanz Bahn, in km'))
    coachcol = ws.col_values(headerrow.index('Tagesdistanz Postauto, in km'))
    tramcol = ws.col_values(headerrow.index('Tagesdistanz Tram und Bus, in km'))
    bikecol = ws.col_values(headerrow.index('Tagesdistanz Velo, in km'))
    cols = [traincol, coachcol, tramcol, bikecol]

    # We will determine pt/bike demand based on household size, income, age of the members and gender. For this we
    # initialize separate dicts:
    hhsize = {}
    hhincome = {}
    hhage = {}
    hhsex = {}

    # We now construct dicts with all the information contained in the microcensus-table
    sizerow = inicol.index('Haushaltsgrösse')
    for i, size in enumerate([1, 2, 3, 4, 5]):
        for j, mode in enumerate(transportmodes):
            hhsize.update({(size, mode): cols[j][sizerow + i]})

    incomerow = inicol.index('Monatliches \nHaushaltseinkommen')
    for i, money in enumerate([2000, 6000, 10000, 14000, 14001]):
        for j, mode in enumerate(transportmodes):
            hhincome.update({(money, mode): cols[j][incomerow + 1 + i]})

    agerow = inicol.index('Alter')
    for i, age in enumerate([17, 24, 44, 64, 65]):
        for j, mode in enumerate(transportmodes):
            hhage.update({(age, mode): cols[j][agerow + i]})

    sexrow = inicol.index('Geschlecht')
    for i, sex in enumerate(['male', 'fem']):
        for j, mode in enumerate(transportmodes):
            hhsex.update({(sex, mode): cols[j][sexrow + i]})

    return {'hhsize': hhsize, 'hhincome': hhincome, 'hhage': hhage, 'hhsex': hhsex}


def estimate_habe_pt_bike(habe_hh, mc_data):
    """
    This function estimates the pt/bike demand for HABE-households based on microcensus-data and based on household
    size, income, age of the members and gender. ATTENTION: we get four different estimates based on these attributes,
    and we take the average of all of these estimates in the end (THIS IS AN ASSUMPTION!)
    :param habe_hh: dict of attributes of an individual household of the HABE. Needed are: 'e_bruttoeink', 'char_nopers'
    and all the male/female-statistics-attributes.
    :param mc_data: dict containing the relevant information for estimating pt/bike-demand (obtainable by get_microcensus_data)
    :return: Estimate of driven kilometers per household and day by train, coach, tram, bike
    """

    # In a first step we compute estimate the pt/bike demands by household size and directly multiply by the household
    # size to get an estimate "per household":
    size = habe_hh['char_nopers'] if habe_hh['char_nopers'] <= 5 else 5
    train_sizeestimate = mc_data['hhsize'][(size, 'train')] * habe_hh['char_nopers']
    coach_sizeestimate = mc_data['hhsize'][(size, 'coach')] * habe_hh['char_nopers']
    tram_sizeestimate = mc_data['hhsize'][(size, 'tram')] * habe_hh['char_nopers']
    bike_sizeestimate = mc_data['hhsize'][(size, 'bike')] * habe_hh['char_nopers']

    # The next estimate is based on income; again, we directly multiply the pkms with the household size
    inc = habe_hh['e_bruttoeink']
    if inc <= 2000:
        inc = 2000
    elif inc <= 6000:
        inc = 6000
    elif inc <= 10000:
        inc = 10000
    elif inc <= 14000:
        inc = 14000
    else:
        inc = 14001
    train_incestimate = mc_data['hhincome'][(inc, 'train')] * habe_hh['char_nopers']
    coach_incestimate = mc_data['hhincome'][(inc, 'coach')] * habe_hh['char_nopers']
    tram_incestimate = mc_data['hhincome'][(inc, 'tram')] * habe_hh['char_nopers']
    bike_incestimate = mc_data['hhincome'][(inc, 'bike')] * habe_hh['char_nopers']

    # Before we can determine bike/pt-demand by age and gender, we need to perform some statistics:
    females = 0
    males = 0
    age_24 = 0
    age_44 = 0
    age_64 = 0
    age_65 = 0
    for attr in habe_hh.keys():
        if '0004' in attr:
            continue
        if 'fem' in attr:
            females += habe_hh[attr]
        elif 'male' in attr:
            males += habe_hh[attr]
        if '0514' in attr or '1524' in attr:
            age_24 += habe_hh[attr]
        elif '2534' in attr or '3544' in attr:
            age_44 += habe_hh[attr]
        elif '4554' in attr or '5564' in attr:
            age_64 += habe_hh[attr]
        elif '6574' in attr or '7599' in attr:
            age_65 += habe_hh[attr]

    # In order to estimate bike/pt based on age, we multiply the age-statistics with the daily distances in order
    # to directly get the estimate "per household"
    train_ageestimate = (mc_data['hhage'][(17, 'train')] + mc_data['hhage'][(24, 'train')]) / 2 * age_24 + \
                        mc_data['hhage'][(44, 'train')] * age_44 + mc_data['hhage'][(64, 'train')] * age_64 + \
                        mc_data['hhage'][(65, 'train')] * age_65
    coach_ageestimate = (mc_data['hhage'][(17, 'coach')] + mc_data['hhage'][(24, 'coach')]) / 2 * age_24 + \
                        mc_data['hhage'][(44, 'coach')] * age_44 + mc_data['hhage'][(64, 'coach')] * age_64 + \
                        mc_data['hhage'][(65, 'coach')] * age_65
    tram_ageestimate = (mc_data['hhage'][(17, 'tram')] + mc_data['hhage'][(24, 'tram')]) / 2 * age_24 + \
                       mc_data['hhage'][(44, 'tram')] * age_44 + mc_data['hhage'][(64, 'tram')] * age_64 + \
                       mc_data['hhage'][(65, 'tram')] * age_65
    bike_ageestimate = (mc_data['hhage'][(17, 'bike')] + mc_data['hhage'][(24, 'bike')]) / 2 * age_24 + \
                       mc_data['hhage'][(44, 'bike')] * age_44 + mc_data['hhage'][(64, 'bike')] * age_64 + \
                       mc_data['hhage'][(65, 'bike')] * age_65

    # Similar to age, we also multiply the genders-statistics directly with the estimated daily distances
    train_sexestimate = mc_data['hhsex'][('male', 'train')] * males + mc_data['hhsex'][('fem', 'train')] * females
    coach_sexestimate = mc_data['hhsex'][('male', 'coach')] * males + mc_data['hhsex'][('fem', 'coach')] * females
    tram_sexestimate = mc_data['hhsex'][('male', 'tram')] * males + mc_data['hhsex'][('fem', 'tram')] * females
    bike_sexestimate = mc_data['hhsex'][('male', 'bike')] * males + mc_data['hhsex'][('fem', 'bike')] * females

    # Finally, we take the average of all estimates (this is an ASSUMPTION!)
    train = (train_sizeestimate + train_incestimate + train_ageestimate + train_sexestimate) / 4
    coach = (coach_sizeestimate + coach_incestimate + coach_ageestimate + coach_sexestimate) / 4
    tram = (tram_sizeestimate + tram_incestimate + tram_ageestimate + tram_sexestimate) / 4
    bike = (bike_sizeestimate + bike_incestimate + bike_ageestimate + bike_sexestimate) / 4

    return train, coach, tram, bike


def write_habe_pt_bike_to_pg(querylist, mcexcel):
    """
    This function estimates the public transport and bike demand for the HABE-households based on the microcensus data.
    This is basically a workaround since the kilometers driven by bike and public transport cannot be derived from HABE
    directly. The relevant microcensus data is stored in the Consumption_Cockpit.xlsx. PLEASE NOTE: we determine the
    bike/pt-demand based on size, income, age- and gender-statistics. In the end, we take a simple average of all four
    estimates. Also note that the results are given in km per household per month!
    :param querylist: list of strings with names of attributes that are needed to estimate the public transport and
    bike demand. Needed are: 'e_bruttoeink', 'char_nopers' and all the male/female-statistics-attributes.
    :param mcexcel: dict (keys: 'path' and 'sheet') with information where the microcensus data can be found
    :return: PLEASE NOTE: the results are directly written to the PG-Database
    """

    # In a first step we retrieve the necessary information for estimating public transportation and bike demand from
    # the HABE-households
    conn = get_pg_connection()
    query = """
    SELECT haushaltid, {} FROM working_tables.habe_hh_prepared_imputed
    """.format(', '.join(querylist))
    cur = conn.cursor(cursor_factory=pge.RealDictCursor)
    cur.execute(query)
    habe_hhs_sql = cur.fetchall()
    cur.close()

    # We then read the microcensus data from the indicated excel-file and get a comprehensive dict.
    mc_data = get_microcensus_data(mcexcel)

    # Create a results container:
    habe_hhs_pt_bike_list = []

    # Core: go through all HABE-households and estimate the public transport and bike demand and append it to the results-
    # list which will be written to the PG-Database
    for hh in habe_hhs_sql:
        train, coach, tram, bike = estimate_habe_pt_bike(hh, mc_data)
        habe_hhs_pt_bike_list.append({'haushaltid': hh['haushaltid'], 'mx6213': bike * 365. / 12.,
                                      'mx622101': train * 365. / 12.,'mx622102': tram * 365. / 12.,
                                      'mx622201': coach * 365. / 12.})

    # Create a new PG-Table to store the bike/pt-demand temporarily
    writecur = conn.cursor()
    query = """
    DROP TABLE IF EXISTS working_tables.habe_pt_bike_temptable
    """
    writecur.execute(query)
    conn.commit()

    query = """
    CREATE TABLE working_tables.habe_pt_bike_temptable
    (haushaltid bigint, mx6213 float, mx622101 float, mx622102 float, mx622201 float,
    CONSTRAINT habeptbike_pkey PRIMARY KEY (haushaltid))
    """
    writecur.execute(query)
    conn.commit()

    # Write the results to the new PG-Table and create an index
    cols = ['haushaltid', 'mx6213', 'mx622101', 'mx622102', 'mx622201']
    query = """
    INSERT INTO working_tables.habe_pt_bike_temptable(%s)
    VALUES (%s(%s)s)
    """ % (', '.join(cols), '%', ')s, %('.join(cols))
    writecur.executemany(query, habe_hhs_pt_bike_list)
    conn.commit()

    writecur.execute("""
    CREATE INDEX habe_ptbike_hhid
    ON working_tables.habe_pt_bike_temptable
    USING btree
    (haushaltid);
    """)
    conn.commit()

    writecur.close()
    conn.close()


def write_archetypes_pt_bike_to_pg():
    """
    This function determines the pt-bike demand of the archetypes based on the estimates for the HABE-households. To store the
    data, a new column is inserted in the PG-archetype-table.
    :return: PLEASE NOTE that the results are directly written to the PG-Database
    """

    # Get a connection to the PG-Database
    conn = get_pg_connection()

    # In a first step, we insert new columns in the archetypes-tables
    writecur = conn.cursor()
    query = """
    ALTER TABLE working_tables.habe_archetypes_weighted
    ADD COLUMN mx6213 float,
    ADD COLUMN mx622101 float,
    ADD COLUMN mx622102 float,
    ADD COLUMN mx622201 float
    """
    writecur.execute(query)
    conn.commit()

    writecur = conn.cursor()
    query = """
        ALTER TABLE working_tables.habe_archetypes_notweighted
        ADD COLUMN mx6213 float,
        ADD COLUMN mx622101 float,
        ADD COLUMN mx622102 float,
        ADD COLUMN mx622201 float
        """
    writecur.execute(query)
    conn.commit()

    # We then extract the estimates for the HABE-households and weight it with the HABE-weightings
    cur = conn.cursor(cursor_factory=pge.RealDictCursor)
    query = """
    SELECT clust.cluster_label_name, SUM(habe.mx6213*std.gewicht10_091011)/SUM(std.gewicht10_091011) AS mx6213,
    SUM(habe.mx622101*std.gewicht10_091011)/SUM(std.gewicht10_091011) AS mx622101,
    SUM(habe.mx622102*std.gewicht10_091011)/SUM(std.gewicht10_091011) AS mx622102,
    SUM(habe.mx622201*std.gewicht10_091011)/SUM(std.gewicht10_091011) AS mx622201
    FROM working_tables.habe_pt_bike_temptable habe LEFT JOIN working_tables.habe_clustering clust ON
    habe.haushaltid=clust.haushaltid LEFT JOIN original_data.habe_standard std ON habe.haushaltid=std.haushaltid
    GROUP BY clust.cluster_label_name
    """
    cur.execute(query)
    weighted_ptbike = cur.fetchall()

    # And we finally write the results to the archetypes-table
    for cl in weighted_ptbike:
        query = """
        UPDATE working_tables.habe_archetypes_weighted
        SET mx6213 = {},
            mx622101 = {},
            mx622102 = {},
            mx622201 = {}
        WHERE cluster_label_name = '{}'
        """.format(cl['mx6213'], cl['mx622101'], cl['mx622102'], cl['mx622201'], cl['cluster_label_name'])
        writecur.execute(query)
        conn.commit()

    # And finally we do the same for the non-weighted archetypes:
    query = """
        SELECT clust.cluster_label_name, AVG(habe.mx6213) AS mx6213,
        AVG(habe.mx622101) AS mx622101,
        AVG(habe.mx622102) AS mx622102,
        AVG(habe.mx622201) AS mx622201
        FROM working_tables.habe_pt_bike_temptable habe LEFT JOIN working_tables.habe_clustering clust ON
        habe.haushaltid=clust.haushaltid
        GROUP BY clust.cluster_label_name
        """
    cur.execute(query)
    notweighted_ptbike = cur.fetchall()

    for cl in notweighted_ptbike:
        query = """
            UPDATE working_tables.habe_archetypes_notweighted
            SET mx6213 = {},
                mx622101 = {},
                mx622102 = {},
                mx622201 = {}
            WHERE cluster_label_name = '{}'
            """.format(cl['mx6213'], cl['mx622101'], cl['mx622102'], cl['mx622201'], cl['cluster_label_name'])
        writecur.execute(query)
        conn.commit()

    writecur.close()
    cur.close()
    conn.close()


class ConsumptionLCA(object):
    """
    This class performs LCA for all functional units of the process models which were defined to represent the consumption
    categories (see consumption-cockpit.xlsx). The idea is to speed up computations by providing "characterization"
    factors for each functional unit which can then be multiplied with the actual demand of consumption Archetypes or
    STATPOP-households.
    """
    def __init__(self, ei_lca, ex_lca, methods, excel, archetypes=False, **kwargs):
        """
        Init-function of the ConsumptionLCA-class
        :param ei_lca: Brightway2-LCA-class for the Brightway-project containing ecoinvent-based databases
        :param ex_lca: Brightway2-LCA-class for the Brightway-project containing EXIOBASE activities
        :param methods: dict with methods for which multiplication factors shall be computed.
        :param excel: dict with information of the excel-file in which the LCA-Modelling can be found. Keys needed are:
        'path' and 'sheet', optionally also 'wwtp', 'heating' if the class is applied to STATPOP-households and if municipality-
        specific scores for wastewater treatments shall be computed.
        :param archetypes: (optional) True/False to indicate if the class is applied to get multiplication factors for
        consumption archetypes of for STATPOP-households.
        """

        self.methods_ = methods

        # In case this class is applied for STATPOP-households, we will extract specific scores per municipality for
        # wastewater treatment plants, petrol car fleets and diesel car fleets.
        if not archetypes:
            self.wwtp_bfsnr_dict_ = self._get_wwtp_bfsnr_dict(excel)

            # Not yet implemented since not really necessary since we merge consumption model directly with MATSim-based mobility model
            #self.petrolfleet_bfsnr_dict_, self.dieselfleet_bfsnr_dict_ = self._get_cantonal_fleets()

            # Read the functional units or the process models respectively from the Excel-file
            self._read_fu_from_excel(excel, ei_lca=ei_lca)
        else:
            # Read the functional units or the process models respectively from the Excel-file
            self._read_fu_from_excel(excel, ei_lca=None)

        # The following attribute will collect all the functional units based on EXIOBASE. This attribute is actually
        # only used for documentation and possible post-analyses.
        self.exiobase_fus_ = []

        # The do_lca-function is called to compute the "characterization" factors.
        self.lca_fu_scores_ = self._do_lca(ei_lca, ex_lca)


    def _get_wwtp_bfsnr_dict(self, excel):
        """
        This function constructs a dict to translate municipality-numbers to the WWTP-size based on the information given
        in an excel-sheet.
        :param excel: Dict with information where to find information on the WWTP-sizes. Keys needed: 'path' and 'wwtp'.
        The latter shall incidate the name of the sheet.
        :return: A translation-dict containing the municipality BFSNR as a key and the WWTP-size as a value
        """

        # Open the excel-file
        ws = xlrd.open_workbook(excel['path']).sheet_by_name(excel['wwtp'])

        # Determine the start for reading the information:
        ara_col = ws.col_values(0)
        startread = ara_col.index('py_start_read') + 2

        # Core: read the information and construct the dict:
        wwtp_dict = {}
        for bfsnr, ara in zip(ws.col_values(5)[startread:], ara_col[startread:]):
            bfsnr = int(bfsnr)
            ara = int(ara)

            # There are municipalities with more than one WWTP. Therefore, we only consider the largest size (this makes
            # sense in view of the fact that the "probability" is larger for a certain HH to be connected to the larger
            # plant). Please note that large WWTPs correspond to a small ARA class-value.
            if bfsnr in wwtp_dict.keys():
                if wwtp_dict[bfsnr] > ara:
                    wwtp_dict[bfsnr] = ara
                else:
                    continue
            else:
                wwtp_dict[bfsnr] = ara

        return wwtp_dict


    def _get_cantonal_fleets(self):
        """
        This function constructs a dict which translates municipal numbers to the respective cantonal fleet. It returns
        separate dicts for the diesel cars and the petrol cars.
        :return: two translation-dicts containing the municipality BFSNR as a key and the cantonal diesel car fleet as
        well as the cantonal petrol car fleet as values.
        """

        # In a first step we get the complete cantonal fleets for all municipalities:
        conn = get_pg_connection()
        cur = conn.cursor(cursor_factory=pge.RealDictCursor)
        cur.execute("""
                SELECT * FROM working_tables.boundaries_gde geo LEFT JOIN working_tables.car_fleets_kt cars
                ON geo.kantonsnum = cars.kantonsnum WHERE geo.bfs_nummer IN (SELECT DISTINCT bfs_nummer FROM
                working_tables.hh)
                """)
        allfleets = cur.fetchall()
        cur.close()
        conn.close()

        # We then split the fleets into petrol and diesel fleets:
        petrolfleets = self._get_cantonal_fuelfleets(allfleets, 'petrol')
        dieselfleets = self._get_cantonal_fuelfleets(allfleets, 'diesel')

        return petrolfleets, dieselfleets


    def _get_cantonal_fuelfleets(self, allfleets, fuel):
        """
        Helper function which recomputes the cantonal fleet specifically for petrol or diesel
        :param allfleets: list of dicts with all cantonal fleets for all municipalities (obtained by
        _get_cantonal_fleets())
        :param fuel: fuel type for which the fleets need to be recomputed (usually just 'diesel' or 'petrol')
        :return: dict which translates the municipality number to the specific cantonal car fleet
        """

        # First: extract all the relevant information for the fuel-type-carfleet:
        fuelfleet = {x['bfs_nummer']: {ky: x[ky] for ky in x.keys() if fuel in ky} for x in allfleets}

        # Then go through all muncipalities and recompute the car-fleet shares:
        for bfsnr in fuelfleet.keys():

            # Get the total sum of the shares
            totshare = 0
            for ky in fuelfleet[bfsnr].keys():
                totshare += fuelfleet[bfsnr][ky]

            # Compute the new share:
            for ky in fuelfleet[bfsnr].keys():
                fuelfleet[bfsnr][ky] /= totshare

        return fuelfleet


    def _read_fu_from_excel(self, excel, ei_lca=None):
        """
        This function reads the functional unit of the process models which can be found in the excel specified
        (usually Consumption_Cockpit.xlsx in sheet LCA-Modelling).
        :param excel: dict with information in which excelfile the LCA-Modelling can be found (see init-function)
        :param ei_lca: (optional) brightway2-LCA-class (see init-function) -> passing this parameter means also that
        the specific multiplication factors for WWTP, heating and car fleets for each municipality shall be computed (or in other
        words: passing this LCA-class means that the ConsumptionLCA-Cockpit shall prepare the factors for STATPOP-
        households)
        """

        # Open the worksheet with the LCA-Modelling
        ws = xlrd.open_workbook(excel['path']).sheet_by_name(excel['sheet'])

        # Read the header row and determine the column which indicates if a certain attribute shall be included in LCA
        # or not:
        header = ws.row_values(0)
        role = ws.col_values(header.index('Role'))

        # Cut the header row to the relevant information
        heads = header[header.index('On 1'):len(header)]

        # Prepare a dict which will contain the demand-to-functional-unit conversion factors of the process models as
        # well as a list of attributes which shows for which attributes LCA-Models exist.
        self.attributes_dict_ = {}
        self.attributes_list_ = []

        # Go through all consumption categories
        for i, conscat in enumerate(role):

            # if the role-value is 0, this means an LCA shall be performed for the consumption category under consideration
            if conscat == 0:

                # Test if an amount-category is available for the consumption category
                if ws.cell_value(i, 1) != '':
                    attr = ws.cell_value(i, 1).lower()
                else:
                    attr = ws.cell_value(i, 0).lower()

                # Fill in the attributes-list as well as the dict containing the conversion factors.
                self.attributes_dict_.update({attr: ws.cell_value(i, 3)})
                self.attributes_list_.append(attr)

                nm = attr + "_fu_"

                # Cut out the row for the modelled activity and trim it to the relevant information
                acts = ws.row_values(i)[header.index('On 1'):len(header)]

                # Core of the function: read the different unit processes, multiply it already with the CFL-factor which
                # comes from Scherer & Pfister to convert the functional unit of the unit process to the product under
                # consideration --> store the functional unit in a dict and then as an attribute of the function
                fu_dict = {}
                for j, head in enumerate(heads):
                    # jump to the 'On'-cell --> if it is =1 then the subsequent cells shall be included in the LCA
                    if 'On' in head:
                        if acts[j] == 1:
                            # Check if there is a demand:
                            if acts[j + 3] * acts[j + 4]:

                                # Construct the correct activity name:
                                actkey = acts[j + 1].replace('(', '').replace(')', '').replace("'", '').split(',')
                                actkey = (actkey[0], ','.join(actkey[1:])[1:])

                                # There is a possibility that the same unit process is used several times within the
                                # same process model:
                                if actkey not in fu_dict.keys():
                                    fu_dict.update({actkey: acts[j + 3] * acts[j + 4]})
                                else:
                                    new_val = fu_dict[actkey] + (acts[j + 3] * acts[j + 4])
                                    fu_dict.update({actkey: new_val})

                setattr(self, nm, fu_dict)

                # In case we want to derive municipality-specific scores for WWTP and car fleets (this is the case if
                # the class shall be applied for STATPOP-households, but not for consumption archetypes), then we enter
                # the following if-loop.
                if ei_lca:
                    if attr == 'mx571203':
                        self.wwtp_lca_dict_ = self._get_wwtp_lca_dict(heads, acts, ei_lca)
                    if attr == 'mx571302':
                        self.heating_lca_dict_ = self._get_heating_lca_dict(excel, ei_lca)
                    # Not implemented yet since not really necessary because we merge consumption model and MATSim-based
                    # mobility model directly
                    # if attr == 'm621501':
                    #     self.petrolcar_lca_dict_ = self._get_petrolcar_lca_dict()
                    # if attr == 'm621501':
                    #     self.dieselcar_lca_dict_ = self._get_dieselcar_lca_dict()


    def _get_wwtp_lca_dict(self, heads, acts, lca):
        """
        This function computes LCA scores for the different WWTP-sizes. This information can be used in combination with
        the translation dict obtained in _get_wwtp_bfsnr_dict to compute municipality-specific LCA-WWTP-scores. The
        function is called within  _read_fu_from_excel().
        :param heads: The list of headers in the excel file relevant for the LCA-modelling
        :param acts: The list of excel-cells which contains the WWTP-information
        :param lca: Brightway2-LCA-class (see init function)
        :return: Dict with WWTP-class size as keys and LCA-score as value.
        """

        wwtp_lca_dict = {}
        i = 0
        bw2.projects.set_current('heia33')

        # Go through all headers, jump to 'On'-header:
        for j, head in enumerate(heads):
            if 'On' in head:
                # We now want to do LCA for all the WWTPs. Therefore, we also need activities for which the 'On'-header
                # is 0 (but of course we want to exclude empty cells)
                if acts[j] == 1 or acts[j] == 0:
                    # i determines the class of the WWTP --> IMPORTANT: the order in the acts-list needs to be correct
                    i += 1

                    # Construct the correct key for the activity
                    actkey = acts[j + 1].replace('(', '').replace(')', '').replace("'", '').split(',')
                    actkey = (actkey[0], ','.join(actkey[1:])[1:])

                    # Go through all methods an do the LCA for the WWTPs
                    for methodkey in self.methods_.keys():
                        lca.switch_method(self.methods_[methodkey])
                        lca.redo_lcia({actkey: 1})
                        wwtp_lca_dict.update({"{}_{}".format(i, methodkey): lca.score})

        return wwtp_lca_dict


    def _get_heating_lca_dict(self, excel, lca):
        """
        This function constructs a dict to translate GWS-code which indicates the concrete energy carrier for heating
        to concrete heating activities. This allows for STATPOP-HHs to use the correct activities for the specific case of
        an individual household. The function is called within  _read_fu_from_excel().
        :param excel: Dict with information where to find information on the GWS-energy carriers. Keys needed: 'path' and 'heating'.
        The latter shall incidate the name of the sheet.
        :return: A translation-dict containing the LCA-FU-score for the different GWS-energy carriers
        """

        bw2.projects.set_current('heia33')

        # Open the excel-file
        ws = xlrd.open_workbook(excel['path']).sheet_by_name(excel['heating'])

        # Determine the start for reading the information:
        gws_col = ws.col_values(2)
        header = ws.row_values(0)

        # Cut the header row to the relevant information
        heads = header[header.index('On 1'):len(header)]

        # Core: read the information and construct the dict:
        heating_dict = {}

        # Go through all GWS-codes
        for i, gws in enumerate(gws_col[1:]):
            # The following if-statement shall break the for-loop in case we reach the last GWS-code-entry
            if gws == '':
                break

            # Read the activities-row
            acts = ws.row_values(i + 1)[header.index('On 1'):len(header)]

            fu_dict = {}
            for j, head in enumerate(heads):
                # jump to the 'On'-cell --> if it is =1 then the subsequent cells shall be included in the LCA
                if 'On' in head:
                    if acts[j] == 1:
                        # Check if there is a demand:
                        if acts[j + 3] * acts[j + 4]:

                            # Construct the correct activity name:
                            actkey = acts[j + 1].replace('(', '').replace(')', '').replace("'", '').split(',')
                            actkey = (actkey[0], ','.join(actkey[1:])[1:])

                            # There is a possibility that the same unit process is used several times within the
                            # same process model:
                            if actkey not in fu_dict.keys():
                                fu_dict.update({actkey: acts[j + 3] * acts[j + 4]})
                            else:
                                new_val = fu_dict[actkey] + (acts[j + 3] * acts[j + 4])
                                fu_dict.update({actkey: new_val})

            # We then directly do LCA for the FU under consideration and for all methods
            for methodkey in self.methods_.keys():
                lca.switch_method(self.methods_[methodkey])
                lca.redo_lcia(fu_dict)
                heating_dict.update({"{}_{}".format(int(gws), methodkey): lca.score * ws.cell_value(i + 1, 3)})

        return heating_dict


    def _get_petrolcar_lca_dict(self):
        raise NotImplementedError


    def _get_dieselcar_lca_dict(self):
        raise NotImplementedError


    def _do_lca(self, ei_lca, ex_lca):
        """
        This is the core-function of the class and performs the actual LCA of the functional units derived in
        _read_fu_from_excel().
        :param ei_lca: Brightway2-LCA-class for the Brightway-project containing ecoinvent-based databases
        :param ex_lca: Brightway2-LCA-class for the Brightway-project containing EXIOBASE activities
        :return: dict with all the LCA-scores for the functional units.
        """

        # Prepare a results-dict-container for the LCA-scores
        lca_fu_scores = {}

        # Go through all processes:
        funits = [attr for attr in dir(self) if attr.endswith('_fu_')]
        for fu in funits:

            # In the following we test if the consumption category was modelled by EXIOBASE or not --> depending on
            # this test we set the brightway2-project accordingly and take the correct brightway2-LCA-class
            ex_test_list = [x[0] for x in getattr(self, fu).keys()]
            if any('EXIOBASE' in ky for ky in ex_test_list):
                bw2.projects.set_current('exiobase_industry_workaround')
                lca = ex_lca
                # For documentation purposes we also list the EXIOBASE-functional units
                self.exiobase_fus_.append(fu)
            else:
                bw2.projects.set_current('heia33')
                lca = ei_lca

            # For a certain process: go through all methods which shall be assessed and do the lca-computations
            for methodkey in self.methods_.keys():
                lca.switch_method(self.methods_[methodkey])
                lca.redo_lcia(getattr(self, fu))  # with gettattr: get the attribute value of "fu" dynamically

                # in the following, a keyname is constructed and the lca-score is saved to a dict
                name = fu.split('_')
                name = '_'.join(name[0:-1])
                lca_fu_scores.update({'_'.join([name, methodkey]): lca.score})

        return lca_fu_scores


class Consumption_Archetype_LCA(object):
    """
    This class performs a LCA for a Consumption-Archetype. For this, it multiplies the
    "characterization factors" which were created by ConsumptionLCA with the concrete demands of a Consumption-
    Archetype.
    """

    def __init__(self, lca_fu_scores, methods, arche_dem, attributes_dict=None, **kwargs):
        """
        Init-function for Consumption_Archetype_LCA.
        :param lca_fu_scores: dict of LCA-scores of the functional units (in principle the mutliplication factors/
        "characterization" factors obtained from ConsumptionLCA).
        :param methods: dict of methods which shall be computed (similar to ConsumptionLCA-class)
        :param arche_dem: dict with demands of consumption archetypes
        :param attributes_dict: dict with funtional-unit-to-demand-conversion factors (e.g. convert 1 CHF to 1 Million
        € in case of EXIOBASE). This can also be obtained by ConsumptionLCA-class.
        """

        # Read in all the demands and the label-names of the archetype under consideration
        for ky in arche_dem.keys():
            if ky == 'cluster_label_name' or ky == 'cluster_label_def':
                setattr(self, ky + '_', arche_dem[ky])
            else:
                setattr(self, ky + '_dem_', arche_dem[ky])

        # The next try-statement tries to consider expenditures for secondary homes. However, this is only possible
        # if certain information is contained in the archetype-demands. If this information is not passed, then the
        # secondary homes expenditures are just not considered.
        try:
            self._sechomes = self._add_sechome_expenditures()
        except:
            pass

        # Core: go through all demands and compute a dict of LCA-scores
        self.lcascores_ = self._compute_arche_score(lca_fu_scores, methods, attributes_dict)

        # In case of enough information is available for the secondary homes, we will also perform an LCA for these
        # expenditures
        try:
            self._sechome_lca(lca_fu_scores, methods, attributes_dict)
        except:
            pass


    def _add_sechome_expenditures(self):
        """
        This function prepares the expenditures for secondary homes for being considered in the LCA as well.
        :return: dict of primary home demands which are proportionally increased according to the secondary home expenditures.
        """

        # In a first step, we sum the expenditures for secondary home (extra dwelling costs plus energy costs) and remove
        # them from further consideration
        sechome_nk = self.a572200_dem_ + self.a572300_dem_
        del self.a572200_dem_, self.a572300_dem_

        # Then we compute the total of the primary home expenditures
        primary_nk_names = ['a571202', 'a571203', 'a571204', 'a571205', 'a571301','a571302', 'a571303']
        tot_prim_nk = 0.
        for nk in primary_nk_names:
            tot_prim_nk += getattr(self, '{}_dem_'.format(nk))

        # We go through all the primary home expenditures, distribute the secondary home costs proportionally to the
        # primary home extra dwelling costs and compute by which factor these categories would increase
        sechome_dem = {}
        for nk in primary_nk_names:
            # Compute how much of the secondary home costs shall be distributed to a certain primary home category:
            addition = sechome_nk * getattr(self, '{}_dem_'.format(nk)) / tot_prim_nk
            # Compute the factor by which the expenditures for the primary home increases. The reason for computing this
            # factor is: we actually do not want to have the expenditures increased but rather the corresponding energy
            # or amount of waste bags, water, etc. demands.
            try:
                factor = (addition + getattr(self, '{}_dem_'.format(nk))) / getattr(self, '{}_dem_'.format(nk))
            except:
                # This try-statement just catches the case that a demand is 0
                factor = 0.

            # a571205 is the only primary home category for which we do not have amounts as a functional unit:
            if not 'a571205' == nk:
                # We now compute the new (increased) value of the corresponding amount of energy, waste bags, etc.
                nk_m = nk
                nk_m = nk_m.replace('a', 'mx')
                new_val = getattr(self, '{}_dem_'.format(nk_m)) * factor

                #...and add this increased demand to a dict.
                sechome_dem['{}_dem_'.format(nk_m)] = new_val

                # The expenditure can be deleted as it will not be further used
                delattr(self, '{}_dem_'.format(nk))

            else:
                new_val = getattr(self, '{}_dem_'.format(nk)) * factor
                sechome_dem['{}_dem_'.format(nk)] = new_val

        return sechome_dem


    def _sechome_lca(self, lca_fu_scores, methods, attributes_dict):
        """
        Similar to _compute_arche_score-function, but separately performed for the secondary home-issue.
        :param lca_fu_scores: see init-function and _compute_arche_score-function
        :param methods: see init-function and _compute_arche_score-function
        :param attributes_dict: see init-function and _compute_arche_score-function
        :return:
        """

        # Prepare a dict for the results of impacts of secondary homes
        sechome_scores = {}

        # go through the primary home demands which were proportionally increased by secondary home expenditures
        for ky in self._sechomes.keys():
            name = ky.split('_')

            # Go through all methods
            for methodkey in methods.keys():
                nm = name[0]

                # Get the corresponding FU-score-multiplier
                fu_score = lca_fu_scores[nm + '_fu_' + methodkey]

                # Get the demand-to-functional-unit-converter
                if attributes_dict:
                    dem2fu = attributes_dict[nm]
                else:
                    dem2fu = 1

                # Compute the score for the increased primary home demand
                new_lca = self._sechomes[ky]*fu_score*dem2fu

                # Then we compute the difference of the new (proportionally increased) primary home demand and the
                # existing primary home demand (without consideration fo secondary home expenditures) and store it to the
                # dict.
                if methodkey in sechome_scores.keys():
                    sechome_scores[methodkey] += (new_lca - self.lcascores_[nm+'_'+methodkey])
                else:
                    sechome_scores[methodkey] = (new_lca - self.lcascores_[nm + '_' + methodkey])

        # Finally, we update the lca-scores-dict with the LCA-score of secondary home expenditures
        for methodkey in methods.keys():
            self.lcascores_.update({'mx5723_'+methodkey: sechome_scores[methodkey]})


    def _compute_arche_score(self, lca_fu_scores, methods, attributes_dict):
        """
        This is the core function of the Consumption_Archetype_LCA-class and computes the LCA-scores of the demands for
        the archetype under consideration
        :param lca_fu_scores: dict of multiplication factors for the LCA (see init-function), obtained by ConsumptionLCA-class
        :param methods: dict of methods (see init-function)
        :param attributes_dict: dict with funtional-unit-to-demand-conversion factors (see init-function), obtained by ConsumptionLCA-class
        :return: dict with the LCA-scores for all demands of the archetype
        """

        # Prepare the dict for the scores of the archetype:
        lcascores = {'cluster_label_name': self.cluster_label_name_}
        lcascores['cluster_label_def'] = self.cluster_label_def_

        # Get all demand-attributes
        dems = [attr for attr in dir(self) if attr.endswith('_dem_')]

        # CORE: go through all demands and compute the LCA-score
        for dem in dems:
            name = dem.split('_')

            # Go through all methods. This is for a general use of the class. If it is called by do_archetype_lca(), then
            # only one method at once will be computed.
            for methodkey in methods.keys():
                nm = '_'.join(name[0:2]) if name[0] == 'cg' else name[0]

                # Extract the functional unit-multiplication-factor
                fu_score = lca_fu_scores[nm+'_fu_'+methodkey]

                # Extract the demand-to-functional-unit-conversion-factor (e.g. 1 CHF to 1 Million € in the case of EXIOBASE)
                if attributes_dict:
                    dem2fu = attributes_dict[nm]
                else:
                    dem2fu = 1

                # Do the actual multiplication and store the results in the results dict
                lcascores.update({nm+'_'+methodkey: getattr(self, dem)*fu_score*dem2fu})

        return lcascores


def do_archetype_lca(conslca, archetypes_demand, tablename):
    """
    This function performs LCA for the consumption archetypes and writes the results directly to the PG-Database.
    PLEASE NOTE that the results are stored as "impact per month per household".
    :param conslca: ConsumptionLCA-instance with the multiplication factors for the demands
    :param archetypes_demand: list of dicts with the archetypes-demands (can be obtained by a SQL-query
    :param tablename: string with the exact table name which shall be created in the PG-Database for the results
    (form: <schema>.<name of table>.
    :return: PLEASE NOTE: the results are directly written to the PG-Database
    """

    # Get connection to the PG_Database
    conn = get_pg_connection()
    writecur = conn.cursor()

    # Go through all methods for which LCA shall be performed:
    for methodkey in conslca.methods_.keys():

        # Create a new table to store the results of the environmental indicator under consideration
        # PLEASE NOTE: indexing and setting a primary key will be done manually afterwards
        tbl = tablename + '_' + methodkey
        query = """
        DROP TABLE IF EXISTS {}
        """.format(tbl)
        writecur.execute(query)
        conn.commit()

        # PLEASE NOTE that we also add the secondary-home-column with mx5723
        createlist = ['{} float'.format(a) for a in conslca.attributes_list_]
        query = """
        CREATE TABLE {} (cluster_label_name varchar, cluster_label_def int, {}, mx5723 float
        )""".format(tbl, ', '.join(createlist))
        writecur.execute(query)
        conn.commit()
        print("DON'T FORGET TO SET A PRIMARY KEY FOR {}".format(tbl))

        # CORE: go through all archetypes and compute LCA -> store results in a list of dicts
        archelcaexport = []
        for arche in archetypes_demand:
            consarchelca = Consumption_Archetype_LCA(conslca.lca_fu_scores_, {methodkey: conslca.methods_[methodkey]}, arche, attributes_dict=conslca.attributes_dict_)
            archescores = {'cluster_label_name': arche['cluster_label_name'], 'cluster_label_def': arche['cluster_label_def']}
            archescores.update({ky.replace('_{}'.format(methodkey),''): consarchelca.lcascores_[ky] for ky in consarchelca.lcascores_.keys()})
            archelcaexport.append(archescores)

        # Finally, write the LCA-scores of all archetypes to the newly created results table in PG-Database
        cols = ['cluster_label_name', 'cluster_label_def']
        cols += conslca.attributes_list_
        cols += ['mx5723']
        query = """
        INSERT INTO %s(%s)
        VALUES (%s(%s)s)
        """ % (tbl, ', '.join(cols), '%', ')s, %('.join(cols))
        writecur.executemany(query, archelcaexport)
        conn.commit()

    writecur.close()
    conn.close()


class RFClassifierTuner(object):
    """
    This class trains a Random Forest Classifier according to the passed tuning parameters and computes different
    classification metrics to evaluate afterwards. PLEASE NOTE: it separates a test set in the beginning (10% of dataset)
    which can be used later on for the calibration of probabilities
    """
    def __init__(self, rf_clf, tuning_params, X, y):
        """
        init-function of the RFClassifierTuner
        :param rf_cfl: SKLEARN-RandomForestClassifier-instance (base-estimator)
        :param tuning_params: dict of dicts. Please note that the "top-keys" will be used for naming the results. The
        lower keys should contain: 'criterion', 'n_estimators' and 'max_features'
        :param X: full dataset of predictors
        :param y: full dataset of class labels
        """

        self.tuning_params_ = tuning_params

        # Set the impurity criterion
        rf_clf.criterion = list(tuning_params.values())[0]['criterion']

        # Prepare containers for the scores and initialize the different combination of max_features/trees which shall
        # be investigated during tuning
        scores = {}
        test_scores = {}

        for ky in tuning_params.keys():
            # Set the max_features:
            rf_clf_ky = copy.deepcopy(rf_clf)
            rf_clf_ky.max_features = tuning_params[ky]['max_features']

            # Create separate RF-instances for each ntree-entry
            for ntree in tuning_params[ky]['n_estimators']:
                rf_clf_ky_ntree = copy.deepcopy(rf_clf_ky)
                rf_clf_ky_ntree.n_estimators = ntree

                # Store all tuning-combinations as attributes and prepare the scores-containers
                setattr(self, "{}_{}".format(ky, ntree), rf_clf_ky_ntree)
                scores["{}_{}".format(ky, ntree)] = {'kappa':[], 'acc':[], 'prec': [], 'rec':[], 'fscore': [], 'oob': []}
                test_scores["{}_{}".format(ky, ntree)] = {'kappa': np.nan, 'acc': np.nan, 'prec': np.nan, 'rec': np.nan, 'fscore': np.nan,
                                                     'oob': np.nan}

        # IMPORTANT: in a first step we separate 10% of the dataset for later testing --> we use a stratified approach
        # to prevent problems with class imbalance
        X_tune, X_test, y_tune, y_test = train_test_split(X, y, stratify=y, test_size=0.1, random_state=12)
        self.X_tune_, self.X_test_, self.y_tune_, self.y_test_ = X_tune.copy(), X_test.copy(), y_tune.copy(), y_test.copy()

        # CORE: we now enter the tuning-process with a stratified (to prevent problems with class imbalance) 10-fold
        # cross validation
        skf = StratifiedKFold(n_splits=10, shuffle=True, random_state=14)

        for train_index, eval_index in skf.split(X_tune, y_tune):
            X_train, X_eval = X_tune[train_index], X_tune[eval_index]
            y_train, y_eval = y_tune[train_index], y_tune[eval_index]

            # Now we go through all tuning-combinations, fit the RF-classifier and compute evaluation metrics
            for ky in scores.keys():
                # Retrieve the correct RF-instance and fit to the training set
                rf = getattr(self, ky)
                rf.fit(X_train, y_train)

                # predict the evaluation set:
                y_pred = rf.predict(X_eval)

                # Compute and store the different metrics
                scores[ky]['oob'].append(rf.oob_score_)
                scores[ky]['acc'].append(accuracy_score(y_eval, y_pred))
                prec, rec, fscore, _ = precision_recall_fscore_support(y_eval, y_pred, beta=1.0, average='weighted')
                scores[ky]['prec'].append(prec)
                scores[ky]['rec'].append(rec)
                scores[ky]['fscore'].append(fscore)
                scores[ky]['kappa'].append(cohen_kappa_score(y_eval, y_pred))

        # We then save the metrics to the class
        self.fold_scores_ = scores

        # In the following we compute the mean scores for the cross-validation and also convert these metrics to a dataframe
        scores_dict = copy.deepcopy(scores)
        for ky in scores_dict.keys():
            for m in scores_dict[ky].keys():
                scores_dict[ky][m] = np.mean(scores_dict[ky][m])
        self.scores_cv_ = scores_dict
        self.scores_cv_df_ = pd.DataFrame.from_dict(scores_dict, orient='index')

        # IMPORTANT: we now compute also all the classification metric statistics based on the separated test-set
        for ky in test_scores.keys():
            # Get the correct RF-instance and fit to the whole tuning set
            rf = getattr(self, ky)
            rf.fit(X_tune, y_tune)

            # Predict the test-set
            y_pred = rf.predict(X_test)

            # Compute and store the different metrics
            test_scores[ky]['oob'] = rf.oob_score_
            test_scores[ky]['acc'] = accuracy_score(y_test, y_pred)
            prec, rec, fscore, _ = precision_recall_fscore_support(y_test, y_pred, beta=1.0, average='weighted')
            test_scores[ky]['prec'] = prec
            test_scores[ky]['rec'] = rec
            test_scores[ky]['fscore'] = fscore
            test_scores[ky]['kappa'] = cohen_kappa_score(y_test, y_pred)

        self.scores_test_ = test_scores
        self.scores_test_df_ = pd.DataFrame.from_dict(test_scores, orient='index')


    def save2excel(self, savepath):
        """
        This function saves the cross-validation scores as well as the scores based on the test-set to an excel file
        :param savepath: string with path to location where excel file shall be stored
        :return: Stores excel file in dedicated location (savepath)
        """

        # First save cross-validation-scores to excel
        self.scores_cv_df_.to_excel(savepath, sheet_name='cross_validation', na_rep='-')

        # Then re-load the workbook in order to insert also scores-based on test-set
        workbook = load_workbook(savepath)
        writer = pd.ExcelWriter(savepath, engine='openpyxl')
        writer.book = workbook
        self.scores_test_df_.to_excel(writer, 'testset_validation', na_rep='-')  # The name needs to be restricted to 30 characters for excel
        writer.save()


    def print_classification_report(self):
        """
        This function allows for printing all the classification reports for all the tuned classifiers based on the test-set
        """
        for ky in self.scores_cv_.keys():
            print("CLASSIFICATION REPORT FOR {}".format(ky))
            rf = getattr(self, ky)
            print(classification_report(self.y_test_, rf.predict(self.X_test_), digits=2))


class HH_STATPOP_stats(object):
    """
    This class computes for each STATPOP-HH the statistics which are needed for the classifier and respectively to compute
    the consumption demands. This function is used by export_statpop_hhs-function
    """
    def __init__(self, pers):
        """
        init-function of the HH_STATPOP_stats-class
        :param pers: dict with the relevant statistics for each person and HH respectively
        """

        self.householdid_ = pers['householdid']
        self.egid_ = pers['egid']

        # IMPORTANT: the car demand corresponds to the mobility model which is based on MATSim (see mobility.py and
        # mobility_cockpit.py
        #self.car_demand =  pers['car_demand']
        self.mobility = pers['mobility']
        self.housing = pers['housing']

        # Compute the geographical statistics (PLEASE NOTE: we decided only to use "Grossregion" and not "Sprachregion"
        # or Kanton). This computation only needs to be done once for a certain HH
        self._hh_wide_geogr(**pers)

        # For the persons-related statistics we first need to initialize the stats and then append the persons-statistics
        self._ini_stats_attr()
        self.append_stats(**pers)


    def _hh_wide_geogr(self, grossregionench, **kwargs):
        """
        This function computes the household-wide (so needs to be done only once per HH) geographic statistics. PLEASE
        NOTE that we decided to only consider "Grossregion" and neglect "Sprachregion" as well as "Kanton" --> if these
        shall be computed nevertheless, the function needs two more parameters: "sprachgebiete" and "kantonsnum".
        :param grossregionench: ARE-number indicating which "Grossregion" the HH belongs to.
        """

        # We decided not to further pursue the localization via sprachgebiete and kantonsnum since preliminary computations
        # showed that this could give certain archetypes a very large weight
        # self.char_langregion_dch = 1 if sprachgebiete == 1 or sprachgebiete == 4 else 0
        # self.char_langregion_fch = 1 if sprachgebiete == 2 else 0
        # self.char_langregion_ich = 1 if sprachgebiete == 3 else 0
        #
        # self.char_kanton_zh = 1 if kantonsnum == 1 else 0
        # self.char_kanton_be = 1 if kantonsnum == 2 else 0
        # self.char_kanton_lu = 1 if kantonsnum == 3 else 0
        # self.char_kanton_sg = 1 if kantonsnum == 17 else 0
        # self.char_kanton_ag = 1 if kantonsnum == 19 else 0
        # self.char_kanton_ti = 1 if kantonsnum == 21 else 0
        # self.char_kanton_vd = 1 if kantonsnum == 22 else 0
        # self.char_kanton_ge = 1 if kantonsnum == 25 else 0
        # self.char_kanton_rest = 1 if kantonsnum not in (1, 2, 3, 17, 19, 21, 22, 25) else 0

        self.char_georegion_ge = 1 if grossregionench == 1 else 0
        self.char_georegion_mit = 1 if grossregionench == 2 else 0
        self.char_georegion_nw = 1 if grossregionench == 3 else 0
        self.char_georegion_zh = 1 if grossregionench == 4 else 0
        self.char_georegion_ost = 1 if grossregionench == 5 else 0
        self.char_georegion_zen = 1 if grossregionench == 6 else 0
        self.char_georegion_ti = 1 if grossregionench == 7 else 0


    def _ini_stats_attr(self):
        """
        This function initializes the statistics which are based on the individual persons belonging to a HH.
        """

        # the following for-loop initializes the gender/age-stats
        for sex in ['fem', 'male']:
            for age in ['0004', '0514', '1524', '2534', '3544', '4554', '5564', '6574', '7599']:
                attrname = 'char_no' + sex + age
                setattr(self, attrname, 0)

        # the following for-loop initializes marital-status-stats
        for i, marit in enumerate(['unwed', 'married', 'wid', 'div']):
            attrname = 'char_no' + marit
            setattr(self, attrname, 0)

        # Finally, we also initialize the Swiss/Non-Swiss-stats
        switchnat = {}
        for i, nat in enumerate(['ch', 'ausl']):
            switchnat.update({i + 1: nat})
            attrname = 'char_no' + nat
            setattr(self, attrname, 0)

        # The switchnat_dict will help to choose the correct attribute for the stats
        self._switchnat_dict_ = switchnat

        self.char_nopers = 0


    def append_stats(self, age, maritalstatus, sex, nationalitycategory, **kwargs):
        """
        This function updates the household-wide statistics with data of a new household member under consideration.
        :param age: age of the person
        :param maritalstatus: marital status of the person
        :param sex: gender of the person
        :param nationalitycategory: indicating if Swiss or non-Swiss
        """

        # Increase the total persons' count
        self.char_nopers += 1

        # Check the age/gender-stats and increase the relevant category by one
        s = 'fem' if sex == 2 else 'male'
        if age <= 4:
            a = '0004'
        elif 5 <= age <= 14:
            a = '0514'
        elif 15 <= age <= 24:
            a = '1524'
        elif 25 <= age <= 34:
            a = '2534'
        elif 35 <= age <= 44:
            a = '3544'
        elif 45 <= age <= 54:
            a = '4554'
        elif 55 <= age <= 64:
            a = '5564'
        elif 65 <= age <= 74:
            a = '6574'
        elif 75 <= age:
            a = '7599'
        attrname = 'char_no' + s + a
        setattr(self, attrname, getattr(self, attrname) + 1)

        # Check the marital status and increase the relevant category by one
        if maritalstatus == 1 or maritalstatus == 5:
            m = 'unwed'
        elif maritalstatus == 2 or maritalstatus == 6:
            m = 'married'
        elif maritalstatus == 3:
            m = 'wid'
        elif maritalstatus == 4 or maritalstatus == 7:
            m = 'div'
        try:
            attrname = 'char_no' + m
            setattr(self, attrname, getattr(self, attrname) + 1)
        except:
            pass

        # With the help of the switch-nat-dict we also classify the nationality status and increase the relevant category by one
        nat = self._switchnat_dict_.get(nationalitycategory)
        attrname = 'char_no' + nat
        setattr(self, attrname, getattr(self, attrname) + 1)


class HH_STATPOP_SKL(object):
    """
    This class converts the list of HH_STATPOP_stats-classed households of a municipality into a matrix that can be
    handled by sci-kit learn. It finally has a similar structure as the SKLDATA-classes (and sub-classes).
    This class is used by export_statpop_hhs.
    """
    def __init__(self, attributes, HHlist):
        """
        init-function of the HH_STATPOP_SKL-class
        :param attributes: list of strings with the attributes contained in the matrix
        :param HHlist: list of HH_STATPOP_stats-classes (constructed in export_statpop_hhs
        """

        # Initialize a similar structure as the SKLData-classes
        self.meta_ = []
        self.data_ = []
        self.attributes_ = attributes

        # Go through all households in the list
        for hh in HHlist:

            # Append household-ID
            self.meta_.append(hh.householdid_)

            # Convert the HH_STATPOP_stats-class to a dict:
            hh_dict = {attr: getattr(hh, attr) for attr in dir(hh) if
                       not attr.startswith('_') and not attr.startswith('append') and not attr == ''
                       and not attr.endswith('_')}

            # Go through all attributes (in order to ensure the correct order for the classifier) and append the
            # the attribute to the matrix
            hh_attrs = []
            for a in attributes:
                hh_attrs.append(hh_dict[a])
            self.data_.append(hh_attrs)

        # Construct np-arrays from the lists
        self.meta_ = np.array(self.meta_)
        self.data_ = np.array(self.data_)


def export_statpop_hhs(bfsnr, attributes, savepath):
    """
    This function retrieves and computes the necessary statistics from STATPOP and exports it as pickle directly to
    the LINUX-Server (or any other specified location). Unfortunately, the memory-restrictions on the PC do not allow
    for computing the demand.
    :param bfsnr: integer representing the BFSNR under investigation
    :param attributes: list of strings with attributes that shall be retrieved from the PG-Database
    :param savepath: path to store the pickles
    """

    # Connect to PG-Database and retrieve the relevant statistics; PLEASE NOTE that we decided only to do geographical
    # matching based on the "Grossregionen" and leave out Kanton and Sprachregion; We also retrieve the MATSim-based results
    # for the car driven kilometers per household
    conn = get_pg_connection()
    cur = conn.cursor(cursor_factory=pge.RealDictCursor)
    # query = """
    # SELECT pers.pers_id, pers.householdid, pers.sex, pers.age, pers.maritalstatus, pers.nationalitycategory, pers.egid,
    # are.grossregionench, mob.car_demand FROM working_tables.persons pers LEFT JOIN
    # original_data.are_raumgliederungen are ON pers.bfs_nummer=are.bfs_nummer LEFT JOIN
    # results.hh_mobility_demand mob ON mob.householdid=pers.householdid WHERE pers.bfs_nummer={}
    # """.format(bfsnr)
    # cur.execute(query)
    # statpop_persons = cur.fetchall()
    # cur.close()
    # del cur
    # conn.close()

    query = """
       SELECT pers.pers_id, pers.householdid, pers.sex, pers.age, pers.maritalstatus, pers.nationalitycategory, pers.egid,
       are.grossregionench, dem.mobility, dem.housing FROM working_tables.persons pers LEFT JOIN
       original_data.are_raumgliederungen are ON pers.bfs_nummer=are.bfs_nummer LEFT JOIN
       working_tables.statpop_housingmobility_quantiles dem ON dem.householdid=pers.householdid WHERE pers.bfs_nummer={}
       """.format(bfsnr)
    cur.execute(query)
    statpop_persons = cur.fetchall()
    cur.close()
    del cur
    conn.close()

    # For the municipality under investigation, we create a list of HH-IDs which were already computed and a list of
    # the household statistics. The first one is necessary since we go through the households person per person and
    # want therefore to know if a new person shall "initialize new statistics" or add to existing household statistics
    statpop_hhid_list = []
    statpop_hh_list = []
    for statpop_pers in statpop_persons:
        if not statpop_pers['householdid'] in statpop_hhid_list:
            statpop_hh_list.append(HH_STATPOP_stats(statpop_pers))
            statpop_hhid_list.append(statpop_pers['householdid'])
        else:
            stap_idx = statpop_hhid_list.index(statpop_pers['householdid'])
            statpop_hh_list[stap_idx].append_stats(**statpop_pers)

    # To also include the MATSim-based mobility demand, we need to add this to the attributes:
    attributesls = list(attributes)
    #attributesls.append('car_demand')
    attributesls.append('mobility')
    attributesls.append('housing')

    # Finally, we convert the list of households into classes which can be read by SKLEARN
    statpop_hh_skl = HH_STATPOP_SKL(attributesls, statpop_hh_list)

    # We then store all the relevant information in a dict:
    statpop_dict = {'statpop_hh_skl': statpop_hh_skl, 'statpop_hh_list': statpop_hh_list, 'statpop_hhid_list': statpop_hhid_list}

    # To use the data on the Linux-Server (decoupled from the PG-DB), we store it as a pickle
    with open(os.path.join(savepath, '{}.pickle'.format(bfsnr)), 'wb') as f:
        pickle.dump(statpop_dict, f)

    del statpop_dict, statpop_hh_skl, statpop_hh_list, statpop_hhid_list, attributes, statpop_persons


def estimate_statpop_pt_bike(HH, hharchetype, mc_data):
    """
    This function is a workaround to estimate public transport and bike demand from microcensus-data. It is therefore very
    similar to estimate_habe_pt_bike, but designed for STATPOP-households instead of HABE-households. PLEASE NOTE that this
    function is in principle not needed if the consumption and the MATSim-based mobility model are merged.
    :param HH: HH_STATPOP_stats-class instance for a specific household
    :param hharchetype: chosen archetype for the household (needed for the income)
    :param mc_data: microcensus-data retrieved by get_microcensus_data()-function
    :return: Estimate of driven kilometers per household and day by train, coach, tram, bike
    """

    # In a first step we compute the pt/bike demands by household size and directly multiply by the household
    # size to get an estimate "per household":
    size = HH.char_nopers if HH.char_nopers <= 5 else 5
    train_sizeestimate = mc_data['hhsize'][(size, 'train')] * HH.char_nopers
    coach_sizeestimate = mc_data['hhsize'][(size, 'coach')] * HH.char_nopers
    tram_sizeestimate = mc_data['hhsize'][(size, 'tram')] * HH.char_nopers
    bike_sizeestimate = mc_data['hhsize'][(size, 'bike')] * HH.char_nopers

    # The next estimate is based on income; again, we directly multiply the pkms with the household size
    inc = hharchetype['e_bruttoeink']
    if inc <= 2000:
        inc = 2000
    elif inc <= 6000:
        inc = 6000
    elif inc <= 10000:
        inc = 10000
    elif inc <= 14000:
        inc = 14000
    else:
        inc = 14001
    train_incestimate = mc_data['hhincome'][(inc, 'train')] * HH.char_nopers
    coach_incestimate = mc_data['hhincome'][(inc, 'coach')] * HH.char_nopers
    tram_incestimate = mc_data['hhincome'][(inc, 'tram')] * HH.char_nopers
    bike_incestimate = mc_data['hhincome'][(inc, 'bike')] * HH.char_nopers

    # Before we can determine bike/pt-demand by age and gender, we need to perform some statistics:
    females = 0
    males = 0
    age_24 = 0
    age_44 = 0
    age_64 = 0
    age_65 = 0
    for attr in dir(HH):
        if '0004' in attr:
            continue
        if 'fem' in attr:
            females += getattr(HH, attr)
        elif 'male' in attr:
            males += getattr(HH, attr)
        if '0514' in attr or '1524' in attr:
            age_24 += getattr(HH, attr)
        elif '2534' in attr or '3544' in attr:
            age_44 += getattr(HH, attr)
        elif '4554' in attr or '5564' in attr:
            age_64 += getattr(HH, attr)
        elif '6574' in attr or '7599' in attr:
            age_65 += getattr(HH, attr)

    # In order to estimate bike/pt based on age, we multiply the age-statistics with the daily distances in order
    # to directly get the estimate "per household"
    train_ageestimate = (mc_data['hhage'][(17, 'train')] + mc_data['hhage'][(24, 'train')]) / 2 * age_24 + \
                        mc_data['hhage'][(44, 'train')] * age_44 + mc_data['hhage'][(64, 'train')] * age_64 + \
                        mc_data['hhage'][(65, 'train')] * age_65
    coach_ageestimate = (mc_data['hhage'][(17, 'coach')] + mc_data['hhage'][(24, 'coach')]) / 2 * age_24 + \
                        mc_data['hhage'][(44, 'coach')] * age_44 + mc_data['hhage'][(64, 'coach')] * age_64 + \
                        mc_data['hhage'][(65, 'coach')] * age_65
    tram_ageestimate = (mc_data['hhage'][(17, 'tram')] + mc_data['hhage'][(24, 'tram')]) / 2 * age_24 + \
                       mc_data['hhage'][(44, 'tram')] * age_44 + mc_data['hhage'][(64, 'tram')] * age_64 + \
                       mc_data['hhage'][(65, 'tram')] * age_65
    bike_ageestimate = (mc_data['hhage'][(17, 'bike')] + mc_data['hhage'][(24, 'bike')]) / 2 * age_24 + \
                       mc_data['hhage'][(44, 'bike')] * age_44 + mc_data['hhage'][(64, 'bike')] * age_64 + \
                       mc_data['hhage'][(65, 'bike')] * age_65

    # Similar to age, we also multiply the genders-statistics directly with the estimated daily distances
    train_sexestimate = mc_data['hhsex'][('male', 'train')] * males + mc_data['hhsex'][('fem', 'train')] * females
    coach_sexestimate = mc_data['hhsex'][('male', 'coach')] * males + mc_data['hhsex'][('fem', 'coach')] * females
    tram_sexestimate = mc_data['hhsex'][('male', 'tram')] * males + mc_data['hhsex'][('fem', 'tram')] * females
    bike_sexestimate = mc_data['hhsex'][('male', 'bike')] * males + mc_data['hhsex'][('fem', 'bike')] * females

    # Finally, we take the average of all estimates (this is an ASSUMPTION!)
    train = (train_sizeestimate + train_incestimate + train_ageestimate + train_sexestimate) / 4
    coach = (coach_sizeestimate + coach_incestimate + coach_ageestimate + coach_sexestimate) / 4
    tram = (tram_sizeestimate + tram_incestimate + tram_ageestimate + tram_sexestimate) / 4
    bike = (bike_sizeestimate + bike_incestimate + bike_ageestimate + bike_sexestimate) / 4

    return train, coach, tram, bike


def do_consumption_demand(bfsnr, clfmob, clfhus, archedict, mc_data, habe_attributes, datapath, savepath):
    """
    This function computes the consumption demands for all households within a certain municpality and stores the
    results in csv-files
    :param bfsnr: integer with official muncipality-number
    :param clf: classifier-instance (should be calibrated for probabilities)
    :param archedict: dict with data for the archetypes (retrieved from PG-DB, still monthly data)
    :param mc_data: microcensus-data retrieved by get_microcensus_data --> please note that when coupling the consumption
    model with the MATSim-based mobility model, this is actually not necessary anymore. But for the sake of
    completeness, we still compute it.
    :param habe_attributes: list of attributes for the demands (to ensure the correct order)
    :param datapath: string with path where to find the preprocessed HH-data (pre-processing was done by export_statpop_hhs())
    :param savepath: string with path where to store the csv-files
    """

    # First we retrieve the preprocessed household-data (pre-processing was done by export_statpop_hhs())
    with open(os.path.join(datapath, '{}.pickle'.format(bfsnr)), 'rb') as f:
        statpop_dict = pickle.load(f)

    statpop_hh_skl = statpop_dict['statpop_hh_skl']
    statpop_hhid_list = statpop_dict['statpop_hhid_list']
    statpop_hh_list = statpop_dict['statpop_hh_list']

    # To reproduce the random choice of an archetype for a certain household, we seed the random-number by the muncipality
    # number
    np.random.seed(bfsnr)

    # Initialize the cluster choices as well as the demands for the households with list-containers
    hh_cluster_choices = []
    hh_dem = []

    # compute the cluster-member-probabilities for all households in a municipality
    # TODO document the following lines
    # TODO test if classes have same order in both classifiers
    assert np.array_equal(np.array(clfmob.classes_), np.array(clfhus.classes_))
    clf_classes = copy.deepcopy(clfmob.classes_)

    probs = np.zeros([statpop_hh_skl.data_.shape[0], len(clf_classes)])

    mobix = [True if pd.isnull(x) else False for x in statpop_hh_skl.data_[:, -1]]
    husix = [True if not pd.isnull(x) else False for x in statpop_hh_skl.data_[:,-1]]

    try:
        probs[mobix, :] = clfmob.predict_proba(statpop_hh_skl.data_[mobix,:-1])
        probs[husix, :] = clfhus.predict_proba(statpop_hh_skl.data_[husix,:])
    except:
        try:
            probs[mobix, :] = clfmob.predict_proba(statpop_hh_skl.data_[mobix, :-1])
        except:
            probs[husix, :] = clfhus.predict_proba(statpop_hh_skl.data_[husix, :])

    del clfmob, clfhus

    #old:
    # probs = clf.predict_proba(statpop_hh_skl.data_)
    # clf_classes = copy.deepcopy(clf.classes_)
    # del clf

    # Go through all households in the pre-processed STATPOP-data:
    for hhidx, hhstatpop in enumerate(statpop_hh_skl.data_):
        # Get a uniform random number between 0 and 1:
        randvar = np.random.uniform(low=0.0, high=1.0, size=None)

        # Compute the cumulated probabilities for a certain household by only considering probabilities > 0
        probs_cum = probs[hhidx][probs[hhidx] > 0].cumsum()

        # Construct a list of class-labels (again only for the HH-specific probabilities > 0)
        clusterlabels = clf_classes[probs[hhidx] > 0]

        # Go through the probabilities and check if the cumulated probability is below or equal to the random variable
        for i, prob in enumerate(probs_cum):
            if randvar <= prob:
                # if true, then we take the cluster as a choice for the consumption behavior
                chosen_cluster = clusterlabels[i]
                break

        # Extract the household-ID because we now compute also the PT-demand based on the microcensus data. PLEASE NOTE
        # that this is not really necessary if the consumption model and the MATSim-based Mobility Model are coupled.
        hhid = statpop_hh_skl.meta_[hhidx]
        idx = statpop_hhid_list.index(hhid)

        # Extract the archetype-demands for the household under consideration
        hharchetype = copy.deepcopy(archedict[chosen_cluster])

        # Computation of the public transport and bike demand based on microcensus data (see comment above)
        train, coach, tram, bike = estimate_statpop_pt_bike(statpop_hh_list[idx], hharchetype, mc_data)
        hharchetype['mx6213'] = bike * 365. / 12.
        hharchetype['mx622101'] = train * 365. / 12.
        hharchetype['mx622102'] = tram * 365. / 12.
        hharchetype['mx622201'] = coach * 365. / 12.

        # Finally we upscale the monthly demands to yearly demands:
        for ky in hharchetype.keys():
            if not ky.startswith('cg') and not ky.startswith('cluster_label'):
                hharchetype[ky] *= 12

        hharchetype['householdid'] = hhid

        # Construct a list with all the probabilities for all clusters for documentation (or post-analysis); these lists
        # will be stored in PG-DB
        hh_cluster_choice = [hhid, chosen_cluster] + list(probs[hhidx])
        hh_cluster_choices.append(hh_cluster_choice)

        # We finally add the demands (ensure correct order of attributes) to the list-container
        hh_dem.append([hharchetype[ky] for ky in habe_attributes if ky in hharchetype.keys()])

    # Finally we store the cluster-choices as well as the demands to CSV-files
    csvname = os.path.join(savepath, "hh_cluster_choices", "hh_cluster_choices_{}.csv".format(bfsnr))
    with open(csvname, 'w') as f:
        writer = csv.writer(f, lineterminator='\n')
        writer.writerows(hh_cluster_choices)

    csvname = os.path.join(savepath, "hh_consumption_demand","hh_consumption_demand_{}.csv".format(bfsnr))
    with open(csvname, 'w') as f:
        writer = csv.writer(f, lineterminator='\n')
        writer.writerows(hh_dem)


class Consumption_HH_LCA(object):
    """
    This class performs a LCA for STATPOP-households. For this, it multiplies the
    "characterization factors" which were created by ConsumptionLCA with the concrete demands of a Consumption-
    Archetype. It also considers secondary homes and is able to use municpality specific WWTP-scores and building-specific
    heating scores.
    """

    def __init__(self, lca_fu_scores, methods, hh_dem,  wwtp_score=None, heating_score=None, attributes_dict=None, **kwargs):
        """
        Init-function for Consumption_HH_LCA.
        :param lca_fu_scores: dict of LCA-scores of the functional units (in principle the mutliplication factors/
        "characterization" factors obtained from ConsumptionLCA).
        :param methods: dict of methods which shall be computed (similar to ConsumptionLCA-class)
        :param hh_dem: dict with demands of STATPOP-HH under consideration
        :param wwtp_score: municipality specific WWTP-score (already adapted to the method if passed with do_consumption_lca)
        :param heating_score: building specific heating-score (already adapted to the method if passed with do_consumption_lca)
        :param attributes_dict: dict with funtional-unit-to-demand-conversion factors (e.g. convert 1 CHF to 1 Million
        € in case of EXIOBASE). This can also be obtained by ConsumptionLCA-class.
        """

        # Read in all the demands and the householdid of the HH under consideration
        for ky in hh_dem.keys():
            if ky == 'householdid':
                self.householdid_ = hh_dem[ky]
            else:
                setattr(self, ky + '_dem_', hh_dem[ky])

        # The next try-statement tries to consider expenditures for secondary homes. However, this is only possible
        # if certain information is contained in the household demands. If this information is not passed, then the
        # secondary homes expenditures are just not considered.
        try:
            self._sechomes = self._add_sechome_expenditures()
        except:
            pass

        # Core: go through all demands and compute a dict of LCA-scores
        self.lcascores_ = self._compute_hh_score(lca_fu_scores, methods, wwtp_score, heating_score, attributes_dict)

        # In case of enough information is available for the secondary homes, we will also perform an LCA for these
        # expenditures
        try:
            self._sechome_lca(lca_fu_scores, methods, wwtp_score, heating_score, attributes_dict)
        except:
            pass


    def _add_sechome_expenditures(self):
        """
        This function prepares the expenditures for secondary homes for being considered in the LCA as well.
        :return: dict of primary home demands which are proportionally increased according to the secondary home expenditures.
        """

        # In a first step, we sum the expenditures for secondary home (extra dwelling costs plus energy costs) and remove
        # them from further consideration
        sechome_nk = self.a572200_dem_ + self.a572300_dem_
        del self.a572200_dem_, self.a572300_dem_

        # Then we compute the total of the primary home expenditures
        primary_nk_names = ['a571202', 'a571203', 'a571204', 'a571205', 'a571301','a571302', 'a571303']
        tot_prim_nk = 0.
        for nk in primary_nk_names:
            tot_prim_nk += getattr(self, '{}_dem_'.format(nk))

        # We go through all the primary home expenditures, distribute the secondary home costs proportionally to the
        # primary home extra dwelling costs and compute by which factor these categories would increase
        sechome_dem = {}
        for nk in primary_nk_names:
            # Compute how much of the secondary home costs shall be distributed to a certain primary home category:
            addition = sechome_nk * getattr(self, '{}_dem_'.format(nk)) / tot_prim_nk
            # Compute the factor by which the expenditures for the primary home increases. The reason for computing this
            # factor is: we actually do not want to have the expenditures increased but rather the corresponding energy
            # or amount of waste bags, water, etc. demands.
            try:
                factor = (addition + getattr(self, '{}_dem_'.format(nk))) / getattr(self, '{}_dem_'.format(nk))
            except:
                # This try-statement just catches the case that a demand is 0
                factor = 0.

            # a571205 is the only primary home category for which we do not have amounts as a functional unit:
            if not 'a571205' == nk:
                # We now compute the new (increased) value of the corresponding amount of energy, waste bags, etc.
                nk_m = nk
                nk_m = nk_m.replace('a', 'mx')
                new_val = getattr(self, '{}_dem_'.format(nk_m)) * factor

                #...and add this increased demand to a dict.
                sechome_dem['{}_dem_'.format(nk_m)] = new_val

                # The expenditure can be deleted as it will not be further used
                delattr(self, '{}_dem_'.format(nk))

            else:
                new_val = getattr(self, '{}_dem_'.format(nk)) * factor
                sechome_dem['{}_dem_'.format(nk)] = new_val

        return sechome_dem


    def _sechome_lca(self, lca_fu_scores, methods, wwtp_score, heating_score, attributes_dict):
        """
        Similar to _compute_hh_score-function, but separately performed for the secondary home-issue.
        :param lca_fu_scores: see init-function and _compute_hh_score-function
        :param methods: see init-function and _compute_hh_score-function
        :param attributes_dict: see init-function and _compute_hh_score-function
        """

        # Prepare a dict for the results of impacts of secondary homes
        sechome_scores = {}

        # go through the primary home demands which were proportionally increased by secondary home expenditures
        for ky in self._sechomes.keys():
            name = ky.split('_')

            # Go through all methods
            for methodkey in methods.keys():
                nm = name[0]

                # Get the corresponding FU-score-multiplier
                fu_score = lca_fu_scores[nm + '_fu_' + methodkey]

                # Get the demand-to-functional-unit-converter
                if attributes_dict:
                    dem2fu = attributes_dict[nm]
                else:
                    dem2fu = 1

                # Compute the score for the increased primary home demand
                new_lca = self._sechomes[ky]*fu_score*dem2fu

                # PLEASE NOTE the difference to Consumption_Archetype_LCA-class: here we compute the score of the primary
                # explicitely with the "Swiss heating mix" instead of the building-specific heating or wwtp score; otherwise
                # we could get troubles because the secondary home gets overestimated (e.g. when we assume the Swiss
                # mix for the secondary home and a GSHP for the primary home...)
                old_lca = getattr(self, "{}_dem_".format(nm))*fu_score*dem2fu

                # Then we compute the difference of the new (proportionally increased) primary home demand and the
                # existing primary home demand (without consideration fo secondary home expenditures) and store it to the
                # dict.
                if methodkey in sechome_scores.keys():
                    sechome_scores[methodkey] += (new_lca - old_lca)
                else:
                    sechome_scores[methodkey] = (new_lca - old_lca)

        # Finally, we update the lca-scores-dict with the LCA-score of secondary home expenditures
        for methodkey in methods.keys():
            self.lcascores_.update({'mx5723_'+methodkey: sechome_scores[methodkey]})


    def _compute_hh_score(self, lca_fu_scores, methods, wwtp_score, heating_score, attributes_dict):
        """
        This is the core function of the Consumption_HH_LCA-class and computes the LCA-scores of the demands for
        the STATPOP-household under consideration
        :param lca_fu_scores: dict of multiplication factors for the LCA (see init-function), obtained by ConsumptionLCA-class
        :param methods: dict of methods (see init-function)
        :param wwtp_score: municipality-specific WWTP-score (see init-function)
        :param heating_score: building-specific heating-score (see init-function)
        :param attributes_dict: dict with funtional-unit-to-demand-conversion factors (see init-function), obtained by ConsumptionLCA-class
        :return: dict with the LCA-scores for all demands of the archetype
        """

        # Prepare the dict for the scores of the household:
        lcascores = {'householdid': self.householdid_}

        # Get all demand-attributes
        dems = [attr for attr in dir(self) if attr.endswith('_dem_')]

        # CORE: go through all demands and compute the LCA-score
        for dem in dems:
            name = dem.split('_')

            # Go through all methods. This is for a general use of the class. If it is called by do_consumption_lca(), then
            # only one method at once will be computed.
            for methodkey in methods.keys():
                nm = '_'.join(name[0:2]) if name[0] == 'cg' else name[0]

                # Extract the functional unit-multiplication-factor
                fu_score = lca_fu_scores[nm+'_fu_'+methodkey]

                # If there are municipality-specific WWTP-scores and or building-specific heating scores passed, then
                # we overwrite the multiplication-factor
                if nm == "mx571203" and wwtp_score:
                    fu_score = wwtp_score
                if nm == "mx571302" and heating_score:
                    fu_score = heating_score

                # Extract the demand-to-functional-unit-conversion-factor (e.g. 1 CHF to 1 Million € in the case of EXIOBASE)
                # PLEASE NOTE: the dem2fu is already considered in the building-specific heating_score
                if attributes_dict and not (nm == 'mx571302' and heating_score):
                    dem2fu = attributes_dict[nm]
                else:
                    dem2fu = 1

                # Do the actual multiplication and store the results in the results dict
                lcascores.update({nm+'_'+methodkey: getattr(self, dem)*fu_score*dem2fu})

        return lcascores


def do_consumption_lca(bfsnr, conslca, savepath):
    """
    This function performs LCA for the consumption demand estimated by do_consumption_demand() for the STATPOP-HHs. It is
    similar to do_archetype_lca(), but it creates csv-files which will be imported in the Postgres-DB later on. It also
    operates only municipality-wise.
    PLEASE NOTE that the results are stored as "impact per YEAR per household".
    :param bfsnr: municipality-number
    :param conslca: ConsumptionLCA-instance with the multiplication factors for the demands
    :param savepath: string with path to location where the csv-files shall be stored
    """

    # Connect to the PG-DB and extract all household-IDs in the municipality
    conn = get_pg_connection()
    cur = conn.cursor()
    query = """
    SELECT householdid, egid FROM working_tables.hh WHERE bfs_nummer={}
    """.format(bfsnr)
    cur.execute(query)
    hhids_sql = cur.fetchall()
    hhids =[x[0] for x in hhids_sql]
    cur.close()
    egid_dict = {x[0]: x[1] for x in hhids_sql}

    # For the following attributes, no LCA-scores exist. However, if we want to do an LCA of expenditures for secondary
    # homes, then we need also to pass these attributes.
    attributes_for_secondaryhomes = ['a572200', 'a572300', 'a571202', 'a571203', 'a571204', 'a571205', 'a571301','a571302', 'a571303']

    # Extracting the household consumption demands by using a real-dict-cursor. Of course, this could have been done
    # directly with a JOIN-statement including the HH-IDs and skipping the HH-ID-extraction above.
    # However, this (more complicated) approach seemed to be faster.
    cur = conn.cursor(cursor_factory=pge.RealDictCursor)
    query = """
    SELECT householdid, {} FROM results.hh_consumption_demand WHERE householdid in ({})
    """.format(', '.join(conslca.attributes_list_ + attributes_for_secondaryhomes), ', '.join([str(hhid) for hhid in hhids]))
    cur.execute(query)
    hhdem = cur.fetchall()
    cur.close()

    # Extracting the household consumption demands by using a real-dict-cursor. Of course, this could have been done
    # directly with a JOIN-statement including the HH-IDs and skipping the HH-ID-extraction above.
    # However, this (more complicated) approach seemed to be faster.
    cur = conn.cursor()
    query = """
        SELECT egid, genhzs FROM original_data.gws_geb
        WHERE egid in ({})""".format(', '.join(list(set([str(egid_dict[ky]) for ky in egid_dict.keys()]))))
    cur.execute(query)
    gwsen = cur.fetchall()
    cur.close()
    conn.close()
    gws_dict = {x[0]: x[1] for x in gwsen}
    hh_gws_dict = {}
    for ky in egid_dict.keys():
        # The try-statement is necessary because not all HH have a proper EGID and not all EGIDs have an energy carrier
        try:
            hh_gws_dict[ky] = gws_dict[egid_dict[ky]]
        except:
            hh_gws_dict[ky] = 7200

    # The following list shall include ensure the correc order of the scores and also include secondary scores:
    attrlist = list(conslca.attributes_list_) + ['mx5723']

    # Go through all methods for which the score shall be computed
    for methodkey in conslca.methods_.keys():

        # Next we retrieve the specific score for the wastewater treatment of the municipality (we use the correct wwtp-class)
        wwtp_score = conslca.wwtp_lca_dict_["{}_{}".format(conslca.wwtp_bfsnr_dict_[bfsnr], methodkey)] if bfsnr in conslca.wwtp_bfsnr_dict_.keys() else None

        # Create the folder where the results shall be written to
        csvpath = savepath + methodkey
        if not os.path.exists(csvpath):
            os.makedirs(csvpath)

        # preparation of the list of lists which will be written to the CSV-file
        hhlcaexport = []

        # Go through all HH-dicts in the dict-results list:
        for hh in hhdem:

            # We then retrieve the specific LCA-score for specific circumstances of the households living situation (take
            # the correct heating energy carrier). PLEASE NOTE: the try-statement should not be necessary in principle.
            try:
                heating_score = conslca.heating_lca_dict_["{}_{}".format(hh_gws_dict[hh['householdid']], methodkey)]
            except:
                heating_score = None

            # Perform the actual LCA for the HH and method under consideration:
            conshhlca = Consumption_HH_LCA(conslca.lca_fu_scores_, {methodkey: conslca.methods_[methodkey]}, hh, wwtp_score, heating_score, attributes_dict=conslca.attributes_dict_)

            # retrieve the results:
            hh = conshhlca.lcascores_

            # First entry in the list: HH-ID
            hh_scores = [hh['householdid']]

            # Extract the results for the method under consideration and for all demands
            scores = [hh[ky+'_'+methodkey] for ky in attrlist]

            # Append the scores to the HH-ID
            hh_scores.extend(scores)

            # Append the HH under consideration to the list which will be written to the CSV-file
            hhlcaexport.append(hh_scores)

        # Create name of the CSV-file
        csvname = "{}\consumption_{}_{}.csv".format(csvpath, methodkey, bfsnr)

        # Write the results of the method under consideration to the CSV-files:
        with open(csvname, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerows(hhlcaexport)
            del writer

        del hhlcaexport, scores, hh_scores
        del conshhlca

    del conslca, bfsnr, savepath, hhids, hhdem, wwtp_score





#----------------------------------------------------------------------------------------------------------------------
# Deprecated version:
#----------------------------------------------------------------------------------------------------------------------


# class SKLData(object):
#     def __init__(self, conn, sqltable, attributes=None, meta=None, joining=None,  xlhabe=False, xlfilter=False, **kwargs):
#         if xlhabe:
#             wb = xlrd.open_workbook(r"D:\froemelt\Documents\Andi\02_Doktorat\03_Projects\05_HEIA\03_Computations\Consumption\Datascreening.xlsx")
#             self.ws = wb.sheet_by_name('Datascreening')
#             if sqltable.split(".")[1] == 'habe_mengen':
#                 self.codes_ = self.ws.col_values(6)
#             else:
#                 self.codes_ = self.ws.col_values(2)
#             self.varnames_ = {}
#             for i, var in enumerate(self.codes_):
#                 if var not in ['Variable-Code', '', '1: yes', '0: no', 'Amounts', 0.0]:
#                     self.varnames_.update({var: self.ws.cell(i, 1).value})
#             self.codes_ = self.codes_[5:]
#
#         else:
#             self.varnames_ = {}
#             for a in attributes:
#                 self.varnames_.update({a: a})
#
#         self.data_ = self._get_data(conn, sqltable, joining, attributes, xlhabe, xlfilter)
#
#         if meta:
#             self.meta_name_ = meta
#             self.meta_ = self._get_meta()
#
#     def _get_data(self, conn, sqltable, joining, attributes, xlhabe, xlfilter):
#         if xlhabe:
#             tabl = sqltable.split(".")
#             switchtabls = {
#                 'habe_ausgaben': [77, 608],
#                 'habe_konsumgueter': [608, 628],
#                 'habe_mengen': [77, 608]
#             }
#             ind = switchtabls.get(tabl[1])
#
#             self.codes_ = self.codes_[ind[0]:ind[1]]
#
#             if xlfilter:
#                 role = self.ws.col_values(4)
#                 del self.ws
#                 role = role[5:]
#                 role = role[ind[0]:ind[1]]
#                 indices = [i for i, x in enumerate(role) if x not in('', -1)]
#                 self.codes_ = [self.codes_[i] for i in indices]
#             if tabl[1] == 'habe_mengen':
#                 self.codes_ = [c for c in self.codes_ if c not in (0.0, '')]
#
#             self.attributes_ = self.codes_
#             del self.codes_
#
#         else:
#             self.attributes_ = attributes
#
#         if joining:
#             for attribs in joining[2]:
#                 self.attributes_ += attribs
#
#         query = """
#         SELECT {} FROM {}
#         """.format((", ".join(self.attributes_)).lower(), sqltable)
#
#         if joining:
#             for i, tbl in enumerate(joining[0]):
#                 query += " LEFT JOIN {} ON {}".format(tbl, "{}.{} = {}.{}".format(tbl, joining[1][i],
#                                                                                   sqltable, joining[1][i]))
#
#         cur = conn.cursor()
#         cur.execute(query)
#         sqldata = cur.fetchall()
#         cur.close()
#
#         return np.array(sqldata)
#
#
#     def _get_meta(self):
#         i = self.attributes_.index(self.meta_name_)
#         metavalues = np.copy(self.data_[:, i])
#         self.data_ = np.delete(self.data_, i, 1)
#         del self.attributes_[i]
#         return metavalues
#
#     def standardize_data(self):
#         self.data_scaled_ = preprocessing.scale(self.data_)